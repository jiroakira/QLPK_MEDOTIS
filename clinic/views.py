
from medicine.serializers import ThuocSerializer
import time
import pytz

from django.db.models.expressions import Func, OuterRef, Subquery
from clinic.excelUtils import ExportExcelBaoCaoXNTThuoc, ExportExcelDSThuocSapHetDate, WriteToExcel, writeToExcelDichVu, writeToExcelThuoc
from finance.models import (
    HoaDonChuoiKham, 
    HoaDonLamSang,
    HoaDonNhapHang, 
    HoaDonThuoc, 
    HoaDonTong, 
    HoaDonVatTu,
    NhapHang
)
from clinic.forms import (
    BaiDangForm,
    MauHoaDonForm, 
    MauPhieuForm, 
    PhongKhamForm, 
    ThuocForm, 
    UserForm, 
    CongTyForm, 
    DichVuKhamForm, 
    PhongChucNangForm, 
    BacSiForm, 
    VatTuForm
)
from medicine.models import (
    DonThuoc, 
    KeDonThuoc,
    NhomThuoc, 
    Thuoc, 
    ThuocLog, 
    TrangThaiDonThuoc, 
    VatTu, 
    NhomVatTu,
    KeVatTu,
    CongTy,
    NhomThau,
)
from django.http.response import  HttpResponseRedirect, JsonResponse
from django.http import HttpResponse
from rest_framework.response import Response
from django.db.models.functions import TruncDay
from django.db.models import Count, F, Sum, Q
from django.db import models
from clinic.models import (
    BacSi, 
    BaiDang, 
    ChiSoXetNghiem, 
    ChiTietChiSoXetNghiem, 
    ChuoiKham, 
    DanhMucBenh, 
    DanhMucChuongBenh, 
    DanhMucLoaiBenh, 
    DanhMucNhomBenh, 
    DichVuKham, 
    District, 
    DoTuoiXetNghiem, 
    DoiTuongXetNghiem, 
    FileKetQua, 
    FileKetQuaChuyenKhoa, 
    FileKetQuaTongQuat, 
    FilePhongKham, 
    HtmlKetQua, 
    KetQuaChuyenKhoa, 
    KetQuaTongQuat, 
    KetQuaXetNghiem, 
    LichHenKham, 
    LichSuChuoiKham, 
    LichSuTrangThaiKhoaKham, 
    MauPhieu, 
    NhomChiPhi,
    NhomChiSoXetNghiem, 
    PhanKhoaKham, 
    PhongChucNang, 
    PhongKham, 
    Province, 
    TrangThaiChuoiKham, 
    TrangThaiKhoaKham, 
    TrangThaiLichHen, 
    User, 
    Ward
)
from django.shortcuts import render
import json
from django.shortcuts import resolve_url
from django.shortcuts import get_object_or_404
from django.contrib.auth import authenticate, login as auth_login
from rest_framework.views import APIView
from django.utils import timezone
from django.contrib import messages
from datetime import datetime, timedelta, date
import decimal
from django.db.models import Max
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db import transaction
from django.contrib.auth import views as auth_views
from asgiref.sync import async_to_sync
from channels.layers import get_channel_layer
from clinic.serializers import FilterHoaDonChuoiKhamBaoHiemSerializer, GroupSerializer, HoaDonChuoiKhamSerializer, PermissionSerializer
from django.contrib.auth import logout
from django.urls import reverse
import locale 
from django.contrib.auth.models import Group, Permission
from django.contrib.contenttypes.models import ContentType

from dateutil.relativedelta import relativedelta

format = '%m/%d/%Y %H:%M %p'

format_2 = '%d/%m/%Y %H:%M'

format_3 = '%d/%m/%Y'

timezone.activate(pytz.timezone("Asia/Ho_Chi_Minh"))

local_time = timezone.localtime(timezone.now())

def getSubName(name):
    lstChar = []
    lstString = name.split(' ')
    for i in lstString:
        lstChar.append(i[0].upper())
    subName = "".join(lstChar)
    return subName

user_login_required = user_passes_test(lambda user: user.is_staff or user.has_perm('clinic.general_view'), login_url='/dang_nhap/')

def staff_user_required(view_func):
    decorated_view_func = login_required(user_login_required(view_func))
    return decorated_view_func

@login_required(login_url='/dang_nhap/')
def index(request):
    if request.user.is_superuser or request.user.is_admin or request.user.has_perm('clinic.general_view'):
        nguoi_dung = User.objects.filter(chuc_nang=1)

        phong_chuc_nang = PhongChucNang.objects.all()
        # * danh sách bệnh nhân chưa được khám
        trang_thai = TrangThaiLichHen.objects.get_or_create(ten_trang_thai="Đã Đặt Trước")[0]

        now = timezone.localtime(timezone.now())
        tomorrow = now + timedelta(1)

        lich_hen = LichHenKham.objects.filter(trang_thai = trang_thai).annotate(relevance=models.Case(
            models.When(thoi_gian_bat_dau__gte=now, then=1),
            models.When(thoi_gian_bat_dau__lt=now, then=2),
            output_field=models.IntegerField(),
        )).annotate(
        timediff=models.Case(
            models.When(thoi_gian_bat_dau__gte=now, then= F('thoi_gian_bat_dau') - now),
            models.When(thoi_gian_bat_dau__lt=now, then=now - F('thoi_gian_bat_dau')),
            # models.When(thoi_gian_bat_dau__lte=today_end - F('thoi_gian_bat_dau')),
            output_field=models.DurationField(),
        )).order_by('relevance', 'timediff').order_by('-thoi_gian_tao')
        
        upcoming_events = []
        past_events = []
        # today_events = LichHenKham.objects.filter(trang_thai = trang_thai, thoi_gian_bat_dau__lte=today_end)
        for lich in lich_hen:
            if lich.relevance == 1:
                upcoming_events.append(lich)
            elif lich.relevance == 2:
                past_events.append(lich)

        now = datetime.now()
        delta = timedelta(days=1)

        bai_dang = BaiDang.objects.filter(thoi_gian_ket_thuc__gt=now)
        starting_day = datetime.now() - timedelta(days=7)

        user_trong_ngay = User.objects.filter(thoi_gian_tao__gte=starting_day).order_by('-thoi_gian_tao')

        tong_tien_hoa_don_chuoi_kham = []
        tong_tien_hoa_don_thuoc = []
        tong_tien_hoa_don_lam_sang = []
        so_luong_nguoi_dung = []
        time_series = []

        while starting_day <= now:
            starting_day += delta
            time_start = starting_day.date()
            time_end = (starting_day + delta).date()
            time_str = starting_day.strftime("%d/%m/%Y")
            danh_sach_hoa_don = HoaDonChuoiKham.objects.filter(thoi_gian_tao__range=[time_start, time_end]).exclude(Q(tong_tien__isnull=True) | Q(tong_tien=0.000))
            danh_sach_hoa_don_thuoc = HoaDonThuoc.objects.filter(thoi_gian_tao__range=[time_start, time_end]).exclude(Q(tong_tien__isnull=True) | Q(tong_tien=0.000))
            danh_sach_hoa_don_lam_sang = HoaDonLamSang.objects.filter(thoi_gian_tao__range=[time_start, time_end]).exclude(Q(tong_tien__isnull=True) | Q(tong_tien=0.000))
            danh_sach_nguoi_dung = User.objects.filter(thoi_gian_tao__range=[time_start, time_end]).filter(chuc_nang = '1')
            total_count_user = User.get_count_in_day(danh_sach_nguoi_dung)
            total_value_chuoi_kham = HoaDonChuoiKham.analysis_total_value(danh_sach_hoa_don)
            total_value_don_thuoc = HoaDonThuoc.analysis_total_value(danh_sach_hoa_don_thuoc)
            total_value_lam_sang = HoaDonLamSang.analysis_total_value(danh_sach_hoa_don_lam_sang)
            tong_tien_hoa_don_chuoi_kham.append(str(total_value_chuoi_kham))
            tong_tien_hoa_don_thuoc.append(str(total_value_don_thuoc))
            tong_tien_hoa_don_lam_sang.append(str(total_value_lam_sang))
            so_luong_nguoi_dung.append(str(total_count_user))
            time_series.append(time_str)

        data = {
            'user': request.user,
            'upcoming_events': upcoming_events[:6],
            'past_events': past_events[:6],
            'nguoi_dung': nguoi_dung,
            'user_trong_ngay': user_trong_ngay[:10],
            'tong_tien_chuoi_kham' : tong_tien_hoa_don_chuoi_kham,
            'tong_tien_thuoc' : tong_tien_hoa_don_thuoc,
            'tong_tien_lam_sang': tong_tien_hoa_don_lam_sang,
            'so_luong_nguoi_dung': so_luong_nguoi_dung,
            'thoi_gian': time_series,
            'bai_dang' : bai_dang,
            'phong_chuc_nang' : phong_chuc_nang,
        }
        return render(request, 'index.html', context=data)
    else:
        phong_chuc_nang = PhongChucNang.objects.all()
        context = {
            'phong_chuc_nang': phong_chuc_nang
        }
        return render(request, 'no_permission_view.html', context)
    
@login_required(login_url='/dang_nhap/')
def danh_sach_benh_nhan(request):
    danh_sach_benh_nhan = User.objects.filter(chuc_nang = 1)
    trang_thai = TrangThaiLichHen.objects.all()
    phong_chuc_nang = PhongChucNang.objects.all()
    tinh = Province.objects.all()
    data = {
        'danh_sach_benh_nhan': danh_sach_benh_nhan,
        'trang_thai': trang_thai,
        'phong_chuc_nang' : phong_chuc_nang,
        'province': tinh
    }
    return render(request, 'le_tan/danh_sach_benh_nhan.html', context=data)

@login_required(login_url='/dang_nhap/')
def update_benh_nhan(request, **kwargs):
    id_benh_nhan = kwargs.get('id')
    instance = get_object_or_404(User, id=id_benh_nhan)
    form = UserForm(request.POST or None, instance=instance)
    phong_chuc_nang = PhongChucNang.objects.all()
    provinces = Province.objects.all()
    data = {
        'benh_nhan': instance,
        'form': form,
        'id_benh_nhan': id_benh_nhan,
        'phong_chuc_nang': phong_chuc_nang,
        'province': provinces
    }
    return render(request, 'le_tan/update_benh_nhan.html', context=data)

def xoa_benh_nhan(request, **kwargs):
    if request.method == 'POST':
        if request.user.is_authenticated and request.user.is_staff:
            id_benh_nhan = request.POST.get('id')
            user = get_object_or_404(User, id=id_benh_nhan)
            user.delete()
            response = {
                'status': 200,
                'message': "Xóa Bệnh Nhân Thành Công"
            }
        else:
            response = {
                'status': 400,
                'message': "Bạn Không Có Quyền Xóa Bệnh Nhân"
            }
            return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")
    else:
        response = {
            'status': 400,
            'message': "Không thể xóa bệnh nhân "
        }
    return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")

@login_required(login_url='/dang_nhap/')
def cap_nhat_thong_tin_benh_nhan(request):
    if request.method == "POST":
        id_benh_nhan   = request.POST.get('id_benh_nhan')
        ho_ten         = request.POST.get('ho_ten')
        so_dien_thoai  = request.POST.get('so_dien_thoai')
        email          = request.POST.get('email')
        cmnd_cccd      = request.POST.get('cmnd_cccd')
        ngay_sinh      = request.POST.get('ngay_sinh')
        gioi_tinh      = request.POST.get('gioi_tinh')
        dan_toc        = request.POST.get('dan_toc')
        ma_so_bao_hiem = request.POST.get('ma_so_bao_hiem')
        dia_chi        = request.POST.get('dia_chi')
        tinh_id = request.POST.get('tinh') 
        huyen_id = request.POST.get('huyen') 
        xa_id = request.POST.get('xa')

        ma_dkbd = request.POST.get('ma_dkbd')
        gt_the_tu = request.POST.get('gt_the_tu')
        lien_tuc_5_nam_tu = request.POST.get('lien_tuc_5_nam_tu')
        can_nang = request.POST.get('can_nang')
        if can_nang != '':
            can_nang = request.POST.get('can_nang')
        else:
            can_nang = 0
    
        ngay_sinh = datetime.strptime(ngay_sinh, format_3)
        ngay_sinh = ngay_sinh.strftime("%Y-%m-%d")
    
        benh_nhan = get_object_or_404(User, id=id_benh_nhan)
        benh_nhan.ho_ten         = ho_ten
        benh_nhan.so_dien_thoai  = so_dien_thoai
        benh_nhan.cmnd_cccd      = cmnd_cccd
        benh_nhan.dia_chi        = dia_chi
        benh_nhan.ngay_sinh      = ngay_sinh
        benh_nhan.gioi_tinh      = gioi_tinh
        benh_nhan.dan_toc        = dan_toc
        benh_nhan.ma_so_bao_hiem = ma_so_bao_hiem

        if email != "":
            benh_nhan.email = email
        else:
            benh_nhan.email = ''

        if tinh_id != "":
            tinh = Province.objects.filter(id=tinh_id).first()       
            benh_nhan.tinh = tinh
        else:
            response = {
                'status': 404,
                'message': 'Không thể thiếu thông tin Tỉnh/Thành Phố'
            }
            return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")

        if huyen_id != "":
            huyen = District.objects.filter(id=huyen_id).first()
            benh_nhan.huyen = huyen
        else:
            response = {
                'status': 404,
                'message': 'Không thể thiếu thông tin Quận/Huyện'
            }
            return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")

        if xa_id != "":
            xa = Ward.objects.filter(id=xa_id).first()     
            benh_nhan.xa = xa
        else:
            response = {
                'status': 404,
                'message': 'Không thể thiếu thông tin Phường/Xã'
            }
            return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")

        benh_nhan.can_nang = can_nang
        benh_nhan.ma_dkbd = ma_dkbd

        if gt_the_tu != '':
            gt_the_tu = datetime.strptime(gt_the_tu, format_3)
            gt_the_tu = gt_the_tu.strftime("%Y-%m-%d")
            benh_nhan.gt_the_tu = gt_the_tu

        if lien_tuc_5_nam_tu != '':
            lien_tuc_5_nam_tu = datetime.strptime(lien_tuc_5_nam_tu, format_3)
            lien_tuc_5_nam_tu = lien_tuc_5_nam_tu.strftime("%Y-%m-%d")
            benh_nhan.lien_tuc_5_nam_tu = lien_tuc_5_nam_tu

        benh_nhan.save()

        from actstream import action 
        action.send(request.user, verb="đã cập nhật thông tin người dùng", target=benh_nhan)

        response = {
            'status': 200,
            'message': 'Cập Nhật Thông Tin Thành Công'
        }
    else:
        response = {
            'status': 404,
            'message': 'Có lỗi xảy ra'
        }
    return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")

class LoginView(auth_views.LoginView):
    template_name = 'registration/login.html'

    def get_success_url(self):
        return resolve_url('trang_chu')

# * Chức năng Lễ Tân
# TODO đăng kí tài khoản cho bệnh nhân tại giao diện dashboard
# TODO xem lịch hẹn khám của bệnh nhân, sau đó xác nhận lại vs bác sĩ và cập nhật lại thời gian cho lịch hẹn khám

@login_required(login_url='/dang_nhap/')
def lich_hen_kham_list(request):
    """ Trả về danh sách lịch hẹn của người dùng mới đặt lịch khám """
    trang_thai = TrangThaiLichHen.objects.get(ten_trang_thai="Đã đặt trước")
    lich_kham = LichHenKham.objects.filter(trang_thai=trang_thai).order_by("-thoi_gian_tao")
    data = {
        'lich_kham': lich_kham
    }
    return render(request, 'index.html', context=data)

def cap_nhat_lich_kham(request, *args, **kwargs):
    # NOTE function này chỉ là dùng khi lễ tân bấm xác nhận, sẽ thay đổi trạng thái của lịch hẹn,
    # NOTE chưa bao gồm thay đổi thời gian của lịch hẹn 
    pk = kwargs.get('pk')
    trang_thai_xac_nhan = TrangThaiLichHen.objects.get_or_create(ten_trang_thai="Xác Nhận")[0]
    lich_hen = get_object_or_404(LichHenKham, pk=pk)
    lich_hen.trang_thai = trang_thai_xac_nhan
    lich_hen.nguoi_phu_trach = request.user
    lich_hen.save()
    return JsonResponse({
        'message': 'Cap Nhat Thanh Cong'
    })

class CapNhatLichKhamAPIToggle(APIView):
    # NOTE: sử dụng toggle này khi kết hợp với ajax, với mục đích là xác nhận lịch hẹn khám
    def get(self, request, format=None, *kwargs):
        pk = kwargs.get('pk')
        obj = get_object_or_404(LichHenKham, pk=pk)
        user_ = self.request.user
        trang_thai_xac_nhan = TrangThaiLichHen.objects.get_or_create(ten_trang_thai="Xác Nhận")[0]
        obj.trang_thai = trang_thai_xac_nhan
        obj.nguoi_phu_trach = user_
        obj.save()
        data = {"message": "Cap Nhat Thanh Cong"}
        return Response(data)
        
# tạo người dùng

# tạo người dùng

def create_user(request):
    if request.user.is_authenticated and request.user.is_staff:
        if request.method == "POST":
            ho_ten         = request.POST.get('ho_ten')
            so_dien_thoai  = request.POST.get('so_dien_thoai')
            password       = request.POST.get("password")
            email          = request.POST.get('email')
            cmnd_cccd      = request.POST.get('cmnd_cccd')
            ngay_sinh      = request.POST.get('ngay_sinh')
            gioi_tinh      = request.POST.get('gioi_tinh')
            dan_toc        = request.POST.get('dan_toc')
            ma_so_bao_hiem = request.POST.get('ma_so_bao_hiem')
            dia_chi        = request.POST.get('dia_chi')

            tinh_id = request.POST.get('tinh')
            huyen_id = request.POST.get('huyen') 
            xa_id = request.POST.get('xa')

            if email == '':
                email = ''

            ma_dkbd = request.POST.get('ma_dkbd')
            gt_the_tu = request.POST.get('gt_the_tu')
            lien_tuc_5_nam_tu = request.POST.get('lien_tuc_5_nam_tu')
            can_nang = request.POST.get('can_nang')

            if can_nang != '':
                can_nang = request.POST.get('can_nang')
            else:
                can_nang = 0

            ngay_sinh = datetime.strptime(ngay_sinh, format_3)
            ngay_sinh = ngay_sinh.strftime("%Y-%m-%d")
        
            if User.objects.filter(so_dien_thoai=so_dien_thoai).exists():
                response = {
                    'status': 404, 
                    'message': "Số Điện Thoại Này Đã Tồn Tại"
                }
                return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")

            benh_nhan = User.objects.create_nguoi_dung(
                ho_ten         = ho_ten.upper(), 
                so_dien_thoai  = so_dien_thoai, 
                password       = password,
                dia_chi        = dia_chi,
                ngay_sinh      = ngay_sinh,
                gioi_tinh      = gioi_tinh,
                dan_toc        = dan_toc,    
                ma_so_bao_hiem = ma_so_bao_hiem,
            )

            if cmnd_cccd != '':
                benh_nhan.cmnd_cccd = cmnd_cccd

            benh_nhan.ngay_sinh      = ngay_sinh
            benh_nhan.gioi_tinh      = gioi_tinh
            benh_nhan.dan_toc        = dan_toc
            benh_nhan.ma_so_bao_hiem = ma_so_bao_hiem

            benh_nhan.can_nang = can_nang
            benh_nhan.ma_dkbd = ma_dkbd

            if gt_the_tu != '':
                gt_the_tu = datetime.strptime(gt_the_tu, format_3)
                gt_the_tu = gt_the_tu.strftime("%Y-%m-%d")
                benh_nhan.gt_the_tu = gt_the_tu

            if lien_tuc_5_nam_tu != '':
                lien_tuc_5_nam_tu = datetime.strptime(lien_tuc_5_nam_tu, format_3)
                lien_tuc_5_nam_tu = lien_tuc_5_nam_tu.strftime("%Y-%m-%d")
                benh_nhan.lien_tuc_5_nam_tu = lien_tuc_5_nam_tu
            
            if tinh_id != "null":
                tinh = Province.objects.filter(id=tinh_id).first()       
                benh_nhan.tinh = tinh
            else:
                response = {
                    'status': 404, 
                    'message': "Không Thể Thiếu Tỉnh Thành"
                }
                return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")
            
            if huyen_id != "null":
                huyen = District.objects.filter(id=huyen_id).first()
                benh_nhan.huyen = huyen
            else:
                response = {
                    'status': 404, 
                    'message': "Không Thể Thiếu Quận Huyện"
                }
                return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")

            if xa_id != "null":
                xa = Ward.objects.filter(id=xa_id).first()
                benh_nhan.xa = xa
            else:
                response = {
                    'status': 404, 
                    'message': "Không Thể Thiếu Phường Xã"
                }
                return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")

            if dia_chi != '':
                benh_nhan.dia_chi        = dia_chi
            else:
                response = {
                    'status': 404, 
                    'message': "Không Thể Thiếu Địa Chỉ Cụ Thể"
                }
                return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")

            benh_nhan.save()

            from actstream import action
            action.send(request.user, verb="tạo mới người dùng", target=benh_nhan)
        
            response = {
                'status': 200,
                "message": "Đăng Kí Người Dùng Thành Công",
            }

            return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")
        else:
            response = {
                'status': 404,
                "message": "Đăng Kí Người Dùng Không Thành Công",
            }
            return HttpResponse(
                json.dumps(response),
                content_type="application/json"
            )
    else:
        response = {
            'status': 400,
            'message': "Bạn Không Có Quyền Thêm Người Dùng"
        }
        return HttpResponse(
            json.dumps(response),
            content_type="application/json", charset="utf-8"
        )

def add_lich_hen(request):      
    if request.method == "POST":
        id_benh_nhan      = request.POST.get('id_benh_nhan')
        thoi_gian_bat_dau = request.POST.get('thoi_gian_bat_dau', None)
        if thoi_gian_bat_dau == '':
            response = {
                'status': 400,
                'message': "Thời gian bắt đầu không được thiếu"
            }
            return HttpResponse(json.dumps(response), content_type='application/json; charset=utf-8')
        ly_do             = request.POST.get('ly_do')
        loai_dich_vu      = request.POST.get('loai_dich_vu')
        user              = User.objects.get(id=id_benh_nhan)

        thoi_gian_bat_dau = datetime.strptime(thoi_gian_bat_dau, format_2)
        thoi_gian = thoi_gian_bat_dau.strftime("%Y-%m-%d %H:%M")

        if loai_dich_vu == 'kham_chua_benh' or loai_dich_vu == 'kham_suc_khoe':
            trang_thai = TrangThaiLichHen.objects.get_or_create(ten_trang_thai = "Chờ Thanh Toán Lâm Sàng")[0]
            lich_hen = LichHenKham.objects.create(
                benh_nhan         = user,
                nguoi_phu_trach   = request.user,
                thoi_gian_bat_dau = thoi_gian,
                ly_do             = ly_do,
                loai_dich_vu      = loai_dich_vu,
                trang_thai        = trang_thai, 
            )
            lich_hen.save()

            response = {
                'message': "Bệnh nhân " + user.ho_ten
            }
        elif loai_dich_vu == 'kham_theo_yeu_cau':
            trang_thai = TrangThaiLichHen.objects.get_or_create(ten_trang_thai = "Chờ Phân Khoa")[0]
            lich_hen = LichHenKham.objects.create(
                benh_nhan         = user,
                nguoi_phu_trach   = request.user,
                thoi_gian_bat_dau = thoi_gian,
                ly_do             = ly_do,
                loai_dich_vu      = loai_dich_vu,
                trang_thai        = trang_thai,
                thanh_toan_sau = True,
            )
            lich_hen.save()
            response = {
                'message': "Bệnh nhân " + user.ho_ten
            }
        
    else: 
        response = {
            'message': "Có lỗi xảy ra"
        }
    return HttpResponse(json.dumps(response), content_type='application/json; charset=utf-8')

def create_thuoc(request):
    if request.method == "POST":
        id_cong_ty        = request.POST.get('id_cong_ty')
        ma_hoat_chat      = request.POST.get('ma_hoat_chat')
        ten_hoat_chat     = request.POST.get('ten_hoat_chat')
        ham_luong         = request.POST.get('ham_luong')
        duong_dung        = request.POST.get('duong_dung')
        ma_thuoc          = request.POST.get("ma_thuoc")
        ten_thuoc         = request.POST.get("ten_thuoc")
        so_dang_ky        = request.POST.get('so_dang_ky')
        dong_goi          = request.POST.get('dong_goi')
        don_vi_tinh       = request.POST.get('don_vi_tinh')
        don_gia           = request.POST.get("don_gia")
        don_gia_tt        = request.POST.get("don_gia_tt")
        gia_bhyt        = request.POST.get("gia_bhyt")
        so_lo             = request.POST.get('so_lo')
        hang_sx           = request.POST.get("hang_san_xuat")
        nuoc_sx           = request.POST.get("nuoc_san_xuat")
        quyet_dinh        = request.POST.get("quyet_dinh")
        cong_bo           = request.POST.get("cong_bo")
        loai_thuoc        = request.POST.get("loai_thuoc")
        han_su_dung       = request.POST.get("han_su_dung")
        ngay_san_xuat     = request.POST.get("ngay_san_xuat")

        han_su_dung = datetime.strptime(han_su_dung, format_3)
        han_su_dung = han_su_dung.strftime("%Y-%m-%d")

        ngay_san_xuat = datetime.strptime(ngay_san_xuat, format_3)
        ngay_san_xuat = ngay_san_xuat.strftime("%Y-%m-%d")        
    
        cong_ty = CongTy.objects.get(id=id_cong_ty)

        Thuoc.objects.create(
            cong_ty           = cong_ty,
            ma_hoat_chat      = ma_hoat_chat,
            ten_hoat_chat     = ten_hoat_chat,
            ham_luong         = ham_luong,
            duong_dung        = duong_dung,
            ma_thuoc          = ma_thuoc,
            ten_thuoc         = ten_thuoc,
            so_dang_ky        = so_dang_ky,
            dong_goi          = dong_goi,
            don_vi_tinh       = don_vi_tinh,
            don_gia           = don_gia,
            don_gia_tt        = don_gia_tt,
            so_lo             = so_lo,
            so_luong_kha_dung = 0,
            hang_sx           = hang_sx,
            nuoc_sx           = nuoc_sx,
            quyet_dinh        = quyet_dinh,
            cong_bo           = cong_bo,
            loai_thuoc        = loai_thuoc,
            han_su_dung       = han_su_dung,
            ngay_san_xuat     = ngay_san_xuat
        )

        response = {
            'status' : 200,
            'message' : 'Tạo Thành Công'
        }
    else:
        response = {
            'status': 404,
            'message': "Có Lỗi Xảy Ra"
        }    
    return HttpResponse(json.dumps(response), content_type='application/json; charset=utf-8')
        
    # return HttpResponse('upload')
# * function này không cần dùng tới template

def create_dich_vu(request):
    if request.method == "POST":
        ma_dvkt          = request.POST.get("ma_dvkt")
        ten_dvkt         = request.POST.get("ten_dvkt")
        don_gia          = request.POST.get("don_gia")
        ma_gia           = request.POST.get("ma_gia")
        quyet_dinh       = request.POST.get("quyet_dinh")
        cong_bo          = request.POST.get("cong_bo")
        phong_chuc_nang  = request.POST.get("phong_chuc_nang")

        phong_chuc_nang = PhongChucNang.objects.get(id=phong_chuc_nang)

        DichVuKham.objects.create(
            ma_dvkt          = ma_dvkt, 
            ten_dvkt         = ten_dvkt, 
            don_gia          = don_gia, 
            ma_gia           = ma_gia, 
            quyet_dinh       = quyet_dinh, 
            cong_bo          = cong_bo, 
            phong_chuc_nang  = phong_chuc_nang,  
        )

        response = {
            'status' : 200,
            'message' : 'Tạo Thành Công',
        }
    else:
        response = {
            'status' : 404,
            'message' : 'Có Lỗi Xảy Ra',
        }    
    return HttpResponse(json.dumps(response), content_type='application/json; charset=utf-8')


class BatDauChuoiKhamAPIToggle(APIView):
    """ 
    * Khi bác sĩ lâm sàng bấm bắt đầu khám cho một bệnh nhân nào đó, thì chuỗi khám sẽ tự động được tạo
    """
    def get(self, request, format=None, **kwargs):
        user_id = kwargs.get('id')
        user = get_object_or_404(User, id=user_id)
        bac_si_lam_sang = self.request.user
        now = timezone.localtime(timezone.now())
        trang_thai = TrangThaiChuoiKham.objects.get_or_create(trang_thai_chuoi_kham="Đang thực hiện")[0]
        chuoi_kham = ChuoiKham.objects.create(
            benh_nhan = user,
            bac_si_dam_nhan = bac_si_lam_sang,
            thoi_gian_bat_dau = now,
            trang_thai = trang_thai
        )
        chuoi_kham.save()
        messages.success(request, 'Bắt đầu chuỗi khám')
        data = {
            "message": "Chuỗi Khám Bắt Đầu"
        }
        return Response(data)

class KetThucChuoiKhamAPI(APIView):
    """ 
    * Khi bác sĩ lâm sàng bấm kết thúc chuỗi khám thì chuỗi khám của người dùng sẽ được cập nhật thêm thời gian kết thúc và sẽ chuyển trạng thái sang Hoàn thành
    """

    def get(self, request, format=None, **kwargs):
        chuoi_kham_id = kwargs.get('id', None)
        if chuoi_kham_id == None:
            messages.error(request, 'Không thể kết thúc chuỗi khám')
            messages.debug(request, f'Chuỗi khám không tồn tại (id:{chuoi_kham_id})')
            return Response({
                "message": "Khong The Ket Thuc Chuoi Kham Vi Chua Bat Dau Chuoi Kham"
            })
        else:
            trang_thai = TrangThaiChuoiKham.objects.get_or_create(trang_thai_chuoi_kham="Hoàn Thành")[0]
            chuoi_kham = get_object_or_404(ChuoiKham, pk = chuoi_kham_id)
            chuoi_kham.thoi_gian_ket_thuc = timezone.localtime(timezone.now())
            chuoi_kham.trang_thai = trang_thai
            chuoi_kham.save()
            messages.success(request, 'Kết thúc chuỗi khám')
        data = {
            "message": "Ket Thuc Chuoi Kham"
        }
        return Response(data)

    # * Đoạn này dùng để sắp xếp lịch hẹn theo trình từ upcoming-past
    # * upcoming events: ascendant order
    # * past events: descendant order
    #     LichHenKham.objects.annotate(relevance=models.Case(
    #     models.When(thoi_gian_bat_dau__gte=now, then=1),
    #     models.When(thoi_gian_bat_dau__lt=now, then=2),
    #     output_field=models.IntegerField(),
    # )).annotate(
    # timediff=models.Case(
    #     models.When(thoi_gian_bat_dau__gte=now, then=F('thoi_gian_bat_dau') - now),
    #     models.When(thoi_gian_bat_dau__lt=now, then=now - F('thoi_gian_bat_dau')),
    #     output_field=models.DurationField(),
    # )).order_by('relevance', 'timediff')

# TODO hiển thị danh sách bệnh nhân chờ khám đối với từng phòng chức năng
@login_required(login_url='/dang_nhap/')
def danh_sach_benh_nhan_cho(request):
    trang_thai = TrangThaiLichHen.objects.all()
    trang_thai_ck = TrangThaiChuoiKham.objects.all()
    phong_chuc_nang = PhongChucNang.objects.all()

    return render(request, 'bac_si_lam_sang/danh_sach_benh_nhan_cho.html', context={"trang_thai": trang_thai, "trang_thai_ck": trang_thai_ck, "phong_chuc_nang": phong_chuc_nang})

@login_required(login_url='/dang_nhap/')
def phong_chuyen_khoa(request, *args, **kwargs):
    # phong_chuc_nang = PhongChucNang.objects.values('ten_phong_chuc_nang').distinct()
    id_phong_chuc_nang = kwargs.get('id_phong_chuc_nang')
    phong = get_object_or_404(PhongChucNang, id=id_phong_chuc_nang)
    codename_perm = f'can_view_{phong.slug}'
    if request.user.has_perm(f'clinic.{codename_perm}'):
        phong_chuc_nang_detail = PhongChucNang.objects.get(id=id_phong_chuc_nang)
        phong_chuc_nang = PhongChucNang.objects.all()
        trang_thai = TrangThaiKhoaKham.objects.all()
        data = {
            'phong_chuc_nang': phong_chuc_nang,
            'trang_thai': trang_thai,
            'id_phong_chuc_nang' : id_phong_chuc_nang,
            'phong_chuc_nang_detail': phong_chuc_nang_detail
        }
        return render(request, 'bac_si_chuyen_khoa/phong_chuyen_khoa.html', context=data)
    else:
        phong_chuc_nang = PhongChucNang.objects.all()
        data = {
            'phong_chuc_nang': phong_chuc_nang,
        }
        return render(request, 'do_not_have_permission.html', context=data)
    

@login_required(login_url='/dang_nhap/')
def phan_khoa_kham(request, **kwargs):
    id_lich_hen = kwargs.get('id_lich_hen', None)
    lich_hen    = get_object_or_404(LichHenKham, id=id_lich_hen)
    benh_nhan = lich_hen.benh_nhan
    phong_chuc_nang = PhongChucNang.objects.all()
    mau_hoa_don = MauPhieu.objects.filter(codename='phieu_chi_dinh').first()
    bac_si_chi_dinh = request.user.ho_ten
    thoi_gian = datetime.now()

    data = {
        'id_lich_hen': id_lich_hen,
        'lich_hen'   : lich_hen,
        'phong_chuc_nang' : phong_chuc_nang,
        'benh_nhan': benh_nhan,
        'mau_hoa_don': mau_hoa_don,
        'bac_si_chi_dinh': bac_si_chi_dinh,
        'thoi_gian': f"{thoi_gian.strftime('%H:%m')} Ngày {thoi_gian.strftime('%d')} Tháng {thoi_gian.strftime('%m')} Năm {thoi_gian.strftime('%Y')}",
        'benh_nhan': f"Họ tên: {benh_nhan.ho_ten}",
        'so_dien_thoai': f"SĐT: {benh_nhan.get_so_dien_thoai()}",
        'dia_chi': f"Đ/C: {benh_nhan.get_dia_chi()}",
    }
    return render(request, 'bac_si_lam_sang/phan_khoa_kham.html', context=data)

def store_phan_khoa(request):
    if request.method == "POST":
        request_data = request.POST.get('data', None)
        user         = request.POST.get('user', None)
        id_lich_hen  = request.POST.get('id_lich_hen', None)
        data         = json.loads(request_data)

        now       = datetime.now()
        date_time = now.strftime("%m%d%y%H%M%S")

        bulk_create_data = []
        bao_hiem = False
        try:
            user = User.objects.get(id = user)
        except User.DoesNotExist:
            response = {
                'status' : 404,
                'message': "Bệnh nhân này không tồn tại, vui lòng thử lại",
            }

            return HttpResponse(json.dumps(response), content_type='application/json; charset=utf-8')
        subName = getSubName(user.ho_ten)
        ma_hoa_don = "HD" + "-" + subName + '-' + date_time
        ma_lk = f"CK{date_time}"

        try:
            lich_hen = LichHenKham.objects.get(id = id_lich_hen)
            trang_thai_lich_hen = TrangThaiLichHen.objects.get_or_create(ten_trang_thai = "Đã Phân Khoa")[0]
            lich_hen.trang_thai = trang_thai_lich_hen
            lich_hen.save()

            trang_thai = TrangThaiChuoiKham.objects.get_or_create(trang_thai_chuoi_kham="Chờ Thanh Toán")[0]
            chuoi_kham = ChuoiKham.objects.get_or_create(bac_si_dam_nhan=request.user, benh_nhan=user, trang_thai=trang_thai, lich_hen = lich_hen, ma_lk=ma_lk)[0]
            chuoi_kham.save()

        except LichHenKham.DoesNotExist:
            response = {
                'status' : 404,
                'message': "Lịch hẹn này không tồn tại, vui lòng thử lại",
            }
            return HttpResponse(json.dumps(response), content_type='application/json; charset=utf-8')
        
        trang_thai_phan_khoa = TrangThaiKhoaKham.objects.get_or_create(trang_thai_khoa_kham='Chờ Khám')[0]

        for i in data:
            if i['obj']['bao_hiem'] == "True":
                bao_hiem = True
            index = data.index(i)
            priority = index + 1
            dich_vu = DichVuKham.objects.only('id').filter(id=i['obj']['id']).first()
            bac_si = request.user
            bulk_create_data.append(
                PhanKhoaKham(
                    benh_nhan=user, 
                    dich_vu_kham=dich_vu, 
                    bao_hiem=i['obj']['bao_hiem'], 
                    bac_si_lam_sang=bac_si, 
                    chuoi_kham=chuoi_kham,
                    priority=priority,
                    trang_thai=trang_thai_phan_khoa
                    )
                )

        hoa_don = HoaDonChuoiKham.objects.create(chuoi_kham=chuoi_kham, ma_hoa_don=ma_hoa_don, bao_hiem = bao_hiem)

        PhanKhoaKham.objects.bulk_create(bulk_create_data)

        from actstream import action
        action.send(request.user, verb='phân khoa thành công cho bệnh nhân', target=user)

        response = {
            'status' : 200,
            'message': "Phân Khoa Khám Thành Công!",
            'url' : '/bac_si_lam_sang/danh_sach_benh_nhan_cho/'
        }

        channel_layer = get_channel_layer()
        async_to_sync(channel_layer.group_send)(
            f"user_{user.id}", {
                'type':'checkup_process_info'
            }
        )
        async_to_sync(channel_layer.group_send)(
            f"checkup_process_user_{lich_hen.benh_nhan.id}", {
                'type':'checkup_process_notification'
            }
        )

        
        return HttpResponse(json.dumps(response), content_type='application/json; charset=utf-8')
    else:
        response = {
            'status' : 404,
            'message': "Xảy Ra Lỗi Trong Quá Trình Xử Lí",
        }
        return HttpResponse(json.dumps(response), content_type='application/json; charset=utf-8')

def store_ke_don(request):
    if request.method == "POST":
        request_data = request.POST.get('data', None)
        user = request.POST.get('user', None)
        id_chuoi_kham = request.POST.get('id_chuoi_kham', None)
        data = json.loads(request_data)

        now = datetime.now()
        date_time = now.strftime("%m%d%y%H%M%S")

        try:
            chuoi_kham = ChuoiKham.objects.get(id=id_chuoi_kham)

        except ChuoiKham.DoesNotExist:
            repsonse = {
                "status": 404,
                "message": "Chuỗi Khám Này Không Tồn Tại",
            }
            return HttpResponse(json.dumps(repsonse), content_type='application/json; charset=utf-8')
        
        try:
            user = User.objects.get(id=user)
            subName = getSubName(user.ho_ten)
            ma_don_thuoc = subName + '-' + date_time
            trang_thai = TrangThaiDonThuoc.objects.get_or_create(trang_thai="Chờ Thanh Toán")[0]

            don_thuoc = DonThuoc.objects.get_or_create(benh_nhan=user, bac_si_ke_don=request.user, trang_thai=trang_thai, ma_don_thuoc=ma_don_thuoc, chuoi_kham = chuoi_kham)[0]
        
        except User.DoesNotExist:
            repsonse = {
                "status": 404,
                "message": "Bệnh Nhân Này Không Tồn Tại",
            }
            return HttpResponse(json.dumps(repsonse), content_type='application/json; charset=utf-8')
            
        bulk_create_data = []
        for i in data:
            thuoc = Thuoc.objects.only('id').get(id=i['obj']['id'])
            so_luong=i['obj']['so_luong']

            if thuoc.kha_dung and int(so_luong) > 0:
                ke_don_thuoc = KeDonThuoc(don_thuoc=don_thuoc, thuoc=thuoc, so_luong=i['obj']['so_luong'], cach_dung=i['obj']['duong_dung'], ghi_chu=i['obj']['ghi_chu'], bao_hiem=i['obj']['bao_hiem'])
                bulk_create_data.append(ke_don_thuoc)
            else:
                don_thuoc.delete()

                repsonse = {
                    "status": 404,
                    "message": f"Vui lòng kiểm tra lại số lượng tồn kho của thuốc: {thuoc.ten_thuoc}",
                }
                return HttpResponse(json.dumps(repsonse), content_type='application/json; charset=utf-8')
                
        KeDonThuoc.objects.bulk_create(bulk_create_data)

        from actstream import action
        action.send(request.user, verb='kê đơn thành công cho bệnh nhân', target=user)
        
        channel_layer = get_channel_layer()
        async_to_sync(channel_layer.group_send)(
            f"prescription_user_{user.id}", {
                'type':'prescription_notification'
            }
        )
        response = {'status': 200, 'message': 'Kê Đơn Thành Công', 'url': '/danh_sach_kham/'}
    else:
       response = {'message': 'oke'} 
    return HttpResponse(json.dumps(response), content_type='application/json; charset=utf-8')

def chinh_sua_don_thuoc(request):
    if request.method == "POST":
        request_data = request.POST.get('data')
        ma_don_thuoc = request.POST.get('ma_don_thuoc')
        data = json.loads(request_data)
        new_dict = {}
        so_luong = data[0]
        ghi_chu = data[1]
        # print(data)
        don_thuoc = DonThuoc.objects.get(ma_don_thuoc=ma_don_thuoc)
        for k in so_luong.keys():
            new_dict[k] = tuple(new_dict[k] for new_dict in data)
        
        for k in new_dict.keys():
            id_thuoc = k
            so_luong_thuoc = new_dict[k][0]
            ghi_chu_thuoc = new_dict[k][1]
            id_ke_don = new_dict[k][2]
            ke_don = KeDonThuoc.objects.get(id=id_ke_don)
            ke_don.so_luong = so_luong_thuoc
            ke_don.ghi_chu = ghi_chu_thuoc
            ke_don.save()

        from actstream import action
        action.send(request.user, verb='đã chỉnh sửa đơn thuốc', action_object=don_thuoc ,target=don_thuoc.benh_nhan)

        response = {'status': 200, 'message': 'Cập Nhật Thành Công'}
        return HttpResponse(json.dumps(response), content_type='application/json; charset=utf-8')

def files_upload_view(request):
    print(request.FILES)
    if request.method == "POST":
        ma_ket_qua = request.POST.get('ma_ket_qua', None)
        mo_ta = request.POST.get('mo_ta', None)
        ket_luan = request.POST.get('ket_qua', None)
        id_chuoi_kham = request.POST.get('id_chuoi_kham')
        print(id_chuoi_kham)

        if ma_ket_qua == '':
            HttpResponse({'status': 404, 'message': 'Mã Kết Quả Không Được Để Trống'})

        if mo_ta == '':
            HttpResponse({'status': 404, 'message': 'Mô Tả Không Được Để Trống'})

        if ket_luan == '':
            HttpResponse({'status': 404, 'message': 'Kết Luận Không Được Để Trống'})

        # chuoi_kham = ChuoiKham.objects.get(id=16)
        # ket_qua_tong_quat = KetQuaTongQuat.objects.get_or_create(chuoi_kham=chuoi_kham, ma_ket_qua=ma_ket_qua, mo_ta=mo_ta, ket_luan=ket_luan)[0]
        
        # for value in request.FILES.values():
        #     file = FileKetQua.objects.create(file=value)
        #     file_kq_tong_quat = FileKetQuaTongQuat.objects.create(ket_qua_tong_quat=ket_qua_tong_quat, file=file)
        #     file_kq_tong_quat.save()

        return HttpResponse('upload')
    return JsonResponse({'post': False})

def upload_files_chuyen_khoa(request):
    if request.method == "POST":
        ma_ket_qua    = request.POST.get('ma_ket_qua', None)
        mo_ta         = request.POST.get('mo_ta', None)
        ket_luan      = request.POST.get('ket_qua', None)
        id_chuoi_kham = request.POST.get('id_chuoi_kham')
        id_phan_khoa = request.POST.get('id_phan_khoa')

        if ma_ket_qua == '':
            HttpResponse({'status': 404, 'message': 'Mã Kết Quả Không Được Để Trống'})

        if mo_ta == '':
            HttpResponse({'status': 404, 'message': 'Mô Tả Không Được Để Trống'})

        if ket_luan == '':
            HttpResponse({'status': 404, 'message': 'Kết Luận Không Được Để Trống'})

        try:
            chuoi_kham = ChuoiKham.objects.get(id=id_chuoi_kham)
        except ChuoiKham.DoesNotExist:
            response = {
                'status': 404,
                'message': 'Chuỗi Khám Không Tồn Tại'
            }
            return HttpResponse(json.dumps(response), content_type='application/json; charset=utf-8')


        trang_thai = TrangThaiKhoaKham.objects.get_or_create(trang_thai_khoa_kham='Đã Tải Lên Kết Quả')[0]
        
        try:
            phan_khoa = PhanKhoaKham.objects.get(id=id_phan_khoa)
        except PhanKhoaKham.DoesNotExist:
            response = {
                'status': 404,
                'message': 'Phân Khoa Khám Không Tồn Tại'
            }
            return HttpResponse(json.dumps(response), content_type='application/json; charset=utf-8')

        phan_khoa.trang_thai = trang_thai
        phan_khoa.save()
        ket_qua_tong_quat = KetQuaTongQuat.objects.get_or_create(chuoi_kham=chuoi_kham)[0]
        ket_qua_chuyen_khoa = KetQuaChuyenKhoa.objects.create(phan_khoa_kham=phan_khoa, ket_qua_tong_quat=ket_qua_tong_quat, ma_ket_qua=ma_ket_qua, mo_ta=mo_ta, ket_luan=ket_luan)
    
        for value in request.FILES.values():
            file = FileKetQua.objects.create(file=value)
            file_kq_chuyen_khoa = FileKetQuaChuyenKhoa.objects.create(ket_qua_chuyen_khoa=ket_qua_chuyen_khoa, file=file)
        
        response = {
            'status': 200,
            'message' : 'Upload Thành Công!'
        }
        return HttpResponse(json.dumps(response), content_type='application/json; charset=utf-8')
    else:
        response = {
            'status': 404,
            'message': 'Có lỗi xảy ra trong quá trình xử lý'
        }
        return HttpResponse(json.dumps(response), content_type='application/json; charset=utf-8')

def upload_files_lam_sang(request):
    print(request.FILES)
    if request.method == "POST":
        ma_ket_qua = request.POST.get('ma_ket_qua', None)
        mo_ta = request.POST.get('mo_ta', None)
        ket_luan = request.POST.get('ket_qua', None)
        id_chuoi_kham = request.POST.get('id_chuoi_kham')

        if ma_ket_qua == '':
            HttpResponse({'status': 404, 'message': 'Mã Kết Quả Không Được Để Trống'})

        if mo_ta == '':
            HttpResponse({'status': 404, 'message': 'Mô Tả Không Được Để Trống'})

        if ket_luan == '':
            HttpResponse({'status': 404, 'message': 'Kết Luận Không Được Để Trống'})

        chuoi_kham = ChuoiKham.objects.get(id=id_chuoi_kham)
        ket_qua_tong_quat = KetQuaTongQuat.objects.get_or_create(chuoi_kham=chuoi_kham)[0]
        ket_qua_tong_quat.ma_ket_qua = ma_ket_qua
        ket_qua_tong_quat.mo_ta = mo_ta
        ket_qua_tong_quat.ket_luan = ket_luan
        ket_qua_tong_quat.save()

        for value in request.FILES.values():
            file = FileKetQua.objects.create(file=value)
            file_ket_qua_tong_quat = FileKetQuaTongQuat.objects.create(file=file, ket_qua_tong_quat=ket_qua_tong_quat)

        return HttpResponse('upload')
    return JsonResponse({'post': False})

@login_required(login_url='/dang_nhap/')
def upload_view(request, **kwargs):
    id_chuoi_kham = kwargs.get('id')
    chuoi_kham = ChuoiKham.objects.get(id = id_chuoi_kham)
    id_phong_chuc_nang = kwargs.get('id_phong_chuc_nang')
    phong_chuc_nang = PhongChucNang.objects.all()
    ten_phong_chuc_nang = PhongChucNang.objects.get(id = id_phong_chuc_nang)
    ho_ten_benh_nhân = chuoi_kham.benh_nhan.ho_ten
    ten_phong_chuc_nang = ten_phong_chuc_nang.ten_phong_chuc_nang
    now       = datetime.now()
    date_time = now.strftime("%m%d%y%H%M%S")

    ma_ket_qua = str(id_phong_chuc_nang) +'-'+ getSubName(ho_ten_benh_nhân) + '-' + str(date_time)
    data = {
        'id_chuoi_kham' : id_chuoi_kham,
        'chuoi_kham' : chuoi_kham,
        'phong_chuc_nang' : phong_chuc_nang,
        'id_phong_chuc_nang' : id_phong_chuc_nang,
        'ma_ket_qua': ma_ket_qua,
    }
    return render(request, 'bac_si_chuyen_khoa/upload.html', context=data)

@login_required(login_url='/dang_nhap/')
def upload_view_lam_sang(request, **kwargs):
    id_chuoi_kham = kwargs.get('id')
    chuoi_kham = ChuoiKham.objects.get(id=id_chuoi_kham)
    benh_nhan = chuoi_kham.benh_nhan
    now = datetime.now()
    date_time = now.strftime("%m%d%y%H%M%S")
 
    ma_ket_qua = getSubName(benh_nhan.ho_ten) +'-'+ str(date_time)
    phong_chuc_nang = PhongChucNang.objects.all()
    ngay_kham = datetime.now()
    mau_phieu = MauPhieu.objects.filter(codename='phieu_ket_qua_lam_sang').first()

    data_dict = {}
    data_dict['{benh_nhan}'] = benh_nhan.ho_ten.upper()
    data_dict['{gioi_tinh}'] = benh_nhan.get_gioi_tinh()
    data_dict['{tuoi}'] = benh_nhan.tuoi()
    data_dict['{dia_chi}'] = benh_nhan.get_dia_chi()
    data_dict['{bac_si_lam_sang}'] = request.user.ho_ten.upper()
    data_dict['{ngay_kham}'] = "Ngày " + str(ngay_kham.strftime('%d')) + " Tháng " + str(ngay_kham.strftime('%m')) + " Năm " + str(ngay_kham.strftime('%Y'))

    data = {
        'id_chuoi_kham' : id_chuoi_kham,
        'phong_chuc_nang' : phong_chuc_nang,
        'mau_phieu': mau_phieu,
        'ma_ket_qua' : ma_ket_qua,
        'data_dict': data_dict,
    }
    return render(request, 'bac_si_lam_sang/upload_ket_qua_lam_sang.html', context=data)

@login_required(login_url='/dang_nhap/')
def phong_tai_chinh_danh_sach_cho(request):
    trang_thai = TrangThaiChuoiKham.objects.all()
    phong_chuc_nang = PhongChucNang.objects.all()
    phong_kham = PhongKham.objects.all().first()
    mau_hoa_don = MauPhieu.objects.filter(codename='hoa_don_lam_sang').first()

    nguoi_thanh_toan = request.user.ho_ten
    thoi_gian_thanh_toan = datetime.now()
    
    data = {
        'trang_thai' : trang_thai,
        'phong_chuc_nang' : phong_chuc_nang,
        'phong_kham' : phong_kham,
        'mau_hoa_don': mau_hoa_don,
        'nguoi_thanh_toan': nguoi_thanh_toan,
        'thoi_gian_thanh_toan': f"{thoi_gian_thanh_toan.strftime('%H:%m')} Ngày {thoi_gian_thanh_toan.strftime('%d')} Tháng {thoi_gian_thanh_toan.strftime('%m')} Năm {thoi_gian_thanh_toan.strftime('%Y')}",
    }
    return render(request, 'phong_tai_chinh/danh_sach_thanh_toan.html', context= data)

@login_required(login_url='/dang_nhap/')
def phong_thuoc_danh_sach_cho(request):
    phong_chuc_nang = PhongChucNang.objects.all()
    trang_thai = TrangThaiDonThuoc.objects.all()

    data = {
        'phong_chuc_nang' : phong_chuc_nang,
        'trang_thai': trang_thai,
    }
    return render(request, 'phong_thuoc/danh_sach_cho.html', context=data)

@login_required(login_url='/dang_nhap/')
def hoa_don_dich_vu(request, **kwargs):
    id_chuoi_kham = kwargs.get('id_chuoi_kham')
    chuoi_kham = get_object_or_404(ChuoiKham, id=id_chuoi_kham)
    check_da_thanh_toan = chuoi_kham.check_thanh_toan()

    print(timezone.localtime(timezone.now()))

    if check_da_thanh_toan == True:
        hoa_don_dich_vu = chuoi_kham.hoa_don_dich_vu
        print(hoa_don_dich_vu)
        tong_tien_hoa_don = hoa_don_dich_vu.tong_tien
        if hoa_don_dich_vu.nguoi_thanh_toan is not None:
            nguoi_thuc_hien = hoa_don_dich_vu.nguoi_thanh_toan.ho_ten.upper()
        else:
            nguoi_thuc_hien = '-'

        if hoa_don_dich_vu.discount is not None:
            discount = hoa_don_dich_vu.discount
            tien_giam_gia = int(tong_tien_hoa_don) * (int(discount) / 100)
        else:
            tien_giam_gia = 0

        ma_hoa_don = hoa_don_dich_vu.ma_hoa_don
        danh_sach_phan_khoa = chuoi_kham.phan_khoa_kham.all()
        tong_tien = []
        bao_hiem = []
        for khoa_kham in danh_sach_phan_khoa:
            if khoa_kham.bao_hiem:
                gia = khoa_kham.dich_vu_kham.don_gia_bhyt
                bao_hiem.append(gia)
            else:
                gia = khoa_kham.dich_vu_kham.don_gia
            tong_tien.append(gia)
        total_spent = sum(tong_tien)
        tong_bao_hiem = sum(bao_hiem)
        tien_phai_thanh_toan  = int(tong_tien_hoa_don) - int(tien_giam_gia) - int(tong_bao_hiem)
        tong_tien.clear()
        bao_hiem.clear()
        phong_chuc_nang = PhongChucNang.objects.all()
        mau_hoa_don = MauPhieu.objects.filter(codename='hoa_don_dich_vu').first()
        thoi_gian_thanh_toan = hoa_don_dich_vu.thoi_gian_cap_nhat + timedelta(hours=7)
        print(thoi_gian_thanh_toan)
        
        benh_nhan = chuoi_kham.benh_nhan
        danh_sach_dich_vu = [f"{i.dich_vu_kham.ten_dvkt}" for i in danh_sach_phan_khoa]
        danh_sach_bao_hiem = ['Áp Dụng' if i.bao_hiem else 'Không Áp Dụng' for i in danh_sach_phan_khoa]
        danh_sach_gia_tien = [f"{i.get_dich_vu_gia()}" for i in danh_sach_phan_khoa]

        data_dict = {}
        data_dict['{benh_nhan}'] = benh_nhan.ho_ten.upper()
        data_dict['{so_dien_thoai}'] = benh_nhan.get_so_dien_thoai()
        data_dict['{dia_chi}'] = benh_nhan.get_dia_chi()
        data_dict['{thoi_gian_thanh_toan}'] = f"Ngày {thoi_gian_thanh_toan.strftime('%d')} Tháng {thoi_gian_thanh_toan.strftime('%m')} Năm {thoi_gian_thanh_toan.strftime('%Y')}"
        data_dict['{tong_tien}'] = "{:,}".format(int(total_spent))
        data_dict['{tong_tien_bao_hiem}'] = "{:,}".format(int(tong_bao_hiem))
        data_dict['{nguoi_thanh_toan}'] = nguoi_thuc_hien
        data_dict['{giam_gia}'] = "{:,}".format(int(tien_giam_gia))
        data_dict['{tien_thanh_toan}'] = "{:,}".format(int(tien_phai_thanh_toan))

        data = {
            'phong_chuc_nang'    : phong_chuc_nang,
            'mau_hoa_don': mau_hoa_don,
            'danh_sach_dich_vu': danh_sach_dich_vu,
            'danh_sach_bao_hiem': danh_sach_bao_hiem,
            'danh_sach_gia_tien': danh_sach_gia_tien,
            'tong_tien': "{:,}".format(int(total_spent)),
            'tong_tien_bao_hiem': "{:,}".format(int(tong_bao_hiem)),        
            'nguoi_thuc_hien': nguoi_thuc_hien,
            'data_tong_tien': total_spent,
            'ma_hoa_don': ma_hoa_don,
            'id_chuoi_kham': id_chuoi_kham,
            'check_da_thanh_toan': check_da_thanh_toan,
            'discount': "{:,}".format(int(tien_giam_gia)),
            'tien_phai_thanh_toan': "{:,}".format(int(tien_phai_thanh_toan)),
            'data_dict': data_dict,

        }
        return render(request, 'phong_tai_chinh/hoa_don_dich_vu.html', context=data)
    else:
        hoa_don_dich_vu = chuoi_kham.hoa_don_dich_vu
        ma_hoa_don = hoa_don_dich_vu.ma_hoa_don
        danh_sach_phan_khoa = chuoi_kham.phan_khoa_kham.all()
        tong_tien = []
        bao_hiem = []
        for khoa_kham in danh_sach_phan_khoa:
            if khoa_kham.bao_hiem:
                gia = khoa_kham.dich_vu_kham.don_gia_bhyt
                bao_hiem.append(gia)
            else:
                gia = khoa_kham.dich_vu_kham.don_gia
            tong_tien.append(gia)
        total_spent = sum(tong_tien)
        tong_bao_hiem = sum(bao_hiem)
        thanh_tien = total_spent - tong_bao_hiem
        tong_tien.clear()
        bao_hiem.clear()
        phong_chuc_nang = PhongChucNang.objects.all()
        mau_hoa_don = MauPhieu.objects.filter(codename='hoa_don_dich_vu').first()
        thoi_gian_thanh_toan = local_time
       
        benh_nhan = chuoi_kham.benh_nhan
        danh_sach_dich_vu = [f"{i.dich_vu_kham.ten_dvkt}" for i in danh_sach_phan_khoa]
        danh_sach_bao_hiem = ['Áp Dụng' if i.bao_hiem else 'Không Áp Dụng' for i in danh_sach_phan_khoa]
        danh_sach_gia_tien = [f"{i.get_dich_vu_gia()}" for i in danh_sach_phan_khoa]

        data_dict = {}
        data_dict['{benh_nhan}'] = benh_nhan.ho_ten.upper()
        data_dict['{so_dien_thoai}'] = benh_nhan.get_so_dien_thoai()
        data_dict['{dia_chi}'] = benh_nhan.get_dia_chi()
        data_dict['{thoi_gian_thanh_toan}'] = f"Ngày {thoi_gian_thanh_toan.strftime('%d')} Tháng {thoi_gian_thanh_toan.strftime('%m')} Năm {thoi_gian_thanh_toan.strftime('%Y')}"
        data_dict['{tong_tien}'] = "{:,}".format(int(total_spent))
        data_dict['{tong_tien_bao_hiem}'] = "{:,}".format(int(tong_bao_hiem))
        data_dict['{nguoi_thanh_toan}'] = request.user.ho_ten.upper()
        data_dict['{giam_gia}'] = "0"
        data_dict['{tien_thanh_toan}'] = "{:,}".format(int(thanh_tien))

        data = {
            'phong_chuc_nang'    : phong_chuc_nang,
            'mau_hoa_don': mau_hoa_don,
            'danh_sach_dich_vu': danh_sach_dich_vu,
            'danh_sach_bao_hiem': danh_sach_bao_hiem,
            'danh_sach_gia_tien': danh_sach_gia_tien,
            'tong_tien': "{:,}".format(int(total_spent)),
            'tong_tien_bao_hiem': "{:,}".format(int(tong_bao_hiem)),
            'thanh_tien': "{:,}".format(int(thanh_tien)),
            'nguoi_thuc_hien': request.user.ho_ten.upper(),
            'data_tong_tien': total_spent,
            'data_thanh_tien': thanh_tien,
            'ma_hoa_don': ma_hoa_don,
            'id_chuoi_kham': id_chuoi_kham,
            'check_da_thanh_toan': check_da_thanh_toan,
            'data_dict': data_dict,

        }
        return render(request, 'phong_tai_chinh/hoa_don_dich_vu.html', context=data)

@login_required(login_url='/dang_nhap/')
def hoa_don_thuoc(request, **kwargs):
    id_don_thuoc = kwargs.get('id_don_thuoc')

    don_thuoc = get_object_or_404(DonThuoc, id=id_don_thuoc)
    danh_sach_thuoc = don_thuoc.ke_don.all()
    check_da_thanh_toan = don_thuoc.check_thanh_toan()

    if check_da_thanh_toan == True:
        hoa_don_thuoc = don_thuoc.hoa_don_thuoc
        if hoa_don_thuoc.nguoi_thanh_toan is not None:
            nguoi_thuc_hien = hoa_don_thuoc.nguoi_thanh_toan.ho_ten.upper()
        else:
            nguoi_thuc_hien = '-'
        thoi_gian_thanh_toan = hoa_don_thuoc.thoi_gian_tao
    else:
        nguoi_thuc_hien = request.user.ho_ten.upper()
        thoi_gian_thanh_toan = datetime.now()

    _danh_sach_thuoc = []
    _danh_sach_thuc_pham_chuc_nang = []
    bao_hiem = []
    tong_tien = []
    for thuoc_instance in danh_sach_thuoc:
        if thuoc_instance.bao_hiem:
            gia = int(thuoc_instance.thuoc.don_gia_tt) * \
                thuoc_instance.so_luong
            bao_hiem.append(gia)
        else:
            gia = int(thuoc_instance.thuoc.don_gia_tt) * \
                thuoc_instance.so_luong

        if thuoc_instance.thuoc.check_loai_thuoc:
            _danh_sach_thuc_pham_chuc_nang.append(thuoc_instance)
        else:
            _danh_sach_thuoc.append(thuoc_instance)

        tong_tien.append(gia)

    total_spent = sum(tong_tien)
    tong_bao_hiem = sum(bao_hiem)
    thanh_tien = total_spent - tong_bao_hiem

    tong_tien.clear()
    bao_hiem.clear()
    
    phong_chuc_nang = PhongChucNang.objects.all()
    mau_hoa_don_thuoc = MauPhieu.objects.filter(codename='hoa_don_thuoc').first()
    mau_hoa_don_tphtdt = MauPhieu.objects.filter(codename='hoa_don_thuc_pham_ho_tro_dieu_tri').first()

    benh_nhan = don_thuoc.benh_nhan
    
    ds_thuoc = [f'{i.thuoc.ten_thuoc}' for i in _danh_sach_thuoc]
    ds_thuc_pham_chuc_nang = [f'{i.thuoc.ten_thuoc}' for i in _danh_sach_thuc_pham_chuc_nang]
    danh_sach_bao_hiem_thuoc = ['Áp Dụng' if i.bao_hiem else 'Không Áp Dụng' for i in _danh_sach_thuoc]
    danh_sach_bao_hiem_thuc_pham_cn = ['Áp Dụng' if i.bao_hiem else 'Không Áp Dụng' for i in _danh_sach_thuc_pham_chuc_nang]
    danh_sach_don_gia_thuoc = [f'{i.thuoc.get_don_gia_tt()}' for i in _danh_sach_thuoc]
    danh_sach_don_gia_thuc_pham_cn = [f'{i.thuoc.get_don_gia_tt()}' for i in _danh_sach_thuc_pham_chuc_nang]
    danh_sach_so_luong_thuoc = [f'{i.so_luong}' for i in _danh_sach_thuoc]
    danh_sach_so_luong_thuc_pham_cn = [f'{i.so_luong}' for i in _danh_sach_thuc_pham_chuc_nang]
    danh_sach_don_vi_tinh_thuoc = [f'{i.thuoc.don_vi_tinh}' for i in _danh_sach_thuoc]
    danh_sach_don_vi_tinh_thuc_pham_cn = [f'{i.thuoc.don_vi_tinh}' for i in _danh_sach_thuc_pham_chuc_nang]
    danh_sach_thanh_tien_thuoc = [f'{i.get_thanh_tien()}' for i in _danh_sach_thuoc]
    danh_sach_thanh_tien_thuc_pham_cn = [f'{i.get_thanh_tien()}' for i in _danh_sach_thuc_pham_chuc_nang]
    tong_tien_thuoc = [i.get_tong_tien() for i in _danh_sach_thuoc]
    tong_tien_thuoc_str = "{:,}".format(int(sum(tong_tien_thuoc)))
    tong_tien_thuc_pham_cn = [i.get_tong_tien() for i in _danh_sach_thuc_pham_chuc_nang]
    tong_tien_thuc_pham_cn_str = "{:,}".format(int(sum(tong_tien_thuc_pham_cn)))
    tong_bao_hiem_thuoc = [i.get_tong_tien_bao_hiem() for i in _danh_sach_thuoc]
    bao_hiem_thuoc_str = "{:,}".format(int(sum(tong_bao_hiem_thuoc)))
    tong_bao_hiem_thuc_pham_cn = [i.get_tong_tien_bao_hiem() for i in _danh_sach_thuc_pham_chuc_nang]
    bao_hiem_thuc_pham_cn_str = "{:,}".format(int(sum(tong_bao_hiem_thuc_pham_cn)))
    thanh_toan_thuoc = sum(tong_tien_thuoc) - sum(tong_bao_hiem_thuoc)
    thanh_toan_thuoc_str = "{:,}".format(int(thanh_toan_thuoc))
    thanh_toan_thuc_pham_cn = sum(tong_tien_thuc_pham_cn) - sum(tong_bao_hiem_thuc_pham_cn)
    thanh_toan_thuc_pham_cn_str = "{:,}".format(int(thanh_toan_thuc_pham_cn))
    
    if don_thuoc.benh_nhan is not None:
        ten_benh_nhan = f"Họ tên: {benh_nhan.ho_ten.upper()}"
        so_dien_thoai = f"SĐT: {benh_nhan.get_so_dien_thoai()}"
        dia_chi = f"Đ/C: {benh_nhan.get_dia_chi()}"
    else:
        ten_benh_nhan = f"Họ tên: {don_thuoc.benh_nhan_vang_lai}"
        so_dien_thoai = f"SĐT: Không có"
        dia_chi = f'Đ/C: Không có'
    
    data_dict = {}
    data_dict['{benh_nhan}'] = ten_benh_nhan
    data_dict['{so_dien_thoai}'] = so_dien_thoai
    data_dict['{dia_chi}'] = dia_chi
    data_dict['{tong_tien}'] = tong_tien_thuoc_str
    data_dict['{tong_tien_bao_hiem}'] = bao_hiem_thuoc_str
    data_dict['{tien_thanh_toan}'] = thanh_toan_thuoc_str
    data_dict['{nguoi_thanh_toan}'] = nguoi_thuc_hien
    data_dict['{thoi_gian_thanh_toan}'] = f"Ngày {thoi_gian_thanh_toan.strftime('%d')} Tháng {thoi_gian_thanh_toan.strftime('%m')} Năm {thoi_gian_thanh_toan.strftime('%Y')}"

    data = { 
        'phong_chuc_nang': phong_chuc_nang,
        'don_thuoc': don_thuoc,
        'flag': don_thuoc.check_tphtdt,
        'mau_hoa_don_thuoc': mau_hoa_don_thuoc,
        'mau_hoa_don_tphtdt': mau_hoa_don_tphtdt,
        'thoi_gian_thanh_toan': f"{thoi_gian_thanh_toan.strftime('%H:%m')} Ngày {thoi_gian_thanh_toan.strftime('%d')} Tháng {thoi_gian_thanh_toan.strftime('%m')} Năm {thoi_gian_thanh_toan.strftime('%Y')}",
        'benh_nhan': ten_benh_nhan,
        'so_dien_thoai': so_dien_thoai,
        'dia_chi': dia_chi,
        'danh_sach_thuoc': ds_thuoc,
        'danh_sach_thuc_pham_chuc_nang': ds_thuc_pham_chuc_nang,
        'danh_sach_bao_hiem_thuoc': danh_sach_bao_hiem_thuoc,
        'danh_sach_bao_hiem_thuc_pham_cn': danh_sach_bao_hiem_thuc_pham_cn,
        'danh_sach_don_gia_thuoc': danh_sach_don_gia_thuoc,
        'danh_sach_don_gia_thuc_pham_cn': danh_sach_don_gia_thuc_pham_cn,
        'danh_sach_so_luong_thuoc': danh_sach_so_luong_thuoc,
        'danh_sach_so_luong_thuc_pham_cn': danh_sach_so_luong_thuc_pham_cn,
        'danh_sach_don_vi_tinh_thuoc': danh_sach_don_vi_tinh_thuoc,
        'danh_sach_don_vi_tinh_thuc_pham_cn': danh_sach_don_vi_tinh_thuc_pham_cn,
        'danh_sach_thanh_tien_thuoc': danh_sach_thanh_tien_thuoc,
        'danh_sach_thanh_tien_thuc_pham_cn': danh_sach_thanh_tien_thuc_pham_cn,
        'tong_tien_thuoc_str': tong_tien_thuoc_str,
        'tong_tien_thuc_pham_cn_str': tong_tien_thuc_pham_cn_str,
        'bao_hiem_thuoc_str': bao_hiem_thuoc_str,
        'bao_hiem_thuc_pham_cn_str': bao_hiem_thuc_pham_cn_str,
        'thanh_toan_thuoc_str': thanh_toan_thuoc_str,
        'thanh_toan_thuc_pham_cn_str': thanh_toan_thuc_pham_cn_str,
        'nguoi_thuc_hien': nguoi_thuc_hien,
        'id_don_thuoc': id_don_thuoc,
        'tong_tien_thanh_toan': thanh_tien,
        'check_da_thanh_toan': check_da_thanh_toan,
        'data_dict': data_dict
    }
    return render(request, 'phong_tai_chinh/hoa_don_thuoc.html', context=data)

@login_required(login_url='/dang_nhap/')
def don_thuoc(request, **kwargs):
    id_don_thuoc = kwargs.get('id_don_thuoc')
    don_thuoc = get_object_or_404(DonThuoc, id=id_don_thuoc)

    danh_sach_thuoc = don_thuoc.ke_don.all()

    if don_thuoc.benh_nhan is not None:
        benh_nhan = "Họ và tên: " + don_thuoc.benh_nhan.ho_ten.upper()
        so_dien_thoai = 'SĐT: ' + str(don_thuoc.benh_nhan.get_so_dien_thoai())
        dia_chi = "Đ/c: " + don_thuoc.benh_nhan.get_dia_chi()

    elif don_thuoc.benh_nhan_vang_lai is not None:
        benh_nhan = "Họ và tên: " + don_thuoc.benh_nhan_vang_lai.upper()
        so_dien_thoai = 'SĐT: -'
        dia_chi = "Đ/c: -"

    _danh_sach_thuoc = []
    _danh_sach_thuc_pham_chuc_nang = []

    for thuoc_instance in danh_sach_thuoc:
        if thuoc_instance.thuoc.check_loai_thuoc:
            _danh_sach_thuc_pham_chuc_nang.append(thuoc_instance)
        else:
            _danh_sach_thuoc.append(thuoc_instance)

    phong_chuc_nang = PhongChucNang.objects.all()

    mau_hoa_don = MauPhieu.objects.filter(codename='don_thuoc').first()
    mau_hoa_don_thuc_pham_cn = MauPhieu.objects.filter(codename='don_thuoc_ho_tro_dieu_tri').first()

    nguoi_thanh_toan = don_thuoc.bac_si_ke_don.ho_ten
    thoi_gian_thanh_toan = don_thuoc.thoi_gian_tao

    ds_thuoc = [f'{i.thuoc.ten_thuoc}' for i in _danh_sach_thuoc]
    ds_thuc_pham_chuc_nang = [f'{i.thuoc.ten_thuoc}' for i in _danh_sach_thuc_pham_chuc_nang]
    danh_sach_so_luong_thuoc = [f'{i.so_luong}' for i in _danh_sach_thuoc]
    danh_sach_so_luong_thuc_pham_cn = [f'{i.so_luong}' for i in _danh_sach_thuc_pham_chuc_nang]
    danh_sach_duong_dung_thuoc = [f'{i.cach_dung}' for i in _danh_sach_thuoc]
    danh_sach_duong_dung_thuc_pham_cn = [f'{i.cach_dung}' for i in _danh_sach_thuc_pham_chuc_nang]
    danh_sach_ghi_chu_thuoc = [f'{i.ghi_chu}' for i in _danh_sach_thuoc]
    danh_sach_ghi_chu_thuc_pham_cn = [f'{i.ghi_chu}' for i in _danh_sach_thuc_pham_chuc_nang]
    danh_sach_don_vi_tinh_thuoc = [f'{i.thuoc.don_vi_tinh}' for i in _danh_sach_thuoc]
    danh_sach_don_vi_tinh_thuc_pham_cn = [f'{i.thuoc.don_vi_tinh}' for i in _danh_sach_thuc_pham_chuc_nang]

    data_dict = {}
    data_dict['{benh_nhan}'] = benh_nhan
    data_dict['{so_dien_thoai}'] = so_dien_thoai
    data_dict['{dia_chi}'] = dia_chi
    data_dict['{thoi_gian_thanh_toan}'] = f"Ngày {thoi_gian_thanh_toan.strftime('%d')} Tháng {thoi_gian_thanh_toan.strftime('%m')} Năm {thoi_gian_thanh_toan.strftime('%Y')}"
    data_dict['{nguoi_thanh_toan}'] = nguoi_thanh_toan

    data = {
        'danh_sach_thuoc': danh_sach_thuoc,
        'don_thuoc' : don_thuoc,
        'id_don_thuoc': id_don_thuoc,
        'phong_chuc_nang' : phong_chuc_nang,
        'ten_thuoc': ds_thuoc,
        'ten_thuoc_cn': ds_thuc_pham_chuc_nang,
        'so_luong': danh_sach_so_luong_thuoc,
        'so_luong_thuoc_cn': danh_sach_so_luong_thuc_pham_cn,
        'duong_dung': danh_sach_duong_dung_thuoc,
        'duong_dung_thuoc_cn': danh_sach_duong_dung_thuc_pham_cn,
        'ghi_chu': danh_sach_ghi_chu_thuoc,
        'ghi_chu_thuoc_cn': danh_sach_ghi_chu_thuc_pham_cn,
        'don_vi_tinh': danh_sach_don_vi_tinh_thuoc,
        'don_vi_tinh_thuoc_cn': danh_sach_don_vi_tinh_thuc_pham_cn,
        'mau_hoa_don': mau_hoa_don,
        'mau_hoa_don_thuc_pham_cn': mau_hoa_don_thuc_pham_cn,
        'data_dict': data_dict,
    }
    return render(request, 'phong_thuoc/don_thuoc.html', context=data)

@login_required(login_url='/dang_nhap/')
def danh_sach_kham(request):
    trang_thai = TrangThaiLichHen.objects.all()
    trang_thai_ck = TrangThaiChuoiKham.objects.all()
    phong_chuc_nang = PhongChucNang.objects.all()

    return render(request, 'bac_si_lam_sang/danh_sach_kham.html', context={"trang_thai": trang_thai, "trang_thai_ck": trang_thai_ck, 'phong_chuc_nang' : phong_chuc_nang})


def login(request):
    return render(request, 'registration/login.html')

def logout_view(request):
    logout(request)
    return HttpResponseRedirect(reverse('dang_nhap'))

class BatDauChuoiKhamToggle(APIView):
    def get(self, request, format=None, **kwargs):
        id_phan_khoa = request.GET.get('id', None)
        print(id_phan_khoa)
        phan_khoa_kham = PhanKhoaKham.objects.get(id=id_phan_khoa)
        chuoi_kham = phan_khoa_kham.chuoi_kham
        trang_thai_chuoi_kham = TrangThaiChuoiKham.objects.get_or_create(trang_thai_chuoi_kham="Đang Thực Hiện")[0]
        trang_thai_phan_khoa = TrangThaiKhoaKham.objects.get_or_create(trang_thai_khoa_kham="Hoàn Thành")[0]
        priotity = chuoi_kham.phan_khoa_kham.all().aggregate(Max('priority'))
        now = timezone.localtime(timezone.now())
        if phan_khoa_kham.priority == 1:
            chuoi_kham.thoi_gian_bat_dau = now
            chuoi_kham.trang_thai = trang_thai_chuoi_kham
            phan_khoa_kham.thoi_gian_bat_dau = now
            phan_khoa_kham.trang_thai = trang_thai_phan_khoa
            phan_khoa_kham.thoi_gian_ket_thuc = now
            chuoi_kham.save()
            phan_khoa_kham.save()
            dich_vu = phan_khoa_kham.dich_vu_kham.ten_dvkt

            from actstream import action
            action.send(request.user, verb='bắt đầu khám', action_object=phan_khoa_kham.dich_vu_kham, target=phan_khoa_kham.benh_nhan)

            response = {'status': '200', 'message': f'Bắt Đầu Chuỗi Khám, Dịch Vụ Khám Đầu Tiên: {dich_vu}', 'time': f'{now}'}
            
            channel_layer = get_channel_layer()
            async_to_sync(channel_layer.group_send)(
                f"user_{phan_khoa_kham.benh_nhan.id}", {
                    'type':'checkup_process_info'
                }
            )
            async_to_sync(channel_layer.group_send)(
                f"funcroom_service", {
                    'type': 'funcroom_info'
                }
            )
            
            return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")

        elif phan_khoa_kham.priority == priotity['priority__max']:
            chuoi_kham.thoi_gian_ket_thuc = now
            phan_khoa_kham.thoi_gian_bat_dau = now
            phan_khoa_kham.thoi_gian_ket_thuc = now
            chuoi_kham.trang_thai = trang_thai_chuoi_kham
            phan_khoa_kham.trang_thai = trang_thai_phan_khoa
            chuoi_kham.save()
            phan_khoa_kham.save()
            dich_vu = phan_khoa_kham.dich_vu_kham.ten_dvkt
            from actstream import action
            action.send(request.user, verb='bắt đầu khám', action_object=phan_khoa_kham.dich_vu_kham, target=phan_khoa_kham.benh_nhan)

            response = {'status': '200', 'message': f'Bắt Đầu Chuỗi Khám, Dịch Vụ Khám Đầu Tiên: {dich_vu}', 'time': f'{now}'}
            
            channel_layer = get_channel_layer()
            async_to_sync(channel_layer.group_send)(
                f"user_{phan_khoa_kham.benh_nhan.id}", {
                    'type':'checkup_process_info'
                }
            )
            async_to_sync(channel_layer.group_send)(
                f"funcroom_service", {
                    'type': 'funcroom_info'
                }
            )
            
            return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")
        else:
            phan_khoa_kham.thoi_gian_bat_dau = now
            phan_khoa_kham.thoi_gian_ket_thuc = now
            chuoi_kham.trang_thai = trang_thai_chuoi_kham
            phan_khoa_kham.trang_thai = trang_thai_phan_khoa
            chuoi_kham.save()
            phan_khoa_kham.save()
            dich_vu = phan_khoa_kham.dich_vu_kham.ten_dvkt

            from actstream import action
            action.send(request.user, verb='bắt đầu khám', action_object=phan_khoa_kham.dich_vu_kham, target=phan_khoa_kham.benh_nhan)

            response = {'status': '200', 'message': f'Bắt Đầu Dịch Vụ: {dich_vu}', 'time': f'{now}'}
            
            channel_layer = get_channel_layer()
            async_to_sync(channel_layer.group_send)(
                f"user_{phan_khoa_kham.benh_nhan.id}", {
                    'type':'checkup_process_info'
                }
            )
            async_to_sync(channel_layer.group_send)(
                f"funcroom_service", {
                    'type': 'funcroom_info'
                }
            )
            
            return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")
 
class KetThucChuoiKhamToggle(APIView):
    def get(self, request, format=None, **kwargs):
        id_phan_khoa = request.GET.get('id', None)
        print(id_phan_khoa)
        phan_khoa_kham = PhanKhoaKham.objects.get(id=id_phan_khoa)
        chuoi_kham = phan_khoa_kham.chuoi_kham
        trang_thai_chuoi_kham = TrangThaiChuoiKham.objects.get_or_create(trang_thai_chuoi_kham="Đang Thực Hiện")[0]
        trang_thai_phan_khoa = TrangThaiKhoaKham.objects.get_or_create(trang_thai_khoa_kham="Hoàn Thành")[0]
        priotity = chuoi_kham.phan_khoa_kham.all().aggregate(Max('priority'))
        now = timezone.localtime(timezone.now())
        print(priotity['priority__max'])
        if phan_khoa_kham.priority == priotity['priority__max']:
            chuoi_kham.thoi_gian_ket_thuc = now
            phan_khoa_kham.thoi_gian_ket_thuc = now
            chuoi_kham.trang_thai = trang_thai_chuoi_kham
            phan_khoa_kham.trang_thai = trang_thai_phan_khoa
            chuoi_kham.save()
            phan_khoa_kham.save()
            dich_vu = phan_khoa_kham.dich_vu_kham.ten_dvkt

            from actstream import action
            action.send(request.user, verb='kết thúc khám', action_object=phan_khoa_kham.dich_vu_kham, target=phan_khoa_kham.benh_nhan)

            response = {'status': '200', 'message': f'Kết Thúc Chuỗi Khám, Dịch Vụ Khám Cuối Cùng: {dich_vu}', 'time': f'{now}'}
            
            channel_layer = get_channel_layer()
            async_to_sync(channel_layer.group_send)(
                f"user_{phan_khoa_kham.benh_nhan.id}", {
                    'type':'checkup_process_info'
                }
            )
            async_to_sync(channel_layer.group_send)(
                f"funcroom_service", {
                    'type': 'funcroom_info'
                }
            )

            return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")
        else:
            phan_khoa_kham.thoi_gian_ket_thuc = now
            phan_khoa_kham.trang_thai = trang_thai_phan_khoa
            phan_khoa_kham.save()
            dich_vu = phan_khoa_kham.dich_vu_kham.ten_dvkt

            from actstream import action
            action.send(request.user, verb='kết thúc khám', action_object=phan_khoa_kham.dich_vu_kham, target=phan_khoa_kham.benh_nhan)

            response = {'status': '200', 'message': f'Kết Thúc Dịch Vụ: {dich_vu}', 'time': f'{now}'}
            
            channel_layer = get_channel_layer()
            async_to_sync(channel_layer.group_send)(
                f"user_{phan_khoa_kham.benh_nhan.id}", {
                    'type':'checkup_process_info'
                }
            )
            async_to_sync(channel_layer.group_send)(
                f"funcroom_service", {
                    'type': 'funcroom_info'
                }
            )

            return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")

def bat_dau_chuoi_kham(request, **kwargs):
    id_phan_khoa = kwargs.get('id', None)
    phan_khoa_kham = PhanKhoaKham.objects.get(id=id_phan_khoa)
    chuoi_kham = phan_khoa_kham.chuoi_kham
    now = timezone.localtime(timezone.now())
    if phan_khoa_kham.priority == 1:
        chuoi_kham.thoi_gian_bat_dau = now
        phan_khoa_kham.thoi_gian_bat_dau = now
        chuoi_kham.save()
        phan_khoa_kham.save()
        dich_vu = phan_khoa_kham.dich_vu_kham.ten_dvkt
        response = {'status': '200', 'message': f'Bắt Đầu Chuỗi Khám, Dịch Vụ Khám: {dich_vu}', 'time': f'{now}'}
        return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")
    else:
        phan_khoa_kham.thoi_gian_bat_dau = now
        phan_khoa_kham.save()
        dich_vu = phan_khoa_kham.dich_vu_kham.ten_dvkt
        response = {'status': '200', 'message': f'Bắt Đầu Dịch Vụ: {dich_vu}', 'time': f'{now}'}
        return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")

@login_required(login_url='/dang_nhap/')
def danh_sach_thuoc(request):
    phong_chuc_nang = PhongChucNang.objects.all()
    nhom_thuoc = NhomThuoc.objects.all()

    data = {
        'phong_chuc_nang' : phong_chuc_nang,
        'nhom_thuoc': nhom_thuoc,
    }
    return render(request, 'phong_thuoc/danh_sach_thuoc.html', context=data)

@login_required(login_url='/dang_nhap/')
def danh_sach_thuoc_phong_tai_chinh(request):
    nhom_thau = NhomThau.objects.all()
    phong_chuc_nang = PhongChucNang.objects.all()
    nhom_thuoc = NhomThuoc.objects.all()

    data={
        'nhom_thau': nhom_thau,
        'phong_chuc_nang': phong_chuc_nang,
        'nhom_thuoc': nhom_thuoc,
    }
    return render(request, 'phong_tai_chinh/danh_sach_thuoc.html', context = data)

@login_required(login_url='/dang_nhap/')
def them_moi_thuoc_phong_tai_chinh(request):
    phong_chuc_nang = PhongChucNang.objects.all()
    cong_ty = CongTy.objects.all()
    data = {
        'cong_ty': cong_ty,
        'phong_chuc_nang' : phong_chuc_nang,
    }
    return render(request, 'phong_tai_chinh/them_moi_thuoc.html', context=data)

@login_required(login_url='/dang_nhap/')
def cong_ty(request):
    phong_chuc_nang = PhongChucNang.objects.all()

    data = {
        'phong_chuc_nang' : phong_chuc_nang,
    }
    return render(request, 'phong_tai_chinh/nguon_cung.html', context=data)

@login_required(login_url='/dang_nhap/')
def update_lich_hen(request, **kwargs):
    id_lich_hen = kwargs.get('id')
    phong_chuc_nang = PhongChucNang.objects.all()
    lich_hen_kham = LichHenKham.objects.get(id=id_lich_hen)
    data = {
        'lich_hen' : lich_hen_kham,
        'phong_chuc_nang' : phong_chuc_nang,
    }
    return render(request, 'le_tan/update_lich_hen.html', context=data)

@login_required(login_url='/dang_nhap/')
def danh_sach_lich_hen(request):
    trang_thai = TrangThaiLichHen.objects.all()
    phong_chuc_nang = PhongChucNang.objects.all()
    nguoi_dung = User.objects.filter(chuc_nang=1)
    data = {
        'trang_thai' : trang_thai,
        'nguoi_dung' : nguoi_dung,
        'phong_chuc_nang' : phong_chuc_nang,
    }
    return render(request, 'le_tan/danh_sach_lich_hen.html', context=data)

def store_update_lich_hen(request):
    if request.method == 'POST':
        thoi_gian_bat_dau = request.POST.get('thoi_gian_bat_dau')
        id_lich_hen = request.POST.get('id')
        thoi_gian_bat_dau = datetime.strptime(thoi_gian_bat_dau, format_2)
        thoi_gian = thoi_gian_bat_dau.strftime("%Y-%m-%d %H:%M")
        trang_thai = TrangThaiLichHen.objects.get(ten_trang_thai = "Xác Nhận")
        lich_hen = LichHenKham.objects.get(id=id_lich_hen)
        lich_hen.thoi_gian_bat_dau = thoi_gian
        lich_hen.trang_thai = trang_thai
        lich_hen.save()

        from actstream import action
        action.send(request.user, verb='thay đổi lịch hẹn', action_object=lich_hen, target=lich_hen.benh_nhan)

        response = {
            "message" : "Cập Nhật Thành Công"
        }
        return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")

@login_required(login_url='/dang_nhap/')   
def them_thuoc_phong_tai_chinh(request):
    cong_ty = CongTy.objects.all()
    phong_chuc_nang = PhongChucNang.objects.all()
    
    data = {
        'cong_ty' : cong_ty,
        'phong_chuc_nang' : phong_chuc_nang,
    }
    return render(request, 'phong_tai_chinh/them_moi_thuoc.html', context=data)

@transaction.atomic
def xuat(request, id=None, so_luong=None):
    try:
        thuoc = Thuoc.objects.filter(id=id)
        if thuoc[0].kha_dung:
            thuoc.update(so_luong_kha_dung=F('so_luong_kha_dung') - so_luong)
            ThuocLog.objects.create(thuoc=thuoc[0], ngay=timezone.now(), quy_trinh=ThuocLog.OUT, so_luong=so_luong)
        else:
            return Response({"error": True, "message": "So Luong Thuoc Kha Dung = 0, Khong The Xuat Thuoc"})  
        return Response({"error": False, "message": f"Xuat Thuoc Thanh Cong: {so_luong} {thuoc[0].ten_thuoc}"})
    except:
        return Response({"error": True, "message": "Loi Tao Log Thuoc"})

@transaction.atomic
def nhap(id, so_luong):
    try:
        thuoc = Thuoc.objects.filter(id=id)
        print(thuoc)
        print(so_luong)
        thuoc.update(so_luong_kha_dung=F('so_luong_kha_dung') + so_luong)
        print('updated')
        ThuocLog.objects.create(thuoc=thuoc[0], ngay=timezone.now(), quy_trinh=ThuocLog.IN, so_luong=so_luong)
        return Response({'error': False, 'message': 'Nhập Thuốc Thành Công'})
    except:
        return Response({'error': True, 'message': 'Không Thể Nhập Thuốc'})

class ThanhToanHoaDonThuocToggle(APIView):
    def get(self, request, format=None):
        id_don_thuoc    = request.GET.get('id', None)
        don_thuoc       = get_object_or_404(DonThuoc, id=id_don_thuoc)
        danh_sach_thuoc = don_thuoc.ke_don.all()
        tong_tien       = request.GET.get('tong_tien', None)
        bao_hiem = False
        
        
        # try:
        for instance in danh_sach_thuoc:    
            id_thuoc = instance.thuoc.id
            so_luong = instance.so_luong
    
            if instance.bao_hiem == True:
                bao_hiem = True

            xuat(request, id=id_thuoc, so_luong=so_luong)

        trang_thai = TrangThaiDonThuoc.objects.get_or_create(trang_thai="Đã Thanh Toán")[0]
        don_thuoc.trang_thai=trang_thai
        don_thuoc.save()

        now             = datetime.now()
        date_time       = now.strftime("%m%d%y%H%M%S")
        ma_hoa_don      = "HDT-" + date_time
        hoa_don_thuoc, created = HoaDonThuoc.objects.get_or_create(don_thuoc=don_thuoc, ma_hoa_don=ma_hoa_don, tong_tien=tong_tien, bao_hiem=bao_hiem, nguoi_thanh_toan=request.user)
        hoa_don_thuoc.save()

        if don_thuoc.benh_nhan is not None:
            channel_layer = get_channel_layer()
            async_to_sync(channel_layer.group_send)(
                f"charge_prescription_user_{don_thuoc.benh_nhan.id}", {
                    'type':'charge_prescription_notification'
                }
            )
        
        from actstream import action
        action.send(request.user, verb='đã thanh toán đơn thuốc', action_object=don_thuoc, target=don_thuoc.benh_nhan)

        response = {'status': 200, 'message': 'Thanh Toán Thành Công'}
        return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")
        # except:
        #     response = {'status': 404, 'message': 'Xảy Ra Lỗi Trong Quá Trình Thanh Toán'}
        #     return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")

class ThanhToanHoaDonDichVuToggle(APIView):
    def get(self, request, format=None):
        ma_hoa_don      = request.GET.get('ma_hoa_don', None)
        hoa_don_dich_vu = get_object_or_404(HoaDonChuoiKham, ma_hoa_don=ma_hoa_don)
        tong_tien       = request.GET.get('tong_tien', None)
        discount        = request.GET.get('discount', None)

        hoa_don_dich_vu.tong_tien = tong_tien
        hoa_don_dich_vu.discount  = discount
        hoa_don_dich_vu.nguoi_thanh_toan = request.user
        
        hoa_don_dich_vu.save()

        # Set trạng thái chuỗi khám
        
        chuoi_kham = hoa_don_dich_vu.chuoi_kham
        lich_hen = chuoi_kham.lich_hen

        trang_thai_lich_hen = TrangThaiLichHen.objects.get_or_create(ten_trang_thai = "Đã Thanh Toán Dịch Vụ")[0]
        if lich_hen.loai_dich_vu == 'kham_theo_yeu_cau':
            trang_thai_chuoi_kham = TrangThaiChuoiKham.objects.get_or_create(trang_thai_chuoi_kham = "Hoàn Thành")[0]    
        else:
            trang_thai_chuoi_kham = TrangThaiChuoiKham.objects.get_or_create(trang_thai_chuoi_kham = "Đã Thanh Toán")[0]

        chuoi_kham.trang_thai = trang_thai_chuoi_kham
        chuoi_kham.save()
        
        # Set trạng thái lịch hẹn
        lich_hen.trang_thai = trang_thai_lich_hen
        lich_hen.save()

        from actstream import action
        action.send(request.user, verb="đã thanh toán chuỗi khám", action_object=chuoi_kham, target=chuoi_kham.benh_nhan)

        channel_layer = get_channel_layer()
        async_to_sync(channel_layer.group_send)(
            f"charge_process_bill_{chuoi_kham.benh_nhan.id}", {
                'type':'charge_process_bill_notification'
            }
        )
        
        response = {
            'status' : 200 ,
            'message': "Thanh Toán Thành Công!"
        }
        return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")

def thanh_toan_hoa_don_thuoc(request):
    id_don_thuoc = request.GET.get('id', None)
    tong_tien = request.GET.get('tong_tien', None)
    print(tong_tien)
    don_thuoc = DonThuoc.objects.get(id = id_don_thuoc)
    danh_sach_thuoc = don_thuoc.ke_don.all()
    now = datetime.now()
    date_time = now.strftime("%m%d%y%H%M%S")
    ma_hoa_don = "HDT-" + date_time
    print(ma_hoa_don)
    
    try:
        for instance in danh_sach_thuoc:    
            id_thuoc = instance.thuoc.id
            so_luong = instance.so_luong
            thuoc = Thuoc.objects.get(id=id_thuoc)
            ten_thuoc = thuoc.ten_thuoc
            if thuoc.kha_dung:
                thuoc.update(so_luong_kha_dung = F('so_luong_kha_dung') - so_luong)
                thuoc.save()
                ThuocLog.objects.create(thuoc=thuoc, ngay=timezone.now(), quy_trinh=ThuocLog.OUT, so_luong=so_luong)
            else:
                response = {'status': 404, 'message': f'Số Lượng Thuốc Không Khả Dụng, Vui Lòng Kiểm Tra Lại Thuốc: {ten_thuoc}'}
                return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")
        
        hoa_don_thuoc = don_thuoc.hoa_don_thuoc
        hoa_don_thuoc.ma_hoa_don = ma_hoa_don
        hoa_don_thuoc.save()
        trang_thai = TrangThaiDonThuoc.objects.get_or_create(trang_thai="Đã Thanh Toán")[0]
        don_thuoc.update(trang_thai=trang_thai)
        don_thuoc.save()
        
        response = {'status': 200, 'message': 'Thanh Toán Thành Công'}
        return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")
    except:
        response = {'status': 404, 'message': 'Xảy Ra Lỗi Trong Quá Trình Thanh Toán'}
        return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")

def get_client_ip(request):
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip

def check_staff_ip(ip_address):
    phong_kham = PhongKham.objects.all().first()
    if phong_kham is not None:
        tinh_trang_phong_kham = phong_kham.tinh_trang
        if tinh_trang_phong_kham is not None:
            ip_range_start = tinh_trang_phong_kham.ip_range_start
            ip_range_end = tinh_trang_phong_kham.ip_range_end

            range_start = ip_range_start.split('.')[-1]
            range_end = ip_range_end.split('.')[-1]
            staff_ip = ip_address.split('.')[-1]

            ip_range_start_3_elements = '.'.join(ip_range_start.split('.')[:-1])
            # ip_range_end_3_elements = '.'.join(ip_range_end.split('.')[:-1])
            staff_ip_3_elements = '.'.join(ip_address.split('.')[:-1])

            if ip_range_start_3_elements == staff_ip_3_elements:
                if int(staff_ip) >= int(range_start) and int(staff_ip) <= int(range_end):
                    return True
                else:
                    return False
            else:
                return False
        else:
            return False
    else:
        return False
    

def loginUser(request):
    from actstream import action
    so_dien_thoai = request.POST.get('so_dien_thoai')
    password = request.POST.get('password')
    
    # staff_ip = get_client_ip(request)
    user = authenticate(request, username=so_dien_thoai, password=password)
    if user is not None:
        if user.is_active and user.is_staff:
            # if check_staff_ip(staff_ip):
            #     auth_login(request, user)
            #     response = {
            #         'status': 200, 
            #     }
            #     return HttpResponse(json.dumps(response), content_type="application/json")
            # else:
            #     response = {
            #         'status': 400,
            #         'message': "Không thể đăng nhập vì bạn không đăng nhập tại phòng khám"
            #     }
            #     return HttpResponse(json.dumps(response), content_type="application/json")
            auth_login(request, user)
            action.send(user, verb="đăng nhập vào hệ thống")
            response = {
                'status': 200,
                'url': "/trang_chu/"
            }
            return HttpResponse(json.dumps(response), content_type="application/json")
        else:
            response = {
                'status': 400,
                'message': "Bạn không phải nhân viên phòng khám"
            }
            return HttpResponse(json.dumps(response), content_type="application/json")
    else:
        response = {
            'status': 400,
            'message': "Thông tin đăng nhập của bạn không đúng, vui lòng kiểm tra lại"
        }
        return HttpResponse(json.dumps(response), content_type="application/json")

def dung_kham(request):
    if request.method == "POST":
        id_lich_hen = request.POST.get('id_lich_hen')
        ly_do = request.POST.get('ly_do')
        lich_hen = LichHenKham.objects.get(id = id_lich_hen)
        trang_thai_lich_hen = TrangThaiLichHen.objects.get_or_create(ten_trang_thai = "Dừng Khám")[0]
        trang_thai_chuoi_kham = TrangThaiChuoiKham.objects.get_or_create(trang_thai_chuoi_kham = "Dừng Khám")[0]
        chuoi_kham = lich_hen.danh_sach_chuoi_kham.all().first()
        lich_su = LichSuChuoiKham.objects.create(chuoi_kham = chuoi_kham, trang_thai = trang_thai_chuoi_kham, chi_tiet_trang_thai = ly_do)
        lich_hen.trang_thai = trang_thai_lich_hen
        lich_hen.save()
        chuoi_kham.trang_thai = trang_thai_chuoi_kham
        chuoi_kham.save() 
        return HttpResponse(json.dumps({
            'status' : 200,
            'message' : "Đã dừng khám",
            # 'url': '/danh_sach_benh_nhan_cho'
        }), content_type="application/json")

def dung_kham_ket_qua_chuyen_khoa(request):
    if request.method == "POST":
        id_chuoi_kham = request.POST.get('id_chuoi_kham')
        ly_do = request.POST.get('ly_do')
        try:
            chuoi_kham = ChuoiKham.objects.get(id = id_chuoi_kham)
            lich_hen = chuoi_kham.lich_hen
            trang_thai_chuoi_kham = TrangThaiChuoiKham.objects.get_or_create(trang_thai_chuoi_kham = "Dừng Khám")[0]
            trang_thai_lich_hen = TrangThaiLichHen.objects.get_or_create(ten_trang_thai = "Dừng Khám")[0]
            lich_su = LichSuChuoiKham.objects.create(chuoi_kham = chuoi_kham, trang_thai = trang_thai_chuoi_kham, chi_tiet_trang_thai = ly_do)
            lich_hen.trang_thai = trang_thai_lich_hen
            lich_hen.save()
            chuoi_kham.trang_thai = trang_thai_chuoi_kham
            chuoi_kham.save()

            response = {
                'status' : 200,
                'message' : "Đã dừng khám",
            }

            from actstream import action
            action.send(request.user, verb='đã dừng khám chuyên khoa cho bệnh nhân', target=chuoi_kham.benh_nhan)

            return HttpResponse(json.dumps(), content_type="application/json")
        except ChuoiKham.DoesNotExist:
            response = {
                'status': 404,
                'message': 'Chuỗi Khám Không Tồn Tại'
            }
            return HttpResponse(json.dumps(response), content_type="application/json")
    else:
        response = {
            'status': 404,
            'message': 'Không gửi được dữ liệu, vui lòng kiểm tra lại'
        }
        return HttpResponse(json.dumps(response), content_type="application/json")



# TODO Sửa lại phần dừng khám của bác sĩ lâm sàng, vì còn liên quan đến phần lịch hẹn

def dung_kham_chuyen_khoa(request):
    if request.method == "POST":
        id_chuoi_kham = request.POST.get('id_chuoi_kham')
        id_phan_khoa = request.POST.get('id_phan_khoa')
        ly_do = request.POST.get('ly_do')

        chuoi_kham = get_object_or_404(ChuoiKham, id=id_chuoi_kham)
        phan_khoa_kham = get_object_or_404(PhanKhoaKham, id=id_phan_khoa)

        trang_thai = TrangThaiChuoiKham.objects.get_or_create(trang_thai_chuoi_kham = "Dừng khám")[0]
        trang_thai_phan_khoa = TrangThaiKhoaKham.objects.get_or_create(trang_thai_khoa_kham = "Dừng Khám")[0]
        chuoi_kham.trang_thai = trang_thai
        chuoi_kham.save()
        phan_khoa_kham.trang_thai = trang_thai_phan_khoa
        phan_khoa_kham.save()
        lich_su_phan_khoa = LichSuTrangThaiKhoaKham.objects.create(phan_khoa_kham=phan_khoa_kham, trang_thai_khoa_kham=trang_thai_phan_khoa, chi_tiet_trang_thai=ly_do)
        lich_su_chuoi_kham = LichSuChuoiKham.objects.create(chuoi_kham=chuoi_kham, trang_thai=trang_thai,chi_tiet_trang_thai=ly_do)
        return HttpResponse(json.dumps({
            'status' : 200,
            'message' : "Đã dừng khám",
            'url': '/phong_chuyen_khoa'
        }), content_type="application/json")

# def xoa_lich_hen(request):
#     if request.method == "POST":
#         id_lich_hen = request.POST.get('id')
#         lich_hen = LichHenKham.objects.get(id=id_lich_hen)
#         lich_hen.delete()
#         return HttpResponse(json.dumps({
#             'status' : 200,
#             'url': '/danh_sach_lich_hen'
#         }), content_type="application/json")

@login_required(login_url='/dang_nhap/')
def update_nguon_cung(request, **kwargs):
    id_cong_ty = kwargs.get('id')
    instance = get_object_or_404(CongTy, id=id_cong_ty)
    form = CongTyForm(request.POST or None, instance=instance)
    phong_chuc_nang = PhongChucNang.objects.all()

    data = {
        'form': form,
        'id_cong_ty': id_cong_ty,
        'phong_chuc_nang' : phong_chuc_nang,
    }
    return render(request, 'phong_tai_chinh/update_nguon_cung.html', context=data)


def chinh_sua_nguon_cung(request):
    if request.method == "POST":
        id_cong_ty = request.POST.get('id_cong_ty')
        ten_cong_ty = request.POST.get('ten_cong_ty')
        dia_chi = request.POST.get('dia_chi')
        giay_phep_kinh_doanh = request.POST.get('giay_phep_kinh_doanh')
        so_lien_lac = request.POST.get('so_lien_lac')
        email = request.POST.get('email')
        mo_ta = request.POST.get('mo_ta')
        cong_ty = get_object_or_404(CongTy, id=id_cong_ty)
        cong_ty.ten_cong_ty = ten_cong_ty
        cong_ty.dia_chi = dia_chi
        cong_ty.giay_phep_kinh_doanh = giay_phep_kinh_doanh
        cong_ty.so_lien_lac = so_lien_lac
        cong_ty.email = email
        cong_ty.mo_ta = mo_ta
        cong_ty.save()

        response = {
            'status': 200,
            'message': 'Cập Nhật Thông Tin Thành Công'
        }
        return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")
    else:
        response = {
            'status': 404,
            'message': 'Cập Nhật Thông Tin Không Thành Công'
        }
        return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")

@login_required(login_url='/dang_nhap/')
def danh_sach_dich_vu_kham(request):
    phong_chuc_nang = PhongChucNang.objects.all()

    return render(request, 'phong_tai_chinh/dich_vu_kham.html', context = {'phong_chuc_nang': phong_chuc_nang})

@login_required(login_url='/dang_nhap/')
def update_dich_vu_kham(request, **kwargs):
    id = kwargs.get('id')
    print(id)
    instance = get_object_or_404(DichVuKham, id=id)
    dich_vu_kham_form = DichVuKhamForm(request.POST or None, instance=instance)
    phong_chuc_nang = PhongChucNang.objects.all()

    data = {
        'dich_vu_kham_form': dich_vu_kham_form,
        'id'               : id,
        'phong_chuc_nang' : phong_chuc_nang
    }
    return render(request, 'phong_tai_chinh/update_dich_vu_kham.html', context=data)

def store_update_dich_vu_kham(request):
    if request.method == 'POST':
        id               = request.POST.get('id')
        ma_dvkt          = request.POST.get('ma_dvkt')
        ten_dvkt         = request.POST.get('ten_dvkt')
        ma_gia           = request.POST.get('ma_gia')
        don_gia          = request.POST.get('don_gia')
        quyet_dinh       = request.POST.get('quyet_dinh')
        cong_bo          = request.POST.get('cong_bo')
        id_phong_chuc_nang = request.POST.get('id_phong_chuc_nang')

        id_phong_chuc_nang = PhongChucNang.objects.get(id=id_phong_chuc_nang)
        
        dich_vu_kham = get_object_or_404(DichVuKham, id=id)
        dich_vu_kham.ma_dvkt          = ma_dvkt
        dich_vu_kham.ten_dvkt         = ten_dvkt
        dich_vu_kham.ma_gia           = ma_gia
        dich_vu_kham.don_gia          = don_gia
        dich_vu_kham.cong_bo          = cong_bo
        dich_vu_kham.quyet_dinh       = quyet_dinh

        dich_vu_kham.id_phong_chuc_nang = id_phong_chuc_nang
        dich_vu_kham.save()
        
        response = {
            'status': 200,
            'message': 'Cập Nhật Thông Tin Thành Công'
        }

        from actstream import action
        action.send(request.user, verb='đã cập nhật thông tin dịch vụ', target=dich_vu_kham)

        return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")
    else:
        response = {
            'status': 404,
            'message': 'Cập Nhật Thông Tin Không Thành Công'
        }
        return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")


def chinh_sua_phong_chuc_nang(request):
    if request.method == "POST":
        id  = request.POST.get('id')
        ten_phong_chuc_nang = request.POST.get('ten_phong_chuc_nang')
        
        phong_chuc_nang = get_object_or_404(PhongChucNang, id=id)
        phong_chuc_nang.ten_phong_chuc_nang = ten_phong_chuc_nang
        phong_chuc_nang.save()

        response = {
            'status': 200,
            'message': 'Cập Nhật Thông Tin Thành Công'
        }
        return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")

@login_required(login_url='/dang_nhap/')
def update_user(request, **kwargs):
    id_user = kwargs.get('id')
    instance = get_object_or_404(User, id=id_user)
    form = UserForm(request.POST or None, instance=instance)
    phong_chuc_nang = PhongChucNang.objects.all()

    data = {
        'form': form,
        'id_user': id_user,
        'phong_chuc_nang' : phong_chuc_nang,
    }
    return render(request, 'update_user.html', context=data)

def cap_nhat_user(request):
    if request.method == "POST":
        id_user       = request.POST.get('id_user')
        ho_ten        = request.POST.get('ho_ten')
        email         = request.POST.get('email')
        so_dien_thoai = request.POST.get('so_dien_thoai')
        cmnd_cccd     = request.POST.get('cmnd')
        user = get_object_or_404(User, id=id_user)
        user.ho_ten        = ho_ten
        user.so_dien_thoai = so_dien_thoai
        user.email         = email
        user.cmnd_cccd     = cmnd_cccd
        user.save()

        response = {
            'status': 200,
            'message': 'Cập Nhật Thông Tin Thành Công'
        }
        return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")

@login_required(login_url='/dang_nhap/')
def chinh_sua_thuoc(request, **kwargs):
    id_thuoc = kwargs.get('id_thuoc')
    instance = get_object_or_404(Thuoc, id=id_thuoc)
    phong_chuc_nang = PhongChucNang.objects.all()
    form = ThuocForm(request.POST or None, instance=instance)
    data = {
        'form': form,
        'id_thuoc': id_thuoc,
        'phong_chuc_nang' : phong_chuc_nang,
    }
    return render(request, 'phong_tai_chinh/update_thuoc.html', context=data)

def update_thuoc(request):
    if request.method == "POST":
        id_thuoc          = request.POST.get('id_thuoc')
        ma_hoat_chat      = request.POST.get('ma_hoat_chat')
        ten_hoat_chat     = request.POST.get('ten_hoat_chat')
        ma_thuoc          = request.POST.get('ma_thuoc')
        ten_thuoc         = request.POST.get('ten_thuoc')
        ham_luong         = request.POST.get('ham_luong')
        duong_dung        = request.POST.get('duong_dung')
        so_dang_ky        = request.POST.get('so_dang_ky')
        dong_goi          = request.POST.get('dong_goi')
        don_vi_tinh       = request.POST.get('don_vi_tinh')
        don_gia           = request.POST.get('don_gia')
        don_gia_tt        = request.POST.get('don_gia_tt')
        so_lo             = request.POST.get('so_lo')
        so_luong_kha_dung = request.POST.get('so_luong_kha_dung')
        hang_sx           = request.POST.get('hang_sx')
        nuoc_sx           = request.POST.get('nuoc_sx')
        quyet_dinh        = request.POST.get('quyet_dinh')
        loai_thuoc        = request.POST.get('loai_thuoc')
        cong_bo           = request.POST.get('cong_bo')
        han_su_dung       = request.POST.get('han_su_dung')
        ngay_san_xuat     = request.POST.get('ngay_san_xuat')
        id_cong_ty        = request.POST.get('id_cong_ty')

        han_su_dung = datetime.strptime(han_su_dung, format_3)
        han_su_dung = han_su_dung.strftime("%Y-%m-%d")

        ngay_san_xuat = datetime.strptime(ngay_san_xuat, format_3)
        ngay_san_xuat = ngay_san_xuat.strftime("%Y-%m-%d")

        cong_ty = get_object_or_404(CongTy, id=id_cong_ty)
        thuoc = get_object_or_404(Thuoc, id = id_thuoc)

        thuoc.ma_hoat_chat      = ma_hoat_chat
        thuoc.ten_hoat_chat     = ten_hoat_chat
        thuoc.ma_thuoc          = ma_thuoc
        thuoc.ten_thuoc         = ten_thuoc
        thuoc.ham_luong         = ham_luong
        thuoc.duong_dung        = duong_dung
        thuoc.so_dang_ky        = so_dang_ky
        thuoc.dong_goi          = dong_goi
        thuoc.don_vi_tinh       = don_vi_tinh
        thuoc.don_gia           = don_gia
        thuoc.don_gia_tt        = don_gia_tt
        thuoc.so_lo             = so_lo
        thuoc.so_luong_kha_dung = so_luong_kha_dung
        thuoc.hang_sx           = hang_sx
        thuoc.nuoc_sx           = nuoc_sx
        thuoc.quyet_dinh        = quyet_dinh
        thuoc.loai_thuoc        = loai_thuoc
        thuoc.cong_bo           = cong_bo
        thuoc.han_su_dung       = han_su_dung
        thuoc.ngay_san_xuat     = ngay_san_xuat
        thuoc.cong_ty           = cong_ty
        thuoc.save()

        from actstream import action
        action.send(request.user, verb='cập nhật thông tin thuốc', target=thuoc)

        response = {
            'status': 200,
            'message': 'Cập Nhật Thông Tin Thành Công'
        }

        return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")
    else:
        response = {
            'status': 404,
            'message': 'Cập Nhật Thông Tin Không Thành Công'
        }
        return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")

@login_required(login_url='/dang_nhap/')
def chinh_sua_thuoc_phong_thuoc(request, **kwargs):
    id_thuoc = kwargs.get('id_thuoc')
    instance = get_object_or_404(Thuoc, id=id_thuoc)
    form = ThuocForm(request.POST or None, instance=instance)
    phong_chuc_nang = PhongChucNang.objects.all()

    data = {
        'form': form,
        'id_thuoc': id_thuoc,
        'phong_chuc_nang' : phong_chuc_nang
    }
    return render(request, 'phong_thuoc/update_thuoc.html', context=data)

def update_thuoc_phong_thuoc(request):
    if request.method == "POST":
        id_thuoc          = request.POST.get('id_thuoc')
        ma_hoat_chat      = request.POST.get('ma_hoat_chat')
        ten_hoat_chat     = request.POST.get('ten_hoat_chat')
        ma_thuoc          = request.POST.get('ma_thuoc')
        ten_thuoc         = request.POST.get('ten_thuoc')
        ham_luong         = request.POST.get('ham_luong')
        duong_dung        = request.POST.get('duong_dung')
        so_dang_ky        = request.POST.get('so_dang_ky')
        dong_goi          = request.POST.get('dong_goi')
        don_vi_tinh       = request.POST.get('don_vi_tinh')
        so_lo             = request.POST.get('so_lo')
        so_luong_kha_dung = request.POST.get('so_luong_kha_dung')
        hang_sx           = request.POST.get('hang_sx')
        nuoc_sx           = request.POST.get('nuoc_sx')
        quyet_dinh        = request.POST.get('quyet_dinh')
        loai_thuoc        = request.POST.get('loai_thuoc')
        cong_bo           = request.POST.get('cong_bo')
        han_su_dung       = request.POST.get('han_su_dung')
        ngay_san_xuat     = request.POST.get('ngay_san_xuat')

        han_su_dung = datetime.strptime(han_su_dung, format_2)
        han_su_dung = han_su_dung.strftime("%Y-%m-%d")

        ngay_san_xuat = datetime.strptime(ngay_san_xuat, format_2)
        ngay_san_xuat = ngay_san_xuat.strftime("%Y-%m-%d")

        thuoc = get_object_or_404(Thuoc, id = id_thuoc)
        thuoc.ma_hoat_chat      = ma_hoat_chat
        thuoc.ten_hoat_chat     = ten_hoat_chat
        thuoc.ma_thuoc          = ma_thuoc
        thuoc.ten_thuoc         = ten_thuoc
        thuoc.ham_luong         = ham_luong
        thuoc.duong_dung        = duong_dung
        thuoc.so_dang_ky        = so_dang_ky
        thuoc.dong_goi          = dong_goi
        thuoc.don_vi_tinh       = don_vi_tinh
        thuoc.so_lo             = so_lo
        thuoc.so_luong_kha_dung = so_luong_kha_dung
        thuoc.hang_sx           = hang_sx
        thuoc.nuoc_sx           = nuoc_sx
        thuoc.quyet_dinh        = quyet_dinh
        thuoc.loai_thuoc        = loai_thuoc
        thuoc.cong_bo           = cong_bo
        thuoc.han_su_dung       = han_su_dung
        thuoc.ngay_san_xuat     = ngay_san_xuat
        thuoc.save()

        response = {
            'status': 200,
            'message': 'Cập Nhật Thông Tin Thành Công'
        }

        return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")
    else:
        response = {
            'status': 404,
            'message': 'Cập Nhật Thông Tin Không Thành Công'
        }
        return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")

@login_required(login_url='/dang_nhap/')    
def doanh_thu_phong_kham(request):
    phong_chuc_nang = PhongChucNang.objects.all()
    
    return render(request, 'phong_tai_chinh/doanh_thu_phong_kham.html', context = {'phong_chuc_nang': phong_chuc_nang})

@login_required(login_url='/dang_nhap/')
def them_dich_vu_kham_excel(request):
    phong_chuc_nang = PhongChucNang.objects.all()

    data = {
        'phong_chuc_nang': phong_chuc_nang,
    }
    return render(request, 'phong_tai_chinh/them_dich_vu_kham_excel.html', context=data)

@login_required(login_url='/dang_nhap/')
def them_dich_vu_kham(request):
    '''Đây là trường hợp "and"'''
    # bac_si = User.objects.filter(Q(chuc_nang = 4) | Q(chuc_nang = 3))
    bac_si = User.objects.filter(chuc_nang = 4)
    phong_chuc_nang = PhongChucNang.objects.all()

    data = {
        'bac_si': bac_si,
        'phong_chuc_nang' : phong_chuc_nang,
    }
    return render(request, 'phong_tai_chinh/them_dich_vu_kham.html', context=data)


from decimal import Decimal
def import_dich_vu_excel(request):
    import pandas as pd
    if request.method == 'POST':
        # data             = request.POST.get('data')  
        # list_objects     = json.loads(data)
        
        file = FilePhongKham.objects.all()
        file = file[0]
        file_url = file.file.url[1:]
        
        excel_data_df = pd.read_excel(file_url, sheet_name='DuLieuMau')
        json_str = excel_data_df.to_json(orient='records')
        json_objects = json.loads(json_str)

        bulk_create_data = []
        print(json_objects)
        for obj in json_objects:
            stt             = obj['STT']
            ma_gia_key      = "MA_GIA"
            ma_cosokcb_key  = "MA_COSOKCB"
            ma_dvkt         = obj['MA_DVKT']
            ten_dvkt        = obj['TEN_DVKT']
            don_gia         = "DON_GIA"

            don_gia_bhyt    = obj['DON_GIA_BHYT']
            don_gia_bhyt = int(don_gia_bhyt)
            don_gia_bhyt    = Decimal(don_gia_bhyt)
            bao_hiem        = True
            quyet_dinh      = 'QUYET_DINH'
            cong_bo         = 'CONG_BO'
            phong_chuc_nang = obj['PHONG_CHUC_NANG']
            # ma_nhom = obj['MA_NHOM']

            group_phong_chuc_nang = PhongChucNang.objects.get_or_create(ten_phong_chuc_nang = phong_chuc_nang)[0]
            # nhom_chi_phi = NhomChiPhi.objects.get(ma_nhom=ma_nhom)

            if ma_gia_key in obj.keys():
                ma_gia = obj[ma_gia_key]
            else:
                ma_gia = ""

            if ma_cosokcb_key in obj.keys():
                ma_cosokcb = obj[ma_cosokcb_key]
            else:
                ma_cosokcb = ""

            if don_gia in obj.keys():
                don_gia = obj[don_gia]
                gia     = Decimal(don_gia)
            else:
                gia=0
            
            if quyet_dinh in obj.keys():
                quyet_dinh = obj[quyet_dinh]
            else:
                quyet_dinh = ''

            if cong_bo in obj.keys():
                cong_bo = obj[cong_bo]
            else:
                cong_bo = ''

            # print(ma_gia)
            model = DichVuKham(
                stt             = stt,
                ma_dvkt         = ma_dvkt,
                ten_dvkt        = ten_dvkt,
                ma_gia          = ma_gia,
                don_gia         = gia,
                don_gia_bhyt = don_gia_bhyt,
                bao_hiem        = bao_hiem,
                quyet_dinh      = quyet_dinh,
                cong_bo         = cong_bo,
                ma_cosokcb      = ma_cosokcb,
                phong_chuc_nang = group_phong_chuc_nang
            )
            bulk_create_data.append(model)
        
        DichVuKham.objects.bulk_update_or_create(bulk_create_data, [
            'stt',
            'ma_dvkt',
            'ten_dvkt',
            'ma_gia',
            'don_gia',
            'don_gia_bhyt',
            'bao_hiem',
            'quyet_dinh',
            'cong_bo',
            'ma_cosokcb',
            'phong_chuc_nang' 
        ], match_field = 'stt', batch_size=10)

        response = {
            'status': 200,
            'message': 'Import Thanh Cong',
            'url' : '/danh_sach_dich_vu_kham/'
        }
        return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")
    else:
        response = {
            'status': 404,
            'message': 'That Bai',
        }
        return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")

@login_required(login_url='/dang_nhap/')      
def them_thuoc_excel(request):
    phong_chuc_nang = PhongChucNang.objects.all()

    return render(request, 'phong_tai_chinh/them_thuoc_excel.html', context = {'phong_chuc_nang': phong_chuc_nang})

def import_thuoc_excel(request):
    import pandas as pd
    if request.method == 'POST':
        # data = request.POST.get('data')
        # list_objects = json.loads(data)

        file = FilePhongKham.objects.all()
        file = file[1]
        file_url = file.file.url[1:]
        
        excel_data_df = pd.read_excel(file_url, sheet_name='DuLieuMau')
        json_str = excel_data_df.to_json(orient='records')
        json_objects = json.loads(json_str)
        
        bulk_create_data = []
        for obj in json_objects:
            stt = obj['STT']
            ma_thuoc_key      = "MA_THUOC_BV"
            ma_hoat_chat_key  = "MA_HOAT_CHAT"
            ma_cskcb_key      = "MA_CSKCB"
            ten_hoat_chat     = obj['HOAT_CHAT']
            duong_dung_key    = "DUONG_DUNG"
            ham_luong_key     = "HAM_LUONG"
            ten_thuoc         = obj['TEN_THUOC']
            so_dang_ky        = obj['SO_DANG_KY']
            dong_goi          = obj['DONG_GOI']
            don_vi_tinh       = obj['DON_VI_TINH']
            don_gia           = Decimal(obj['DON_GIA'])
            don_gia_tt        = Decimal(obj['DON_GIA_TT'])
            so_luong_kha_dung = obj['SO_LUONG']
            hang_sx           = obj['HANG_SX']
            nuoc_sx_key       = "NUOC_SX"
            quyet_dinh        = obj['QUYET_DINH']
            cong_bo           = obj['CONG_BO']
            loai_thuoc        = obj['LOAI_THUOC']
            loai_thau         = obj['LOAI_THAU']
            nhom_thau         = obj['NHOM_THAU']
            nha_thau          = 'NHA_THAU'
            bao_hiem          = True

            # nhom_chi_phi = obj['NHOM_CHI_PHI']

            group_nhom_thau = NhomThau.objects.get_or_create(ten_nhom_thau=nhom_thau)[0]
            group_cong_ty = CongTy.objects.get_or_create(ten_cong_ty=nha_thau)[0]
            # nhom_chi_phi = NhomChiPhi.objects.get(ma_nhom=nhom_chi_phi)
            
            if ma_hoat_chat_key in obj.keys():
                ma_hoat_chat = obj[ma_hoat_chat_key]
            else:
                ma_hoat_chat = ""

            if ma_cskcb_key in obj.keys():
                ma_cskcb = obj[ma_cskcb_key]
            else:
                ma_cskcb = ""

            if nuoc_sx_key in obj.keys():
                nuoc_sx = obj[nuoc_sx_key]
            else: 
                nuoc_sx = ""

            if ma_thuoc_key in obj.keys():
                ma_thuoc = obj[ma_thuoc_key]
            else:
                ma_thuoc = ""

            if duong_dung_key in obj.keys():
                duong_dung = obj[duong_dung_key]
            else:
                duong_dung = ""

            if ham_luong_key in obj.keys():
                ham_luong = obj[ham_luong_key]
            else:
                ham_luong = ""

            if nha_thau in obj.keys():
                nha_thau = obj[nha_thau]
            else:
                nha_thau = ''

            model = Thuoc(
                stt = stt,
                ma_thuoc          = ma_thuoc,
                ma_hoat_chat      = ma_hoat_chat, 
                ten_hoat_chat     = ten_hoat_chat, 
                duong_dung        = duong_dung,
                ham_luong         = ham_luong,
                ten_thuoc         = ten_thuoc,
                so_dang_ky        = so_dang_ky, 
                dong_goi          = dong_goi,
                don_vi_tinh       = don_vi_tinh,
                don_gia           = don_gia,
                don_gia_tt        = don_gia_tt,
                so_lo             = "",
                so_luong_kha_dung = so_luong_kha_dung,
                ma_cskcb          = ma_cskcb, 
                hang_sx           = hang_sx,
                nuoc_sx           = nuoc_sx,
                quyet_dinh        = quyet_dinh, 
                loai_thuoc        = loai_thuoc, 
                cong_bo           = cong_bo,
                loai_thau         = loai_thau,
                nhom_thau         = group_nhom_thau,
                cong_ty           = group_cong_ty,
                bao_hiem          = bao_hiem,
                
            )

            bulk_create_data.append(model)

        Thuoc.objects.bulk_update_or_create(bulk_create_data, [
            'stt',
            'ma_thuoc',
            'ma_hoat_chat', 
            'ten_hoat_chat', 
            'duong_dung', 
            'ham_luong', 
            'ten_thuoc', 
            'so_dang_ky', 
            'dong_goi', 
            'don_vi_tinh', 
            'don_gia', 
            'don_gia_tt',
            'so_luong_kha_dung',
            'ma_cskcb',
            'hang_sx',
            'nuoc_sx',
            'quyet_dinh',
            'cong_bo',
            'loai_thau',
            'nhom_thau', 

        ], match_field = 'stt')
        response = {
            'status': 200,
            'message': 'Import Thanh Cong'
        }
        return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")
        
def store_cong_ty(request):
    if request.method == 'POST':
        ten_cong_ty          = request.POST.get('ten_cong_ty')
        giay_phep_kinh_doanh = request.POST.get('giay_phep_kinh_doanh')
        so_lien_lac          = request.POST.get('so_lien_lac')
        email                = request.POST.get('email')
        dia_chi              = request.POST.get('dia_chi')

        CongTy.objects.get_or_create(
            ten_cong_ty          = ten_cong_ty,
            giay_phep_kinh_doanh = giay_phep_kinh_doanh,
            so_lien_lac          = so_lien_lac,
            email                = email,
            dia_chi              = dia_chi,
        )[0]

        response = {
            'status': 200,
            "message" : "Thêm nguồn cung thành công",
        }
        return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")
    else:
        response = {
            'status': 404,
            "message" : "Thêm nguồn cung thất bại",
        }
        return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")

@login_required(login_url='/dang_nhap/')
def danh_sach_bai_dang(request):
    phong_chuc_nang = PhongChucNang.objects.all()

    return render(request, "le_tan/danh_sach_bai_dang.html", context={'phong_chuc_nang':phong_chuc_nang})

def store_thanh_toan_lam_sang(request):
    if request.method == 'POST':
        id_lich_hen = request.POST.get('id')
        gia_tien = request.POST.get('gia_tien')

        lich_hen = get_object_or_404(LichHenKham, id=id_lich_hen)

        hoa_don_lam_sang = HoaDonLamSang.objects.create(
            tong_tien = gia_tien, 
            lich_hen = lich_hen,
            nguoi_thanh_toan = request.user,
        )

        hoa_don_lam_sang.save()

        trang_thai = TrangThaiLichHen.objects.get_or_create(ten_trang_thai = "Đã Thanh Toán Lâm Sàng")[0]
        lich_hen.trang_thai = trang_thai
        lich_hen.save()

        channel_layer = get_channel_layer()
        async_to_sync(channel_layer.group_send)(
            f"charge_bill_user_{lich_hen.benh_nhan.id}", {
                'type':'charge_bill_notification'
            }
        )

        response = {
            'status': 200,
            'message': 'Thanh Toán Lâm Sàng Thành Công'
        }

        from actstream import action
        action.send(request.user, verb='đã thanh toán lâm sàng', target=lich_hen.benh_nhan)
        # return redirect('/danh_sach_benh_nhan/')
        return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")

    else:
        response = {
            'status': 400,
            'message': 'Thanh Toán Lâm Sàng Không Thành Công'
        }
        return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")


@login_required(login_url='/dang_nhap/')
def danh_sach_phong_chuc_nang(request):
    phong_chuc_nang = PhongChucNang.objects.all()
    bac_si_phu_trach = User.objects.filter(chuc_nang = 4)
    data = {
        'bac_si_phu_trach' : bac_si_phu_trach,
        'phong_chuc_nang' : phong_chuc_nang,
    }
    return render(request, 'le_tan/danh_sach_phong_chuc_nang.html', context=data)

@login_required(login_url='/dang_nhap/')
def them_phong_chuc_nang(request):
    bac_si_phu_trach = User.objects.filter(chuc_nang=4)
    phong_chuc_nang = PhongChucNang.objects.all()

    data = {
        'bac_si_phu_trach' : bac_si_phu_trach,
        'phong_chuc_nang' : phong_chuc_nang
    }
    return render(request, 'le_tan/them_phong_chuc_nang.html', context=data)


def them_pcn_kem_dich_vu(request):
    if request.method == "POST":
        ten_phong_chuc_nang = request.POST.get("ten_phong_chuc_nang")
        # id_bac_si           = request.POST.get('bac_si_phu_trach') # phan nay la phan id bac si m gui len o template, dat ten nao thi tuy
        request_data        = request.POST.get('data')
        data                = json.loads(request_data)

        phong_chuc_nang = PhongChucNang.objects.create(ten_phong_chuc_nang=ten_phong_chuc_nang)

        content_type = ContentType.objects.get_for_model(PhongChucNang)

        new_group, created = Group.objects.get_or_create(name=f"Nhóm {phong_chuc_nang}")
        codename_perm = f'can_view_{phong_chuc_nang.slug}'

        if not Permission.objects.filter(codename = codename_perm).exists():
            permission = Permission.objects.create(
                codename=codename_perm,
                name = f'Xem {phong_chuc_nang}',
                content_type=content_type
            )
            new_group.permissions.add(permission)

        if data:
            list_id_dich_vu_kham = []
    
            for obj in data:
                id = obj['obj']['id']
                list_id_dich_vu_kham.append(id) # lay tat ca id cua dich vu kham va append vao list
    
            danh_sach_dich_vu = DichVuKham.objects.filter(id__in=list_id_dich_vu_kham) # filter tat ca dich vu co id nam trong               list_id_dich_vu_kham

            for dich_vu in danh_sach_dich_vu:
                dich_vu.phong_chuc_nang = phong_chuc_nang # cap nhat field phong_chuc_nang cho tung dich_vu
            DichVuKham.objects.bulk_update(danh_sach_dich_vu, ['phong_chuc_nang']) # update tat ca cac ban ghi o tren bang 1 query
        
        response = {
            'status': 200,
            'message': 'Thêm Phòng Chức Năng Thành Công'
        }
        return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")

    else:
        response = {
            'status': 404,
            'message': 'Thêm Phòng Chức Năng Không Thành Công'
        }
        return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")

@login_required(login_url='/dang_nhap/')
def update_phong_chuc_nang(request, **kwargs):
    id_phong_chuc_nang = kwargs.get('id')
    instance = get_object_or_404(PhongChucNang, id=id_phong_chuc_nang)
    phong_chuc_nang = PhongChucNang.objects.all()
    form = PhongChucNangForm(request.POST or None, instance=instance)
    data = {
        'form': form,
        'id_phong_chuc_nang': id_phong_chuc_nang,
        'phong_chuc_nang': phong_chuc_nang
    }
    return render(request, 'le_tan/update_phong_chuc_nang.html', context=data)

@login_required(login_url='/dang_nhap/')
def them_bai_dang(request):
    phong_chuc_nang = PhongChucNang.objects.all()

    return render(request, 'le_tan/them_bai_dang.html',context ={'phong_chuc_nang' : phong_chuc_nang})

def upload_bai_dang(request):
    print(request.FILES)
    if request.method == "POST":
        tieu_de            = request.POST.get('tieu_de', None)
        noi_dung_chinh     = request.POST.get('noi_dung_chinh', None)
        noi_dung           = request.POST.get('noi_dung', None)
        thoi_gian_bat_dau  = request.POST.get('thoi_gian_bat_dau', None)
        thoi_gian_ket_thuc = request.POST.get('thoi_gian_ket_thuc')

        thoi_gian_bat_dau = datetime.strptime(thoi_gian_bat_dau, format_2)
        thoi_gian_bat_dau = thoi_gian_bat_dau.strftime("%Y-%m-%d %H:%M")

        thoi_gian_ket_thuc = datetime.strptime(thoi_gian_ket_thuc, format_2)
        thoi_gian_ket_thuc = thoi_gian_ket_thuc.strftime("%Y-%m-%d %H:%M")
        # print(id_chuoi_kham)

        if tieu_de == '':
            HttpResponse({'status': 404, 'message': 'Tiêu đề không được để trống'})

        if noi_dung_chinh == '':
            HttpResponse({'status': 404, 'message': 'Nội dung chính không được để trống'})

        # chuoi_kham = ChuoiKham.objects.get(id=id_chuoi_kham)
        # ket_qua_tong_quat = KetQuaTongQuat.objects.get_or_create(chuoi_kham=chuoi_kham)[0]
        # ket_qua_chuyen_khoa = KetQuaChuyenKhoa.objects.create(ket_qua_tong_quat=ket_qua_tong_quat, ma_ket_qua=ma_ket_qua, mo_ta=mo_ta, ket_luan=ket_luan)
        
        for value in request.FILES.values():
            
            bai_dang = BaiDang.objects.create(
                tieu_de            = tieu_de,
                noi_dung_chinh     = noi_dung_chinh,
                noi_dung           = noi_dung,
                thoi_gian_bat_dau  = thoi_gian_bat_dau,
                thoi_gian_ket_thuc = thoi_gian_ket_thuc,
                hinh_anh           = value,
                nguoi_dang_bai     = request.user,
            )
            bai_dang.save()

        return HttpResponse('upload')
    response = {
        'status': 200,
        'message' : 'Upload Thành Công!'
    }
    return HttpResponse(json.dumps(response), content_type='application/json; charset=utf-8')

@login_required(login_url='/dang_nhap/')
def chi_tiet_bai_dang(request, **kwargs):
    id_bai_dang = kwargs.get('id')
    bai_dang = BaiDang.objects.get(id = id_bai_dang)
    phong_chuc_nang = PhongChucNang.objects.all()

    data = {
        'bai_dang': bai_dang,
        'phong_chuc_nang' : phong_chuc_nang,
    }
    return render(request, 'le_tan/bai_dang.html', context=data)

@login_required(login_url='/dang_nhap/')
def update_don_thuoc(request, **kwargs):
    id_don_thuoc = kwargs.get('id')
    don_thuoc = DonThuoc.objects.get(id = id_don_thuoc)
    danh_sach_thuoc = don_thuoc.ke_don.all()
    phong_chuc_nang = PhongChucNang.objects.all()

    data = {
        'danh_sach_thuoc': danh_sach_thuoc,
        'don_thuoc' : don_thuoc,
        'phong_chuc_nang' : phong_chuc_nang,
    }
    return render(request, 'bac_si_lam_sang/update_don_thuoc.html', context=data)

@login_required(login_url='/dang_nhap/')
def danh_sach_vat_tu(request):
    phong_chuc_nang = PhongChucNang.objects.all()

    return render(request, 'phong_tai_chinh/danh_sach_vat_tu.html', context={'phong_chuc_nang': phong_chuc_nang})

@login_required(login_url='/dang_nhap/')
def them_vat_tu_excel(request):
    phong_chuc_nang = PhongChucNang.objects.all()

    return render(request, 'phong_tai_chinh/them_vat_tu_excel.html', context={'phong_chuc_nang':phong_chuc_nang})

# UPDATE
def import_vat_tu_excel(request):
    if request.method == 'POST':
        data = request.POST.get('data')
        list_objects = json.loads(data)
        bulk_create_data = []

        for obj in list_objects:
            stt            = obj['MA']
            ma_nhom_vtyt   = obj['MA_NHOM_VTYT']
            ten_nhom_vtyt  = obj['TEN_NHOM_VTYT']
            ma_hieu        = obj['MA_HIEU']
            ma_vtyt_bv     = obj['MA_VTYT_BV']
            ten_vtyt_bv    = obj['TEN_VTYT_BV']
            quy_cach       = obj['QUY_CACH']
            hang_sx        = obj['HANG_SX']
            nuoc_sx        = obj['NUOC_SX']
            don_vi_tinh    = obj['DON_VI_TINH']
            don_gia        = Decimal(obj['DON_GIA'])
            don_gia_tt     = Decimal(obj['DON_GIA_TT'])
            nha_thau       = obj['NHA_THAU']
            quyet_dinh     = obj['QUYET_DINH']
            cong_bo        = obj['CONG_BO']
            dinh_muc       = obj['DINH_MUC']
            so_luong       = obj['SO_LUONG']
            loai_thau      = obj['LOAI_THAU']
            bao_hiem       = True

            group_cong_ty = CongTy.objects.get_or_create(ten_cong_ty=nha_thau)[0]
            group_nhom_vat_tu = NhomVatTu.objects.get_or_create(ma_nhom_vtyt = ma_nhom_vtyt, ten_nhom_vtyt = ten_nhom_vtyt)[0]
            
            model = VatTu(
                stt               = stt,
                nhom_vat_tu       = group_nhom_vat_tu,
                ma_hieu           = ma_hieu,
                ma_vtyt_bv        = ma_vtyt_bv,
                ten_vtyt_bv       = ten_vtyt_bv,
                quy_cach          = quy_cach,
                hang_sx           = hang_sx,
                nuoc_sx           = nuoc_sx,
                don_vi_tinh       = don_vi_tinh,
                don_gia           = don_gia,
                don_gia_tt        = don_gia_tt,
                nha_thau          = group_cong_ty,
                quyet_dinh        = quyet_dinh,
                cong_bo           = cong_bo,
                dinh_muc          = dinh_muc,
                so_luong_kha_dung = so_luong,
                loai_thau         = loai_thau,
                bao_hiem          = bao_hiem,
            )

            bulk_create_data.append(model)

        VatTu.objects.bulk_update_or_create(bulk_create_data, [
            'stt',
            'nhom_vat_tu',
            'ma_hieu',
            'ma_vtyt_bv',
            'ten_vtyt_bv',
            'quy_cach',
            'hang_sx',
            'nuoc_sx',
            'don_vi_tinh',
            'don_gia',
            'don_gia_tt',
            'nha_thau',
            'quyet_dinh',
            'cong_bo',
            'dinh_muc',
            'so_luong_kha_dung',
            'loai_thau',
            'bao_hiem',
        ], match_field = 'stt')

        response = {
            'status': 200,
            'message': 'Import Thanh Cong'
        }
        return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")
# END

def nhan_don_thuoc(request):
    if request.method == 'POST':
        id = request.POST.get('id')

        trang_thai_don_thuoc = TrangThaiDonThuoc.objects.get_or_create(trang_thai = "Hoàn Thành")[0]
        don_thuoc = get_object_or_404(DonThuoc, id=id)
        if don_thuoc.chuoi_kham is not None:
            chuoi_kham = don_thuoc.chuoi_kham
            lich_hen = chuoi_kham.lich_hen
            trang_thai_chuoi_kham = TrangThaiChuoiKham.objects.get_or_create(trang_thai_chuoi_kham = "Hoàn Thành")[0]
            trang_thai_lich_hen = TrangThaiLichHen.objects.get_or_create(ten_trang_thai = "Hoàn Thành")[0]
            chuoi_kham.trang_thai = trang_thai_chuoi_kham
            lich_hen.trang_thai = trang_thai_lich_hen
            lich_hen.save()
            chuoi_kham.save()
            don_thuoc.trang_thai = trang_thai_don_thuoc
            don_thuoc.save()
            
            channel_layer = get_channel_layer()
            async_to_sync(channel_layer.group_send)(
                f"process_accomplished_user_{don_thuoc.benh_nhan.id}", {
                    'type':'process_accomplished_notification'
                }
            )
        else:
            don_thuoc.trang_thai = trang_thai_don_thuoc
            don_thuoc.save()
            response={
                'status' : 200,
                'message' : 'Đã Nhận Đơn Thuốc'
            }
            return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")

        response={
            'status' : 200,
            'message' : 'Đã Nhận Đơn Thuốc'
        }
        return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")
        
def xoa_thuoc(request):
    if request.method == 'POST':
        id = request.POST.get('id')
        thuoc = Thuoc.objects.get(id = id)
        thuoc.delete()

        response = {
            'thuoc' : thuoc.ten_thuoc
        }
        return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")

def xoa_dich_vu(request):
    if request.method == 'POST':
        id = request.POST.get('id')
        DichVuKham.objects.get(id=id).delete()

        response = {
            'message' : "Xóa Thành Công"
        }
        return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")

def xoa_vat_tu(request):
    if request.method == 'POST':
        id = request.POST.get('id')
        VatTu.objects.get(id=id).delete()

        response = {
            'message' : "Xóa Thành Công"
        }
        return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")

@login_required(login_url='/dang_nhap/')
def danh_sach_bac_si(request):
    phong_chuc_nang = PhongChucNang.objects.all()
    province = Province.objects.all()
    data={
        'phong_chuc_nang' : phong_chuc_nang,
        'province' : province,
    }

    return render(request, 'danh_sach_bac_si.html', context=data)

def create_bac_si(request):
    if request.method == "POST":
        ho_ten         = request.POST.get("ho_ten",         None)
        so_dien_thoai  = request.POST.get("so_dien_thoai",  None)
        password       = request.POST.get("password",       None)
        username       = request.POST.get('username', None)
        cmnd_cccd      = request.POST.get("cmnd_cccd",      None)
        dia_chi        = request.POST.get("dia_chi",        None)
        ngay_sinh      = request.POST.get("ngay_sinh",      None)
        gioi_tinh      = request.POST.get("gioi_tinh",      None)
        dan_toc        = request.POST.get("dan_toc",        None)
        ma_so_bao_hiem = request.POST.get("ma_so_bao_hiem", None)
        gioi_thieu     = request.POST.get('gioi_thieu',     None)
        chuc_danh      = request.POST.get('chuc_danh',      None)
        chuyen_khoa    = request.POST.get('chuyen_khoa',    None)
        noi_cong_tac   = request.POST.get('noi_cong_tac',   None)
        kinh_nghiem    = request.POST.get('kinh_nghiem',    None)
        loai_cong_viec = request.POST.get('loai_cong_viec', None)
        chuc_nang      = request.POST.get('chuc_nang',      None)
        tinh_id      = request.POST.get('tinh',      None)
        huyen_id      = request.POST.get('huyen',      None)
        xa_id      = request.POST.get('xa',      None)

        ngay_sinh = datetime.strptime(ngay_sinh, format_3)
        ngay_sinh = ngay_sinh.strftime("%Y-%m-%d")
        
        if len(ho_ten) == 0:
            return HttpResponse(json.dumps({'message': "Họ Tên Không Được Trống", 'status': '400'}), content_type='application/json; charset=utf-8')

        if User.objects.filter(so_dien_thoai=so_dien_thoai).exists():
            return HttpResponse(json.dumps({'message': "Số Điện Thoại Đã Tồn Tại", 'status': '409'}), content_type='application/json; charset=utf-8')

        if User.objects.filter(cmnd_cccd=cmnd_cccd).exists():
            return HttpResponse(json.dumps({'message': "Số chứng minh thư đã tồn tại", 'status': '403'}), content_type = 'application/json; charset=utf-8')

        if dia_chi == '':
            response = {
                'status': 400,
                'message': "Thiếu địa chỉ cụ thể"
            }
            return HttpResponse(
                json.dumps(response),
                content_type="application/json"
            )

        user = User.objects.create_nguoi_dung(
            ho_ten         = ho_ten, 
            so_dien_thoai  = so_dien_thoai,
            cmnd_cccd      = cmnd_cccd,
            dia_chi        = dia_chi,
            gioi_tinh      = gioi_tinh,
            dan_toc        = dan_toc,
            ngay_sinh = ngay_sinh,    
            ma_so_bao_hiem = ma_so_bao_hiem,
            password       = password,
        )

        user.chuc_nang = chuc_nang

        if username != '':
            user.username = username
        
        user.staff = True

        if tinh_id != 'null':      
            tinh = Province.objects.filter(id=tinh_id).first()
            user.tinh = tinh
        else:
            response = {
                'status': 400,
                'message': "Không thể thiếu tỉnh thành"
            }
            return HttpResponse(
                json.dumps(response),
                content_type="application/json"
            )
    
        if huyen_id != 'null':
            huyen = District.objects.filter(id=huyen_id).first()
            user.huyen = huyen
        else:
            response = {
                'status': 400,
                'message': "Không thể thiếu quận huyện"
            }
            return HttpResponse(
                json.dumps(response),
                content_type="application/json"
            )

        if xa_id != 'null': 
            xa = Ward.objects.filter(id=xa_id).first()  
            user.xa = xa
        else:
            response = {
                'status': 400,
                'message': "Không thể thiếu phường xã"
            }
            return HttpResponse(
                json.dumps(response),
                content_type="application/json"
            )
   
        user.save()

        bac_si = BacSi.objects.create(
            user           = user,
            gioi_thieu     = gioi_thieu    ,
            chuc_danh      = chuc_danh     ,
            chuyen_khoa    = chuyen_khoa   ,
            noi_cong_tac   = noi_cong_tac  ,
            kinh_nghiem    = kinh_nghiem   ,
            loai_cong_viec = loai_cong_viec,
        )

        bac_si.save()

        response = {
            "message"      : "Thêm Mới Bác Sĩ Thành Công",
            "ho_ten"       : bac_si.user.ho_ten,
            "so_dien_thoai": bac_si.user.so_dien_thoai,
        }

        return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")
    else:
        return HttpResponse(
            json.dumps({"nothing to see": "this isn't happening"}),
            content_type="application/json"
        )

@login_required(login_url='/dang_nhap/')
def update_bac_si(request, **kwargs):
    id_bac_si = kwargs.get('id')
    instance = get_object_or_404(BacSi, id=id_bac_si)
    form = BacSiForm(request.POST or None, instance=instance)
    
    user_id = kwargs.get('user_id')
    instance_user = get_object_or_404(User, id=user_id)
    form_user = UserForm(request.POST or None, instance=instance_user)
    phong_chuc_nang = PhongChucNang.objects.all()
    data = {
        'id_bac_si': id_bac_si,
        'form': form,
        'user_id': user_id,
        'form_user': form_user,
        'phong_chuc_nang': phong_chuc_nang,
    }
    return render(request, 'update_bac_si.html', context=data)

def cap_nhat_thong_tin_bac_si(request):
    if request.method == "POST":
        id   = request.POST.get('id')
        user_id        = request.POST.get('user_id')
        ho_ten         = request.POST.get('ho_ten')
        so_dien_thoai  = request.POST.get('so_dien_thoai')
        cmnd_cccd      = request.POST.get('cmnd_cccd')
        ngay_sinh      = request.POST.get('ngay_sinh')
        gioi_thieu     = request.POST.get('gioi_thieu')
        noi_cong_tac   = request.POST.get('noi_cong_tac')
        chuyen_khoa    = request.POST.get('chuyen_khoa')
        chuc_danh      = request.POST.get('chuc_danh')
        loai_cong_viec = request.POST.get('loai_cong_viec')
        kinh_nghiem    = request.POST.get('kinh_nghiem')
        dia_chi        = request.POST.get('dia_chi')

        ngay_sinh = datetime.strptime(ngay_sinh, format_3)
        ngay_sinh = ngay_sinh.strftime("%Y-%m-%d")

        benh_nhan = get_object_or_404(User, id=user_id)
        benh_nhan.ho_ten         = ho_ten
        benh_nhan.so_dien_thoai  = so_dien_thoai
        benh_nhan.cmnd_cccd      = cmnd_cccd
        benh_nhan.dia_chi        = dia_chi
        benh_nhan.ngay_sinh      = ngay_sinh
        benh_nhan.save()

        bac_si = get_object_or_404(BacSi, id=id)
        bac_si.gioi_thieu     = gioi_thieu    
        bac_si.noi_cong_tac   = noi_cong_tac  
        bac_si.chuyen_khoa    = chuyen_khoa   
        bac_si.chuc_danh      = chuc_danh     
        bac_si.loai_cong_viec = loai_cong_viec
        bac_si.kinh_nghiem    = kinh_nghiem   
        bac_si.save()
        
        from actstream import action
        action.send(request.user, verb='đã cập nhật thông tin bác sĩ', target=bac_si)

        response = {
            'status': 200,
            'message': 'Cập Nhật Thông Tin Thành Công'
        }
        return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")
    else:
        response = {
            'status': 404,
            'message': 'Cập Nhật Thông Tin Không Thành Công'
        }
        return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")

def xoa_lich_hen(request):
    if request.method == 'POST':
        id = request.POST.get('id')
        LichHenKham.objects.get(id = id).delete()

        response = {
            'message' : "Xóa Lịch Hẹn Thành Công!"
        }

        return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")

@login_required(login_url='/dang_nhap/')  
def xuat_bao_hiem(request):
    phong_chuc_nang = PhongChucNang.objects.all()

    data ={
        'phong_chuc_nang' : phong_chuc_nang,
    }

    return render(request, 'phong_tai_chinh/xuat_bao_hiem.html', context = data)

@login_required(login_url='/dang_nhap/')  
def danh_sach_benh_nhan_bao_hiem(request):
    phong_chuc_nang = PhongChucNang.objects.all()

    data ={
        'phong_chuc_nang' : phong_chuc_nang,
    }
    return render(request, 'phong_tai_chinh/danh_sach_benh_nhan_bao_hiem.html', context = data)

@login_required(login_url='/dang_nhap/')
def danh_sach_dich_vu_bao_hiem(request):
    phong_chuc_nang = PhongChucNang.objects.all()

    data ={
        'phong_chuc_nang' : phong_chuc_nang,
    }
    return render(request, 'phong_tai_chinh/danh_sach_dich_vu_bao_hiem.html', context = data)

@login_required(login_url='/dang_nhap/')
def danh_sach_thuoc_bao_hiem(request):
    phong_chuc_nang = PhongChucNang.objects.all()

    data ={
        'phong_chuc_nang' : phong_chuc_nang,
    }
    return render(request, 'phong_tai_chinh/danh_sach_thuoc_bao_hiem.html', context = data)

def upload_ket_qua_lam_sang(request):
    if request.method == "POST":
        ma_ket_qua    = request.POST.get('ma_ket_qua')
        mo_ta         = request.POST.get('mo_ta')
        id_chuoi_kham = request.POST.get('id_chuoi_kham')
        noi_dung = request.POST.get('noi_dung')

        if ma_ket_qua == '':
            HttpResponse({'status': 404, 'message': 'Mã Kết Quả Không Được Để Trống'})

        if mo_ta == '':
            HttpResponse({'status': 404, 'message': 'Mô Tả Không Được Để Trống'})

        try:
            chuoi_kham = ChuoiKham.objects.get(id=id_chuoi_kham)
            
            lich_hen = chuoi_kham.lich_hen
            if lich_hen.loai_dich_vu == 'kham_theo_yeu_cau':
                trang_thai_chuoi_kham = TrangThaiChuoiKham.objects.get(trang_thai_chuoi_kham='Hoàn Thành')
                lich_hen.thanh_toan_sau = False
                chuoi_kham.trang_thai = trang_thai_chuoi_kham
                chuoi_kham.save()
                lich_hen.save()

            ket_qua_tong_quat = KetQuaTongQuat.objects.get_or_create(chuoi_kham=chuoi_kham)[0]
            ket_qua_tong_quat.ma_ket_qua = ma_ket_qua
            ket_qua_tong_quat.mo_ta      = mo_ta
            ket_qua_tong_quat.save()

            HtmlKetQua.objects.create(
                ket_qua_tong_quat = ket_qua_tong_quat,
                noi_dung = noi_dung,
            )

            from actstream import action
            action.send(request.user, verb='tải lên kết quả tổng quát', action_object=ket_qua_tong_quat, target=chuoi_kham.benh_nhan)

        except ChuoiKham.DoesNotExist:
            response = {
                'status': 404,
                'message' : "Upload Kết Quả Thành Công!"
            }
            return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")

        response = {
            'status': 200,
            'message' : "Upload Kết Quả Thành Công!"
        }

        return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")
    else:
        response = {
            'status': 404,
            'message' : "Không Thể Gửi Lên Dữ Liệu, Vui Lòng Kiểm Tra Lại"
        }

        return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")


def upload_ket_qua_chuyen_khoa(request):
    if request.method == "POST":
        ma_ket_qua    = request.POST.get('ma_ket_qua', None)
        mo_ta         = request.POST.get('mo_ta', None)
        ket_luan      = request.POST.get('ket_qua', None)
        id_chuoi_kham = request.POST.get('id_chuoi_kham')
        id_phan_khoa = request.POST.get('id_phan_khoa')

        if ma_ket_qua == '':
            HttpResponse({'status': 404, 'message': 'Mã Kết Quả Không Được Để Trống'})

        if mo_ta == '':
            HttpResponse({'status': 404, 'message': 'Mô Tả Không Được Để Trống'})

        if ket_luan == '':
            HttpResponse({'status': 404, 'message': 'Kết Luận Không Được Để Trống'})

        chuoi_kham = ChuoiKham.objects.filter(id=id_chuoi_kham).first()
        trang_thai = TrangThaiKhoaKham.objects.get_or_create(trang_thai_khoa_kham='Đã Tải Lên Kết Quả')[0]
        phan_khoa = PhanKhoaKham.objects.filter(id=id_phan_khoa).first()
        phan_khoa.trang_thai = trang_thai
        phan_khoa.save()
        ket_qua_tong_quat = KetQuaTongQuat.objects.get_or_create(chuoi_kham=chuoi_kham)[0]
        ket_qua_chuyen_khoa = KetQuaChuyenKhoa.objects.create(
            ket_qua_tong_quat=ket_qua_tong_quat, 
            ma_ket_qua=ma_ket_qua, mo_ta=mo_ta, 
            ket_luan=ket_luan,
            phan_khoa_kham=phan_khoa,
            bac_si_chuyen_khoa=request.user,
        )

        from actstream import action
        action.send(request.user, verb="tải lên kết quả", action_object=ket_qua_chuyen_khoa, target=chuoi_kham.benh_nhan)

        response = {
            'status': 200,
            'message' : 'Upload Thành Công!'
        }

        return HttpResponse(json.dumps(response), content_type='application/json; charset=utf-8')


# * --- update 6/1/2021 ---
@login_required(login_url='/dang_nhap/')
def thong_ke_vat_tu(request):
    phong_chuc_nang = PhongChucNang.objects.all()
    user_id = request.user.id

    data = {
        'phong_chuc_nang': phong_chuc_nang,
        'user_id'        : user_id,
    }
    return render(request, 'phong_tai_chinh/thong_ke_vat_tu_cuoi_thang.html', context=data)

def store_thong_ke_vat_tu(request):
    if request.method == "POST":
        request_data = request.POST.get('data', None)
        user = request.user
        data = json.loads(request_data)
        print(data)

        list_thanh_tien = []
        bulk_create_data = []

        for i in data:
            thanh_tien = int(i['obj']['thanh_tien'])
            list_thanh_tien.append(thanh_tien)
            
        tong_tien = int(sum(list_thanh_tien))
        hoa_don_vat_tu = HoaDonVatTu.objects.get_or_create(nguoi_phu_trach = user, tong_tien = tong_tien)[0]
        # hoa_don_vat_tu = HoaDonVatTu.objects.get_or_create
        hoa_don_vat_tu.save()
        for i in data:
            vat_tu = VatTu.objects.only('id').get(id=i['obj']['id'])    
            ke_don_thuoc = KeVatTu(vat_tu=vat_tu, so_luong=i['obj']['so_luong'], bao_hiem=i['obj']['bao_hiem'], hoa_don_vat_tu = hoa_don_vat_tu)
            bulk_create_data.append(ke_don_thuoc)

        KeVatTu.objects.bulk_create(bulk_create_data)
        response = {'status': 200, 'message': 'Thống Kê Hoàn Tất'}
        from actstream import action
        action.send(request.user, verb='đã thực hiện thống kê vật tư')

    else:
        response = {'status': 404, 'message' : "Có Lỗi Xảy Ra"}
    return HttpResponse(json.dumps(response), content_type='application/json; charset=utf-8')

@login_required(login_url='/dang_nhap/')
def them_vat_tu(request):
    phong_chuc_nang = PhongChucNang.objects.all()
    cong_ty = CongTy.objects.filter(loai_cung = "vat_tu")
    nhom_vat_tu = NhomVatTu.objects.all()
    data = {
        'phong_chuc_nang': phong_chuc_nang,
        'cong_ty'        : cong_ty,
        'nhom_vat_tu'    : nhom_vat_tu,
    }
    return render(request, 'phong_tai_chinh/them_moi_vat_tu.html', context=data)

def create_vat_tu(request):
    if request.method == "POST":
        id_cong_ty        = request.POST.get('id_cong_ty')
        ma_hieu           = request.POST.get('ma_hieu')
        ma_vtyt_bv        = request.POST.get('ma_vtyt_bv')
        quy_cach          = request.POST.get('quy_cach')
        hang_sx           = request.POST.get("hang_san_xuat")
        nuoc_sx           = request.POST.get("nuoc_san_xuat")
        don_vi_tinh       = request.POST.get('don_vi_tinh')
        don_gia           = request.POST.get("don_gia")
        don_gia_tt        = request.POST.get("don_gia_tt")
        id_nhom_vat_tu    = request.POST.get('id_nhom_vat_tu')
        so_luong_kha_dung = request.POST.get("so_luong_kha_dung")
        quyet_dinh        = request.POST.get("quyet_dinh")
        cong_bo           = request.POST.get("cong_bo") 
        ten_vtyt_bv       = request.POST.get('ten_vtyt_bv')   
    
        cong_ty = CongTy.objects.get(id=id_cong_ty)

        nhom_vat_tu = NhomVatTu.objects.get(id = id_nhom_vat_tu)

        vat_tu = VatTu.objects.create(
            nha_thau           = cong_ty,
            don_vi_tinh        = don_vi_tinh,
            don_gia            = don_gia,
            don_gia_tt         = don_gia_tt,
            so_luong_kha_dung  = so_luong_kha_dung,
            hang_sx            = hang_sx,
            nuoc_sx            = nuoc_sx,
            quyet_dinh         = quyet_dinh,
            cong_bo            = cong_bo,
            ma_hieu            = ma_hieu,
            ma_vtyt_bv         = ma_vtyt_bv,
            quy_cach           = quy_cach,
            nhom_vat_tu        = nhom_vat_tu,
            ten_vtyt_bv        = ten_vtyt_bv,
        )

        from actstream import action
        action.send(request.user, verb='đã tạo mới vật tư', target=vat_tu)
        response = {
            'status' : 200,
            'message' : 'Tạo Thành Công'
        }
        return HttpResponse(json.dumps(response), content_type='application/json; charset=utf-8')

@login_required(login_url='/dang_nhap/')
def update_phong_kham(request, **kwargs):
    id = kwargs.get('id')
    instance = get_object_or_404(PhongKham, id=id)
    form = PhongKhamForm(request.POST or None, instance=instance)
    phong_chuc_nang = PhongChucNang.objects.all()

    data = {
        'form': form,
        'id': id,
        'phong_chuc_nang' : phong_chuc_nang,
    }
    return render(request, 'update_phong_kham.html', context=data)

def store_update_phong_kham(request):
    if request.method == "POST":
        id                  = request.POST.get('id')
        ten_phong_kham      = request.POST.get('ten_phong_kham')
        so_dien_thoai       = request.POST.get('so_dien_thoai')
        email               = request.POST.get('email')
        gia_tri_diem_tich   = request.POST.get('gia_tri_diem_tich')
        chu_khoan           = request.POST.get('chu_khoan')
        so_tai_khoan        = request.POST.get('so_tai_khoan')
        thong_tin_ngan_hang = request.POST.get('thong_tin_ngan_hang')

        phong_kham = get_object_or_404(PhongKham, id=id)
        phong_kham.ten_phong_kham      = ten_phong_kham
        phong_kham.so_dien_thoai       = so_dien_thoai
        phong_kham.email               = email
        phong_kham.gia_tri_diem_tich   = gia_tri_diem_tich
        phong_kham.chu_khoan           = chu_khoan
        phong_kham.so_tai_khoan        = so_tai_khoan
        phong_kham.thong_tin_ngan_hang = thong_tin_ngan_hang
        phong_kham.save()

        from actstream import action
        action.send(request.user, verb='đã cập nhật thông tin phòng khám')

        response = {
            'status': 200,
            'message': 'Cập Nhật Thông Tin Thành Công'
        }
        return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")

@login_required(login_url='/dang_nhap/')
def hoa_don_dich_vu_bao_hiem(request, *args, **kwargs):
    id = kwargs.get('id')
    # chuoi_kham = ChuoiKham.objects.filter(benh_nhan__id=user_id, trang_thai__id = 4)[0]
    try:
        chuoi_kham = ChuoiKham.objects.get(id=id)
        hoa_don_dich_vu = chuoi_kham.hoa_don_dich_vu
        if hoa_don_dich_vu is None:
            return render(request, '404.html')
        danh_sach_phan_khoa = chuoi_kham.phan_khoa_kham.filter(bao_hiem = True)
        tong_tien = []
        bao_hiem = []
        for khoa_kham in danh_sach_phan_khoa:
            if khoa_kham.bao_hiem:
                # gia = khoa_kham.dich_vu_kham.don_gia * decimal.Decimal((1 - (khoa_kham.dich_vu_kham.bao_hiem_dich_vu_kham.dang_bao_hiem)/100))
                gia = khoa_kham.dich_vu_kham.don_gia
                bao_hiem.append(gia)
            else:
                gia = khoa_kham.dich_vu_kham.don_gia
            tong_tien.append(gia)
        total_spent = sum(tong_tien)
        tong_bao_hiem = sum(bao_hiem)
        thanh_tien = total_spent - tong_bao_hiem
        tong_tien.clear()
        bao_hiem.clear()
        phong_chuc_nang = PhongChucNang.objects.all()
        phong_kham = PhongKham.objects.all().first()
        data = {
            'chuoi_kham'         : chuoi_kham,
            'tong_tien'          : total_spent,
            'phong_chuc_nang'    : phong_chuc_nang,
            'danh_sach_phan_khoa': danh_sach_phan_khoa,
            'tong_tien'          : total_spent,
            'ap_dung_bao_hiem'   : tong_bao_hiem,
            'thanh_tien'         : thanh_tien,
            'hoa_don_dich_vu'    : hoa_don_dich_vu,
            'phong_kham'         : phong_kham,
        }
        return render(request, 'phong_tai_chinh/hoa_don_dich_vu_bao_hiem.html', context=data)
    except ChuoiKham.DoesNotExist:
        return render(request, '404.html')

@login_required(login_url='/dang_nhap/')
def hoa_don_thuoc_bao_hiem(request, **kwargs):
    id = kwargs.get('id')
    don_thuoc = DonThuoc.objects.get(id = id)
    danh_sach_thuoc = don_thuoc.ke_don.filter(bao_hiem=True)
    # tong_tien = []
    # for thuoc_instance in danh_sach_thuoc:
    #     gia = int(thuoc_instance.thuoc.don_gia_tt) * thuoc_instance.so_luong
    #     tong_tien.append(gia)
    bao_hiem = []
    tong_tien = []
    for thuoc_instance in danh_sach_thuoc:
        if thuoc_instance.bao_hiem:
            gia = int(thuoc_instance.thuoc.don_gia_tt) * \
                thuoc_instance.so_luong
            bao_hiem.append(gia)
        else:
            gia = int(thuoc_instance.thuoc.don_gia_tt) * \
                thuoc_instance.so_luong
        tong_tien.append(gia)

    total_spent = sum(tong_tien)
    tong_bao_hiem = sum(bao_hiem)
    thanh_tien = total_spent - tong_bao_hiem
    
    total_spent = sum(tong_tien)
    tong_tien.clear()
    bao_hiem.clear()
    
    phong_chuc_nang = PhongChucNang.objects.all()
    phong_kham = PhongKham.objects.all().first()

    data = {
        'danh_sach_thuoc': danh_sach_thuoc,
        'tong_tien'      : total_spent,
        'don_thuoc'      : don_thuoc,
        'phong_chuc_nang': phong_chuc_nang,
        'thanh_tien'     : thanh_tien,
        'tong_bao_hiem'  : tong_bao_hiem,
        'phong_kham'     : phong_kham,
    }
    return render(request, 'phong_tai_chinh/hoa_don_thuoc_bao_hiem.html', context=data)

def thong_tin_phong_kham(request, *args, **kwargs):
    phong_chuc_nang = PhongChucNang.objects.all()

    data = {
        'phong_chuc_nang': phong_chuc_nang,
    }
    return render(request, 'thong_tin_phong_kham.html', context = data)

# END

# NEW FROM LONG 13/1
def create_user_index(request):
    if request.method == "POST":
        ho_ten         = request.POST.get("ho_ten", None)
        so_dien_thoai  = request.POST.get("so_dien_thoai", None)
        password       = request.POST.get("password", None)
        cmnd_cccd      = request.POST.get("cmnd_cccd", None)
        dia_chi        = request.POST.get("dia_chi", None)
 
        if len(ho_ten) == 0:
            return HttpResponse(json.dumps({'message': "Họ Tên Không Được Trống", 'status': '400'}), content_type='application/json; charset=utf-8')
 
        if User.objects.filter(so_dien_thoai=so_dien_thoai).exists():
            return HttpResponse(json.dumps({'message': "Số Điện Thoại Đã Tồn Tại", 'status': '409'}), content_type='application/json; charset=utf-8')
 
        if User.objects.filter(cmnd_cccd=cmnd_cccd).exists():
            return HttpResponse(json.dumps({'message': "Số chứng minh thư đã tồn tại", 'status': '403'}), content_type = 'application/json; charset=utf-8')
 
        user = User.objects.create_user(
            ho_ten         = ho_ten.upper(), 
            so_dien_thoai  = so_dien_thoai, 
            password       = password,
            dia_chi        = dia_chi,
        )
        if cmnd_cccd != '':
            user.cmnd_cccd = cmnd_cccd
        user.save()
 
        response = {
            "message": "Đăng Kí Người Dùng Thành Công",
            "ho_ten": user.ho_ten,
            "so_dien_thoai": user.so_dien_thoai,
        }

        from actstream import action
        action.send(request.user, verb='đã tạo mới người dùng', target=user)

        return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")
    else:
        return HttpResponse(
            json.dumps({"nothing to see": "this isn't happening"}),
            content_type="application/json"
        )
 
def update_vat_tu(request, *args, **kwargs):
    id = kwargs.get('id')
    instance = get_object_or_404(VatTu, id=id)
    form = VatTuForm(request.POST or None, instance=instance)
    phong_chuc_nang = PhongChucNang.objects.all()
    data = {
        'form': form,
        'id'  : id,
        'phong_chuc_nang': phong_chuc_nang,
    }
    return render(request, 'phong_tai_chinh/update_vat_tu.html', context=data)
 
def store_update_vat_tu(request):
    if request.method == "POST":
        id                = request.POST.get('id')
        id_cong_ty        = request.POST.get('id_cong_ty')
        ma_hieu           = request.POST.get('ma_hieu')
        ma_vtyt_bv        = request.POST.get('ma_vtyt_bv')
        ten_vtyt_bv       = request.POST.get('ten_vtyt_bv')
        don_vi_tinh       = request.POST.get('don_vi_tinh')
        quy_cach          = request.POST.get('quy_cach')
        don_gia           = request.POST.get('don_gia')
        don_gia_tt        = request.POST.get('don_gia_tt')
        dinh_muc          = request.POST.get('dinh_muc')
        so_luong_kha_dung = request.POST.get('so_luong_kha_dung')
        hang_sx           = request.POST.get('hang_sx')
        nuoc_sx           = request.POST.get('nuoc_sx')
        quyet_dinh        = request.POST.get('quyet_dinh')
        nhom_vat_tu       = request.POST.get('nhom_vat_tu')
        cong_bo           = request.POST.get('cong_bo')
 
        nha_thau = CongTy.objects.get(id = id_cong_ty)
        nhom_vat_tu = NhomVatTu.objects.get(id = nhom_vat_tu)
        vat_tu = get_object_or_404(VatTu, id=id)
 
        vat_tu.ma_hieu           = ma_hieu
        vat_tu.ma_vtyt_bv        = ma_vtyt_bv
        vat_tu.ten_vtyt_bv       = ten_vtyt_bv
        vat_tu.don_vi_tinh       = don_vi_tinh
        vat_tu.quy_cach          = quy_cach
        vat_tu.don_gia           = don_gia
        vat_tu.don_gia_tt        = don_gia_tt
        vat_tu.dinh_muc          = dinh_muc
        vat_tu.so_luong_kha_dung = so_luong_kha_dung
        vat_tu.hang_sx           = hang_sx
        vat_tu.nuoc_sx           = nuoc_sx
        vat_tu.quyet_dinh        = quyet_dinh
        vat_tu.nhom_vat_tu       = nhom_vat_tu
        vat_tu.cong_bo           = cong_bo
        vat_tu.nha_thau          = nha_thau
        vat_tu.save()
 
        from actstream import action
        action.send(request.user, verb='đã cập nhật thông tin vật tư', target=vat_tu)

        response = {
            'status': 200,
            'message': 'Cập Nhật Thông Tin Thành Công'
        }
        return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")
    else:
        response = {
            'message' : "Có Lỗi Xảy Ra Trong Quá Trình Xử Lý"
        }
        return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")


def view_ket_qua_xet_nghiem(request, **kwargs):
    id_chuoi_kham = kwargs.get('id', None)
    id_dich_vu = kwargs.get('id_dich_vu', None)
    id_phan_khoa = kwargs.get('id_phan_khoa', None)
    try:
        chuoi_kham = ChuoiKham.objects.get(id=id_chuoi_kham)
        benh_nhan = chuoi_kham.benh_nhan
        # phong_chuc_nang = PhongChucNang.objects.get(id=id_phong_chuc_nang)
        dich_vu = DichVuKham.objects.get(id=id_dich_vu)
        id_phong_chuc_nang = dich_vu.phong_chuc_nang.id
        if dich_vu.check_chi_so:
            list_chi_so_xet_nghiem = dich_vu.chi_so_xet_nghiem.all()
            id_phong_chuc_nang = dich_vu.phong_chuc_nang.id
            phong_chuc_nang = PhongChucNang.objects.all()
            nhom_chi_so = NhomChiSoXetNghiem.objects.all()
            ho_ten_benh_nhan = benh_nhan.ho_ten.upper()
            now       = datetime.now()
            date_time = now.strftime("%m%d%y%H%M%S")
            bac_si_chuyen_khoa = request.user

            ma_ket_qua = str(id_phong_chuc_nang) +'-'+ getSubName(ho_ten_benh_nhan) + '-' + str(date_time)
            context = {
                'id_chuoi_kham': id_chuoi_kham,
                'dich_vu': dich_vu,
                'benh_nhan': benh_nhan,
                'chi_so_xet_nghiem': list_chi_so_xet_nghiem,
                'ma_ket_qua': ma_ket_qua,
                'phong_chuc_nang': phong_chuc_nang,
                'id_phong_chuc_nang': id_phong_chuc_nang,
                'id_phan_khoa': id_phan_khoa,
                'bac_si_chuyen_khoa': ["Bác Sỹ " + bac_si_chuyen_khoa.ho_ten.upper()],
                'ho_ten_benh_nhan': ho_ten_benh_nhan,
                'nhom_chi_so': nhom_chi_so,
            }
            return render(request, 'bac_si_chuyen_khoa/chi_so_xet_nghiem.html', context=context)
        elif dich_vu.check_html:
            id_phong_chuc_nang = dich_vu.phong_chuc_nang.id
            phong_chuc_nang = PhongChucNang.objects.all()
            danh_sach_mau_phieu = MauPhieu.objects.filter(codename__isnull=False)
            ho_ten_benh_nhan = benh_nhan.ho_ten.upper()
            now       = datetime.now()
            date_time = now.strftime("%m%d%y%H%M%S")
            
            bac_si_chuyen_khoa = request.user
            ngay_kham = datetime.now()
            bac_si_lam_sang = chuoi_kham.bac_si_dam_nhan
            ma_ket_qua = str(id_phong_chuc_nang) +'-'+ getSubName(ho_ten_benh_nhan) + '-' + str(date_time)

            data_dict = {}
            data_dict['{benh_nhan}'] = benh_nhan.ho_ten.upper()
            data_dict['{dia_chi}'] = benh_nhan.get_dia_chi()
            data_dict['{gioi_tinh}'] = benh_nhan.get_gioi_tinh()
            data_dict['{tuoi}'] = str(benh_nhan.tuoi())
            data_dict['{bac_si_chuyen_khoa}'] = bac_si_chuyen_khoa.ho_ten.upper()
            data_dict['{bac_si_lam_sang}'] = bac_si_lam_sang.ho_ten.upper()
            data_dict['{ngay_kham}'] = "Ngày " + str(ngay_kham.strftime('%d')) + " Tháng " + str(ngay_kham.strftime('%m')) + " Năm " + str(ngay_kham.strftime('%Y'))

            context = {
                'ma_ket_qua': ma_ket_qua,
                'id_chuoi_kham': id_chuoi_kham,
                'id_phong_chuc_nang': id_phong_chuc_nang,
                'phong_chuc_nang': phong_chuc_nang,
                'danh_sach_mau_phieu': danh_sach_mau_phieu,
                'id_phan_khoa': id_phan_khoa,
                'dich_vu': dich_vu,
                'ho_ten_benh_nhan' : ho_ten_benh_nhan,
                'data': data_dict,
            }
            return render(request, 'bac_si_chuyen_khoa/upload_phieu_ket_qua.html', context=context)
        else:
            id_phong_chuc_nang = dich_vu.phong_chuc_nang.id
            phong_chuc_nang = PhongChucNang.objects.all()
            ho_ten_benh_nhan = benh_nhan.ho_ten
            now       = datetime.now()
            date_time = now.strftime("%m%d%y%H%M%S")
            bac_si_chuyen_khoa = request.user

            ma_ket_qua = str(id_phong_chuc_nang) +'-'+ getSubName(ho_ten_benh_nhan) + '-' + str(date_time)
            context = {
                'benh_nhan': benh_nhan,
                'ma_ket_qua': ma_ket_qua,
                'id_chuoi_kham': id_chuoi_kham,
                'id_phong_chuc_nang': id_phong_chuc_nang,
                'phong_chuc_nang': phong_chuc_nang,
                'id_phan_khoa': id_phan_khoa,
                'bac_si_chuyen_khoa': bac_si_chuyen_khoa,
                'ho_ten_benh_nhan' : ho_ten_benh_nhan,
            }
            return render(request, 'bac_si_chuyen_khoa/upload.html', context=context)

    except ChuoiKham.DoesNotExist:
        return render(request, 'not_found.html')

def store_ket_qua_xet_nghiem(request):
    if request.method == "POST":
        id_chuoi_kham = request.POST.get('id_chuoi_kham')
        id_phan_khoa = request.POST.get('id_phan_khoa')
        ma_ket_qua = request.POST.get('ma_ket_qua')
        mo_ta = request.POST.get('mo_ta')
        ket_qua = request.POST.get('ket_qua')
        data = request.POST.get('data')
        list_data = json.loads(data)

        chuoi_kham = ChuoiKham.objects.filter(id=id_chuoi_kham).first()
        
        trang_thai = TrangThaiKhoaKham.objects.get_or_create(trang_thai_khoa_kham='Đã Tải Lên Kết Quả')[0]
        phan_khoa_kham = PhanKhoaKham.objects.filter(id=id_phan_khoa).first()
        phan_khoa_kham.trang_thai = trang_thai
        phan_khoa_kham.save()

        ket_qua_tong_quat = KetQuaTongQuat.objects.get_or_create(chuoi_kham=chuoi_kham)[0]
        ket_qua_chuyen_khoa = KetQuaChuyenKhoa.objects.create(
            phan_khoa_kham = phan_khoa_kham,
            ket_qua_tong_quat=ket_qua_tong_quat,
            ma_ket_qua=ma_ket_qua, 
            mo_ta=mo_ta, 
            ket_luan=ket_qua, 
            chi_so=True,
            bac_si_chuyen_khoa = request.user
        )

        bulk_create_data = []
        for i in list_data:
            id = i['key3']
            ket_qua_xet_nghiem = i['value']
            danh_gia = i['value2']
            chi_so = ChiSoXetNghiem.objects.filter(id=id).first()

            model = KetQuaXetNghiem(
                phan_khoa_kham = phan_khoa_kham,
                ket_qua_chuyen_khoa = ket_qua_chuyen_khoa,
                chi_so_xet_nghiem = chi_so,
                ket_qua_xet_nghiem = ket_qua_xet_nghiem,
                danh_gia_chi_so = danh_gia,
            )
            bulk_create_data.append(model)
        KetQuaXetNghiem.objects.bulk_create(bulk_create_data)

        from actstream import action
        action.send(request.user, verb="tải lên kết quả", action_object=ket_qua_chuyen_khoa, target=chuoi_kham.benh_nhan)

        response = {
            'status': 200,
            'message': 'Lưu Dữ Liệu Thành Công'
        }
    else:
        response = {
            'status': 404,
            'message': 'Không gửi được dữ liệu!'
        }
    return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")

def store_ket_qua_chuyen_khoa_html(request):
    if request.method == "POST":
        id_phan_khoa = request.POST.get('id_phan_khoa')
        id_chuoi_kham = request.POST.get("id_chuoi_kham")
        noi_dung = request.POST.get('noi_dung')
        mo_ta = request.POST.get('mo_ta')
        ma_ket_qua    = request.POST.get('ma_ket_qua')

        if noi_dung == "":
            response = {
                'status': 404,
                'message': 'Vui lòng kiểm tra lại phiếu kết quả'
            }
            return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")
        
        trang_thai = TrangThaiKhoaKham.objects.get_or_create(trang_thai_khoa_kham='Đã Tải Lên Kết Quả')[0]
        phan_khoa = PhanKhoaKham.objects.filter(id=id_phan_khoa).first()
        phan_khoa.trang_thai = trang_thai
        phan_khoa.save()
        chuoi_kham = ChuoiKham.objects.filter(id=id_chuoi_kham).first()

        ket_qua_tong_quat = KetQuaTongQuat.objects.get_or_create(chuoi_kham=chuoi_kham)[0]
        ket_qua_chuyen_khoa = KetQuaChuyenKhoa.objects.create(
            ma_ket_qua = ma_ket_qua,
            phan_khoa_kham = phan_khoa,
            ket_qua_tong_quat = ket_qua_tong_quat,
            mo_ta = mo_ta,
            html = True,
            bac_si_chuyen_khoa = request.user,
        )
        HtmlKetQua.objects.create(
            phan_khoa_kham = phan_khoa,
            ket_qua_chuyen_khoa = ket_qua_chuyen_khoa,
            noi_dung = noi_dung
        )

        from actstream import action
        action.send(request.user, verb='tải lên kết quả', action_object=ket_qua_chuyen_khoa, target=chuoi_kham.benh_nhan)

        response = {
            'status': 200,
            'message': 'Lưu Dữ Liệu Thành Công'
        }
        return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")
    else:
        response = {
            'status': 404,
            'message': 'Có Lỗi Xảy Ra'
        }
        return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")


 
def them_ma_benh_excel(request):
    return render(request, 'phong_tai_chinh/them_benh_excel.html')

def import_ma_benh_excel(request):
    import pandas as pd
    try:
        if request.method == "POST":
            # for value in request.FILES.values():
            #     file = FilePhongKham.objects.create(file=value)
            file = FilePhongKham.objects.all()
            file = file[0]
            file_url = file.file.url[1:]
            
            excel_data_df = pd.read_excel(file_url, sheet_name='Sheet1')
            json_str = excel_data_df.to_json(orient='records')
            json_objects = json.loads(json_str)
            
            bulk_create_data = []
            for i in json_objects:
                stt = i['STT']
                ma_chuong = i['MA_CHUONG']
                ten_chuong = i['TEN_CHUONG']
                ma_nhom = i['MA_NHOM_CHINH']
                ten_nhom = i['TEN_NHOM_CHINH']
                ma_loai = i['MA_LOAI']
                ten_loai = i['TEN_LOAI']
                ma_benh = i['MA_BENH']
                ten_benh = i['TEN_BENH']
                ma_nhom_bcao_byt = i['MA_NHOM_BCAO_BYT']
                ma_nhom_chi_tiet = i['MA_NHOM_CHI_TIET']

                chuong_benh = DanhMucChuongBenh.objects.get_or_create(stt=stt, ma_chuong=ma_chuong, ten_chuong=ten_chuong)[0]
                nhom_benh = DanhMucNhomBenh.objects.get_or_create(chuong_benh=chuong_benh, ma_nhom_chinh=ma_nhom, ten_nhom_chinh=ten_nhom)[0]
                loai_benh = DanhMucLoaiBenh.objects.get_or_create(nhom_benh=nhom_benh, ma_loai=ma_loai, ten_loai=ten_loai)[0]
                
                model = DanhMucBenh(loai_benh=loai_benh, ma_benh=ma_benh, ten_benh=ten_benh, ma_nhom_bcao_byt=ma_nhom_bcao_byt, ma_nhom_chi_tiet=ma_nhom_chi_tiet)
                bulk_create_data.append(model)

            DanhMucBenh.objects.bulk_update_or_create(bulk_create_data, [
                'loai_benh',
                'ma_benh', 
                'ten_benh',
                'ma_nhom_bcao_byt',
                'ma_nhom_chi_tiet',
            ], match_field = 'ma_benh')
            response = {
                'status': 200,
                'message': 'Import Thanh Cong'
            }
            return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")
        else:
            response = {
                'status': 404,
                'message': 'Loi'
            }
            return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")
        
    except Exception as e:
        print(e)


def danh_sach_mau_phieu(request):
    phong_chuc_nang = PhongChucNang.objects.all()

    data = {
        'phong_chuc_nang' : phong_chuc_nang,
    }
    return render(request, 'le_tan/danh_sach_mau_phieu.html', context = data)

def them_mau_phieu(request):
    dich_vu = DichVuKham.objects.all()
    phong_chuc_nang = PhongChucNang.objects.all()
    context = {
        'dich_vu': dich_vu,
        'phong_chuc_nang' : phong_chuc_nang,
    }
    return render(request, 'le_tan/them_mau_phieu.html', context=context)

def chi_tiet_mau_phieu(request, **kwargs):
    id = kwargs.get('id')
    # mau_phieu = MauPhieu.objects.get(id=id)
    instance = get_object_or_404(MauPhieu, id=id)
    if instance.codename is None:
        form = MauPhieuForm(request.POST or None, instance=instance)
        context = {
            'mau_phieu': instance,
            'form': form,
            'id_mau_phieu': id
        }
        return render(request, 'le_tan/update_mau_phieu.html', context=context)
    else:
        form = MauHoaDonForm(request.POST or None, instance=instance)
        context = {
            'mau_hoa_don': instance,
            'form': form,
            'id_mau_hoa_don': id
        }
        return render(request, 'le_tan/update_mau_hoa_don.html', context=context)

def create_mau_phieu(request):
    if request.method == "POST":
        id_dich_vu = request.POST.get('id_dich_vu')
        noi_dung = request.POST.get('noi_dung')
        html = request.POST.get("html")
       
        try:
            dich_vu = DichVuKham.objects.filter(id=id_dich_vu).first()
            ten_mau_phieu = dich_vu.ten_dvkt
            if html == '1':
                dich_vu.html = True
                dich_vu.save()
            mau_phieu = MauPhieu.objects.create(dich_vu=dich_vu, ten_mau=ten_mau_phieu, noi_dung=noi_dung)
            from actstream import action
            action.send(request.user, verb='đã tạo mới mẫu phiếu', target=mau_phieu)

        except DichVuKham.DoesNotExist:
            response = {
                'status': 404,
                'message': 'Không Tồn Tại Dịch Vụ Khám'
            }
            return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")
        
        response = {
            'status': 200,
            'message': 'Tạo mới thành công'
        }
        return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")
    else:
        response = {
            'status': 404,
            'message': 'Xảy ra lỗi'
        }
        return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")

def update_mau_phieu(request):
    if request.method == "POST":
        id_dich_vu = request.POST.get('id_dich_vu')
        id_mau_phieu = request.POST.get('id_mau_phieu')
        noi_dung = request.POST.get('noi_dung')
        try:
            dich_vu = DichVuKham.objects.filter(id=id_dich_vu).first()
            # mau_phieu = MauPhieu.objects.update(dich_vu=dich_vu, ten_mau=ten_mau_phieu, noi_dung=noi_dung)
            mau_phieu = MauPhieu.objects.filter(id=id_mau_phieu).first()
            mau_phieu.dich_vu = dich_vu
            mau_phieu.noi_dung = noi_dung
            mau_phieu.save()

            from actstream import action
            action.send(request.user, verb='đã cập nhật mẫu phiếu', target=mau_phieu)
        except DichVuKham.DoesNotExist:
            response = {
                'status': 404,
                'message': 'Không Tồn Tại Dịch Vụ Khám'
            }
            return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")
        
        response = {
            'status': 200,
            'message': 'Cập Nhật Thành Công'
        }
        return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")
    else:
        response = {
            'status': 404,
            'message': 'Xảy ra lỗi'
        }
        return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")

def test_mau_phieu(request):
    chi_so = ChiSoXetNghiem.objects.all()
    ten_chi_so = [i.ten_chi_so for i in chi_so]
    chi_so_binh_thuong = [i.chi_tiet.chi_so_binh_thuong_max for i in chi_so]
    don_vi = [i.chi_tiet.don_vi_do if i.chi_tiet.don_vi_do is not None else '' for i in chi_so]

    mau_phieu = MauPhieu.objects.all()[0]

    context = {
        'chi_so_xet_nghiem': chi_so,
        'list_ten_chi_so': ten_chi_so,
        'list_chi_so_binh_thuong': chi_so_binh_thuong,
        'mau_phieu': mau_phieu,
        'don_vi_do': don_vi
    }
    return render(request, 'test_mau_phieu.html', context=context)

def tat_ca_lich_hen_banh_nhan(request, **kwargs):
    id_benh_nhan = kwargs.get('id')
    context = {
        'id_benh_nhan': id_benh_nhan
    }
    return render(request, 'le_tan/danh_sach_lich_hen_benh_nhan.html', context=context)

def chi_tiet_lich_hen_benh_nhan(request, **kwargs):
    id_lich_hen = kwargs.get('id')
    lich_hen = LichHenKham.objects.get(id=id_lich_hen)
    chuoi_kham = lich_hen.danh_sach_chuoi_kham.all()[0]
    hoa_don_lam_sang = lich_hen.hoa_don_lam_sang.all()[0]
    hoa_don_chuoi_kham = chuoi_kham.hoa_don_dich_vu
    don_thuoc = chuoi_kham.don_thuoc_chuoi_kham.all()[0]
    hoa_don_thuoc = don_thuoc.hoa_don_thuoc
    ket_qua = chuoi_kham.ket_qua_tong_quat.all()[0].ket_qua_chuyen_khoa.all().filter(chi_so=False).count()
    ket_qua_xet_nghiem = chuoi_kham.ket_qua_tong_quat.all()[0].ket_qua_chuyen_khoa.all().filter(chi_so=True).count()
    ket_qua_chuyen_khoa = chuoi_kham.ket_qua_tong_quat.all()[0].ket_qua_chuyen_khoa.all()
    ket_qua_tong_quat = chuoi_kham.ket_qua_tong_quat.all()[0]
    context = {
        'lich_hen': lich_hen,
        'chuoi_kham': chuoi_kham,
        'hoa_don_lam_sang': hoa_don_lam_sang,
        'hoa_don_chuoi_kham': hoa_don_chuoi_kham,
        'hoa_don_thuoc': hoa_don_thuoc,
        'so_ket_qua': ket_qua,
        'so_ket_qua_xet_nghiem': ket_qua_xet_nghiem,
        'ket_qua_chuyen_khoa': ket_qua_chuyen_khoa,
        'ket_qua_tong_quat': ket_qua_tong_quat
    }
    return render(request, 'le_tan/chi_tiet_lich_hen_benh_nhan.html', context=context)

def chi_tiet_ket_qua_xet_nghiem(request, **kwargs):
    id_ket_qua_chuyen_khoa = kwargs.get('id')
    ket_qua_chuyen_khoa = KetQuaChuyenKhoa.objects.filter(id=id_ket_qua_chuyen_khoa).first()
    if ket_qua_chuyen_khoa is not None:
        ket_qua_xet_nghiem = ket_qua_chuyen_khoa.ket_qua_xet_nghiem.all()
        phan_khoa_kham = ket_qua_xet_nghiem[0].phan_khoa_kham
        dich_vu = phan_khoa_kham.dich_vu_kham
        mau_phieu = dich_vu.mau_phieu.all()[0]

        benh_nhan = [phan_khoa_kham.get_ten_benh_nhan()]
        dia_chi = [phan_khoa_kham.get_dia_chi_benh_nhan()]
        tuoi = [str(phan_khoa_kham.get_tuoi_benh_nhan())]
        gioi_tinh = [phan_khoa_kham.get_gioi_tinh_benh_nhan()]
        bac_si_chi_dinh = [phan_khoa_kham.get_bac_si_chi_dinh()]
        chan_doan = [ket_qua_chuyen_khoa.get_ket_luan()]

        ten_chi_so = [i.get_ten_chi_so() for i in ket_qua_xet_nghiem]
        chi_so_binh_thuong = [f'{i.get_chi_so_min()} - {i.get_chi_so_max()}' for i in ket_qua_xet_nghiem]
        don_vi = [i.get_don_vi() for i in ket_qua_xet_nghiem]
        ket_qua = [i.get_ket_qua_xet_nghiem() for i in ket_qua_xet_nghiem]
        ngay_kham = ["Ngày " + str(ket_qua_chuyen_khoa.thoi_gian_tao.strftime('%d')) + " Tháng " + str(ket_qua_chuyen_khoa.thoi_gian_tao.strftime('%m')) + " Năm " + str(ket_qua_chuyen_khoa.thoi_gian_tao.strftime('%Y'))]
        bac_si_chuyen_khoa = ["Bác Sỹ " + ket_qua_chuyen_khoa.bac_si_chuyen_khoa.ho_ten]

        context = {
            'mau_phieu': mau_phieu, 
            'list_ten_chi_so': ten_chi_so,
            'list_chi_so_binh_thuong': chi_so_binh_thuong,
            'list_don_vi': don_vi,
            'list_ket_qua': ket_qua,
            'benh_nhan': benh_nhan,
            'dia_chi': dia_chi,
            'tuoi': tuoi,
            'gioi_tinh': gioi_tinh,
            'bac_si_chi_dinh': bac_si_chi_dinh,
            'chan_doan': chan_doan,
            'ngay_kham': ngay_kham,
            'bac_si_chuyen_khoa': bac_si_chuyen_khoa,
        }
        return render(request, 'mau_phieu.html', context=context)

    else:
        return render(request, '404.html')

def chi_tiet_phieu_ket_qua(request, **kwargs):
    id_ket_qua_chuyen_khoa = kwargs.get('id')
    ket_qua_chuyen_khoa = KetQuaChuyenKhoa.objects.filter(id=id_ket_qua_chuyen_khoa).first()
    if ket_qua_chuyen_khoa is not None:
        html_ket_qua = ket_qua_chuyen_khoa.html_ket_qua.all().first()
        noi_dung = html_ket_qua.noi_dung
        context = {
            'noi_dung': noi_dung,
        }
        return render(request, 'phieu_ket_qua.html', context=context)
    else:
        return render(request, '404.html')

def chi_tiet_phieu_ket_qua_lam_sang(request, **kwargs):
    id_ket_qua_tong_quat = kwargs.get('id')
    ket_qua_tong_quat = KetQuaTongQuat.objects.filter(id=id_ket_qua_tong_quat).first()
    if ket_qua_tong_quat is not None:
        html_ket_qua_tong_quat = ket_qua_tong_quat.html_ket_qua_tong_quat.all().first()
        noi_dung = html_ket_qua_tong_quat.noi_dung
        context = {
            'noi_dung': noi_dung,
        }
        return render(request, 'phieu_ket_qua_tong_quat.html', context=context)
    else:
        return render(request, '404.html')
        
def phieu_ket_qua(request, **kwargs):
    id_ket_qua_chuyen_khoa = kwargs.get('id')
    ket_qua_chuyen_khoa = KetQuaChuyenKhoa.objects.filter(id=id_ket_qua_chuyen_khoa).first()
    html_ket_qua = ket_qua_chuyen_khoa.html_ket_qua.all().first()
    noi_dung = html_ket_qua.noi_dung
    context = {
        'noi_dung': noi_dung,
    }
    return render(request, 'phieu_ket_qua_mobile.html', context=context)

def phieu_ket_qua_tong_quat(request, **kwargs):
    id_ket_qua_tong_quat = kwargs.get('id')
    ket_qua_tong_quat = get_object_or_404(KetQuaTongQuat, id=id_ket_qua_tong_quat)
    html_ket_qua = ket_qua_tong_quat.html_ket_qua_tong_quat.all().first()
    noi_dung = html_ket_qua.noi_dung
    context = {
        'noi_dung': noi_dung,
    }
    return render(request, 'phieu_ket_qua_mobile.html', context=context)

# UPDATE BY LONG
def xoa_bac_si(request):
    if request.method == 'POST':
        id    = request.POST.get('id')
        user_id = request.POST.get('user_id')
        BacSi.objects.get(id = id).delete()
        user = User.objects.get(id = user_id)
        user.delete()

        response = {
            'user' : user.ho_ten
        }
        return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")

def create_lich_tai_kham(request):
    if request.method == "POST":
        user_id           = request.POST.get('user_id')
        thoi_gian_bat_dau = request.POST.get('thoi_gian_bat_dau')
        ly_do             = "Tái Khám"
        user = User.objects.get(id=user_id)

        thoi_gian_bat_dau = datetime.strptime(thoi_gian_bat_dau, format_2)
        thoi_gian = thoi_gian_bat_dau.strftime("%Y-%m-%d %H:%M")
        
        trang_thai = TrangThaiLichHen.objects.get_or_create(ten_trang_thai = "Đã Đặt Trước")[0]
        lich_hen = LichHenKham.objects.create(
            benh_nhan = user,
            nguoi_phu_trach = request.user,
            thoi_gian_bat_dau = thoi_gian,
            ly_do = ly_do,
            trang_thai = trang_thai,
        )
        lich_hen.save()

        from actstream import action 
        action.send(request.user, verb='đã tạo mới lịch tái khám', action_object=lich_hen, target=lich_hen.benh_nhan)

        response = {
            'message': "Thêm lịch tái khám thành công!"
        }
        return HttpResponse(json.dumps(response), content_type='application/json; charset=utf-8')
    else:
        response = {
            'message': "Có Lỗi Xảy Ra!"
        }
        return HttpResponse(json.dumps(response), content_type='application/json; charset=utf-8')

def create_don_thuoc_rieng(request):
    phong_chuc_nang = PhongChucNang.objects.all()
    nguoi_dung = User.objects.filter(chuc_nang=1)
    mau_hoa_don = MauPhieu.objects.filter(codename='don_thuoc').first()
    nguoi_thanh_toan = request.user.ho_ten
    thoi_gian_thanh_toan = datetime.now()
    data = {
        'phong_chuc_nang': phong_chuc_nang,
        'nguoi_dung' : nguoi_dung,
        'mau_hoa_don': mau_hoa_don,
        'nguoi_thanh_toan': nguoi_thanh_toan,
        'thoi_gian_thanh_toan': f"{thoi_gian_thanh_toan.strftime('%H:%m')} Ngày {thoi_gian_thanh_toan.strftime('%d')} Tháng {thoi_gian_thanh_toan.strftime('%m')} Năm {thoi_gian_thanh_toan.strftime('%Y')}",
    }
    return render(request, 'phong_thuoc/tao_don_thuoc.html', context = data)

def store_don_thuoc_rieng(request):
    if request.method == "POST":
        from actstream import action
        request_data = request.POST.get('data', None)
        user_id = request.POST.get('user_id', None)
        ten_khach_vang_lai = request.POST.get("ten_khach_vang_lai", None)
        so_dien_thoai_khach = request.POST.get("so_dien_thoai_khach", None)
        # id_chuoi_kham = request.POST.get('id_chuoi_kham', None)
        data = json.loads(request_data)

        now = datetime.now()
        date_time = now.strftime("%m%d%y%H%M%S")
        
        bulk_create_data = []
        trang_thai = TrangThaiDonThuoc.objects.get_or_create(trang_thai="Chờ Thanh Toán")[0]
        if user_id == "":
            subName = getSubName(ten_khach_vang_lai)
            user = ten_khach_vang_lai + '-' + str(so_dien_thoai_khach)
            ma_don_thuoc = subName + '-' + date_time
            don_thuoc = DonThuoc.objects.get_or_create(benh_nhan_vang_lai=user, bac_si_ke_don=request.user, trang_thai=trang_thai, ma_don_thuoc=ma_don_thuoc)[0]
            
            action.send(request.user, verb='đã tạo mới đơn thuốc lẻ cho khách vãng lai')
        elif ten_khach_vang_lai == "":
            user = get_object_or_404(User, id=user_id)
            subName = getSubName(user.ho_ten)
            ma_don_thuoc = subName + '-' + date_time
            don_thuoc = DonThuoc.objects.get_or_create(benh_nhan=user, bac_si_ke_don=request.user, trang_thai=trang_thai, ma_don_thuoc=ma_don_thuoc)[0]
            
            action.send(request.user, verb='đã tạo mới đơn thuốc lẻ cho bệnh nhân', target=user)
        
        for i in data:
            thuoc = Thuoc.objects.only('id').get(id=i['obj']['id'])
            ke_don_thuoc = KeDonThuoc(don_thuoc=don_thuoc, thuoc=thuoc, so_luong=i['obj']['so_luong'], cach_dung=i['obj']['duong_dung'], ghi_chu=i['obj']['ghi_chu'], bao_hiem=i['obj']['bao_hiem'])
            bulk_create_data.append(ke_don_thuoc)

        KeDonThuoc.objects.bulk_create(bulk_create_data)

        response = {
            'status': 200, 
            'message': 'Kê Đơn Thành Công', 
            'url': '/danh_sach_kham/'
        }
    else:
        response = {
            'status': 404,
            'message': 'Không gửi được dữ liệu, vui lòng kiểm tra lại'
        } 
    return HttpResponse(json.dumps(response), content_type='application/json; charset=utf-8')

# END UPDATE

def dang_ki_tieu_chuan_view(request):
    dich_vu = DichVuKham.objects.all()
    doi_tuong = DoiTuongXetNghiem.objects.all()

    context = {
        'range': range(5),
        'dich_vu': dich_vu,
        'doi_tuong_xet_nghiem': doi_tuong 
    }
    return render(request, 'le_tan/dang_ki_tieu_chuan.html', context=context)

def create_tieu_chuan(request):
    if request.method == "POST":

        nhom_chi_so = request.POST.get('nhom_chi_so')
        data = request.POST.get('data')
        list_data = json.loads(data)

        nhom_chi_so = NhomChiSoXetNghiem.objects.create(ten_nhom = nhom_chi_so)

        list_bulk_create = []

        for i in list_data:
            if i[0]['value'] != '' and i[1]['value'] != '':
                ma_chi_so = i[0]['value']
                ten_chi_so = i[1]['value']
                chi_so_min = i[2]['value']
                chi_so_max = i[3]['value']
                don_vi = i[4]['value']
                ghi_chu = i[5]['value']

                chi_tiet = ChiTietChiSoXetNghiem.objects.create(
                    chi_so_binh_thuong_min = chi_so_min,
                    chi_so_binh_thuong_max = chi_so_max,
                    don_vi_do = don_vi,
                    ghi_chu = ghi_chu
                )

                model = ChiSoXetNghiem(
                    nhom_chi_so = nhom_chi_so,
                    ma_chi_so = ma_chi_so,
                    ten_chi_so = ten_chi_so,
                    chi_tiet = chi_tiet
                )

                list_bulk_create.append(model)

            else:
                response = {
                    'status': 404, 
                    'message': 'Hãy Xóa Những Vùng Dữ Liệu Trống'
                }
                return HttpResponse(json.dumps(response), content_type='application/json; charset=utf-8')

        ChiSoXetNghiem.objects.bulk_create(list_bulk_create)

        from actstream import action
        action.send(request.user, verb='đã đăng kí mới tiêu chuẩn cho dịch vụ kỹ thuật')

        response = {
            'status': 200, 
            'message': 'Đăng kí tiêu chuẩn thành công'
        }
    else:
        response = {
            'status': 404, 
            'message': 'Không thể gửi lên dữ liệu'
        }
    return HttpResponse(json.dumps(response), content_type='application/json; charset=utf-8')
    
def list_tieu_chuan_dich_vu(request):
    phong_chuc_nang = PhongChucNang.objects.all()

    data = {
        'phong_chuc_nang' : phong_chuc_nang,
    }
    return render(request, 'le_tan/danh_sach_tieu_chuan_dich_vu.html', context = data)

def chi_tiet_tieu_chuan_dich_vu(request, **kwargs):
    id_nhom_chi_so = kwargs.get('id')
    nhom_chi_so = get_object_or_404(NhomChiSoXetNghiem, id=id_nhom_chi_so)
    tieu_chuan_dich_vu = nhom_chi_so.chisoxetnghiem_set.all()
    ten_nhom_chi_so = nhom_chi_so.ten_nhom
    phong_chuc_nang = PhongChucNang.objects.all()

    context = {
        'tieu_chuan_dich_vu': tieu_chuan_dich_vu,
        'phong_chuc_nang' : phong_chuc_nang,
        'ten_nhom_chi_so': ten_nhom_chi_so,
        'nhom_chi_so': nhom_chi_so,
    }
    return render(request, 'le_tan/chi_tiet_tieu_chuan.html', context=context)

def chinh_sua_tieu_chuan_dich_vu(request):
    if request.method == "POST":
        data = request.POST.get('data')
        list_data = json.loads(data)
        id_nhom_chi_so = request.POST.get('id_nhom_chi_so')
        ten_nhom_chi_so = request.POST.get('nhom_chi_so')

        try:
            nhom_chi_so_xet_nghiem = NhomChiSoXetNghiem.objects.get(id=id_nhom_chi_so)
            nhom_chi_so_xet_nghiem.ten_nhom = ten_nhom_chi_so
            nhom_chi_so_xet_nghiem.save()

        except NhomChiSoXetNghiem.DoesNotExist:
            response = {
                'status': 404,
                'message': "Không tìm thấy thông tin nhóm chỉ số của tiêu chuẩn"
            }
            return HttpResponse(json.dumps(response), content_type='application/json; charset=utf-8')

        for i in list_data:
            id_chi_so = i['id_chi_so']
            if id_chi_so != '':
                try:
                    chi_so = ChiSoXetNghiem.objects.get(id=id_chi_so)
                    chi_tiet_chi_so = chi_so.chi_tiet

                    chi_so.ma_chi_so = i['ma_chi_so']
                    chi_so.ten_chi_so = i['ten_chi_so']
                    
                    chi_tiet_chi_so.chi_so_binh_thuong_min = i['chi_so_min']
                    chi_tiet_chi_so.chi_so_binh_thuong_max = i['chi_so_max']
                    chi_tiet_chi_so.don_vi_do = i['don_vi']
                    chi_tiet_chi_so.ghi_chu = i['ghi_chu']
                    chi_tiet_chi_so.save()
            
                    chi_so.chi_tiet = chi_tiet_chi_so
                    chi_so.save()
                except ChiSoXetNghiem.DoesNotExist:
                    continue
            else:
                chi_so = ChiSoXetNghiem.objects.create(
                    nhom_chi_so = nhom_chi_so_xet_nghiem,
                    ma_chi_so = i['ma_chi_so'],
                    ten_chi_so = i['ten_chi_so']
                )
                chi_tiet_chi_so = ChiTietChiSoXetNghiem.objects.create(
                    chi_so_binh_thuong_min = i['chi_so_min'],
                    chi_so_binh_thuong_max = i['chi_so_max'],
                    don_vi_do = i['don_vi'],
                    ghi_chu = i['ghi_chu']
                )
                chi_tiet_chi_so.save()
                chi_so.chi_tiet = chi_tiet_chi_so
                chi_so.save()

        from actstream import action
        action.send(request.user, verb='đã cập nhật tiêu chuẩn dịch vụ kỹ thuật')

        response = {
            'status': 200,
            'message': "Cập nhật chỉ số tiêu chuẩn dịch vụ kĩ thuật thành công"
        }
    else: 
        response = {
            'status': 404,
            'message': 'Xảy ra lỗi trong quá trình gửi dữ liệu'
        }
    return HttpResponse(json.dumps(response), content_type='application/json; charset=utf-8')

from django.views.decorators.csrf import csrf_exempt

@csrf_exempt
def export_excel(request):
    if request.method == 'POST':
        startDate = request.POST.get('startDate')
        endDate = request.POST.get('endDate')
        
        start = datetime.strptime(startDate, "%d-%m-%Y")
        tomorrow_start = start + timedelta(1)
        if endDate == '':
            hoa_don_dich_vu = HoaDonChuoiKham.objects.filter(thoi_gian_tao__lte=tomorrow_start, thoi_gian_tao__gt=start)
        else:
            end = datetime.strptime(endDate, "%d-%m-%Y")
            if startDate == endDate:
                end = end + timedelta(1)
            hoa_don_dich_vu = HoaDonChuoiKham.objects.filter(thoi_gian_tao__lte=end, thoi_gian_tao__gt=start)

        serializer = FilterHoaDonChuoiKhamBaoHiemSerializer(hoa_don_dich_vu, many=True, context={'request': request})
        excel_data = serializer.data

        xlsx_data = WriteToExcel(excel_data)
        response = HttpResponse(
            content_type='application/ms-excel'
        )
        response['Content-Disposition'] = 'attachment; filename="Report.xlsx"'
        response.write(xlsx_data)
       
    else:
        response = HttpResponse(json.dumps({'message': "It's not gonna happen"}), content_type='application/json; charset=utf-8')
    return response

@csrf_exempt
def export_dich_vu_bao_hiem_excel(request):
    if request.method == 'POST':
        startDate = request.POST.get('startDate')
        endDate = request.POST.get('endDate')
        start = datetime.strptime(startDate, "%d-%m-%Y")
        tomorrow_start = start + timedelta(1)

        if endDate == '':
            dich_vu = PhanKhoaKham.objects.filter(thoi_gian_tao__lt=tomorrow_start, thoi_gian_tao__gte=start).values('dich_vu_kham__ten_dvkt').annotate(tong_tien=Sum('dich_vu_kham__don_gia')).order_by('dich_vu_kham__ten_dvkt').annotate(dich_vu_kham_count = Count('dich_vu_kham__ten_dvkt')).annotate(ma_dvkt=F('dich_vu_kham__ma_dvkt')).annotate(don_gia=F('dich_vu_kham__don_gia')).annotate(stt=F('dich_vu_kham__stt'))
            list_dich_vu = []
            for i in dich_vu:
                list_dich_vu.append(i)
            response = {
                'data' : list_dich_vu,
            }
        else:
            end = datetime.strptime(endDate, "%d-%m-%Y")
            if startDate == endDate:
                end = end + timedelta(1)
            dich_vu = PhanKhoaKham.objects.filter(thoi_gian_tao__lt=end, thoi_gian_tao__gte=start).filter(bao_hiem=True).values('dich_vu_kham__ten_dvkt').annotate(tong_tien=Sum('dich_vu_kham__don_gia')).order_by('dich_vu_kham__ten_dvkt').annotate(dich_vu_kham_count = Count('dich_vu_kham__ten_dvkt')).annotate(ma_dvkt=F('dich_vu_kham__ma_dvkt')).annotate(don_gia=F('dich_vu_kham__don_gia')).annotate(stt=F('dich_vu_kham__stt'))
            
            list_tong_tien_formatted = ["{:,}".format(int(i['tong_tien'])) if i['tong_tien'] is not None else 0 for i in dich_vu]

            list_dich_vu = []
            for idx, val in enumerate(dich_vu):
                val['tong_tien'] = list_tong_tien_formatted[idx]
                list_dich_vu.append(val)

        xlsx_data = writeToExcelDichVu(list_dich_vu)
        response = HttpResponse(
            content_type='application/ms-excel'
        )
        response['Content-Disposition'] = 'attachment; filename="Report.xlsx"'
        response.write(xlsx_data)
       
    else:
        response = HttpResponse(json.dumps({'message': "It's not gonna happen"}), content_type='application/json; charset=utf-8')
    return response       

@csrf_exempt
def export_thuoc_bao_hiem_excel(request):
    if request.method == 'POST':
        startDate = request.POST.get('startDate')
        endDate = request.POST.get('endDate')
        start = datetime.strptime(startDate, "%d-%m-%Y")
        tomorrow_start = start + timedelta(1)

        if endDate == '':
            thuoc = KeDonThuoc.objects.filter(thoi_gian_tao__lt=tomorrow_start, thoi_gian_tao__gte=start).values('thuoc__ten_thuoc').annotate(tong_tien=Sum('thuoc__don_gia_tt')).order_by('thuoc__ten_thuoc').annotate(thuoc_count = Count('thuoc__ten_thuoc')).annotate(ten_hoat_chat=F('thuoc__ten_hoat_chat')).annotate(duong_dung=F('thuoc__duong_dung')).annotate(ham_luong=F('thuoc__ham_luong')).annotate(so_dang_ky=F('thuoc__so_dang_ky')).annotate(don_vi_tinh=F('thuoc__don_vi_tinh')).annotate(don_gia=F('thuoc__don_gia_tt'))
            
            list_thuoc = []
            for i in thuoc:
                list_thuoc.append(i)
            response = {
                'data' : list_thuoc,
            }
        else:
            end = datetime.strptime(endDate, "%d-%m-%Y")
            if startDate == endDate:
                end = end + timedelta(1)
            thuoc = KeDonThuoc.objects.filter(thoi_gian_tao__lt=end, thoi_gian_tao__gte=start).values('thuoc__ten_thuoc').annotate(tong_tien=Sum('thuoc__don_gia_tt')).order_by('thuoc__ten_thuoc').annotate(thuoc_count = Count('thuoc__ten_thuoc')).annotate(ten_hoat_chat=F('thuoc__ten_hoat_chat')).annotate(duong_dung=F('thuoc__duong_dung')).annotate(ham_luong=F('thuoc__ham_luong')).annotate(so_dang_ky=F('thuoc__so_dang_ky')).annotate(don_vi_tinh=F('thuoc__don_vi_tinh')).annotate(don_gia=F('thuoc__don_gia_tt')).annotate(ma_thuoc=F('thuoc__ma_thuoc'))
            list_tong_tien_formatted = ["{:,}".format(int(i['tong_tien'])) for i in thuoc]

            list_thuoc = []
            for idx, val in enumerate(thuoc):
                val['tong_tien'] = list_tong_tien_formatted[idx]
                list_thuoc.append(val)

        xlsx_data = writeToExcelThuoc(list_thuoc)
        response = HttpResponse(
            content_type='application/ms-excel'
        )
        response['Content-Disposition'] = 'attachment; filename="Report.xlsx"'
        response.write(xlsx_data)
       
    else:
        response = HttpResponse(json.dumps({'message': "It's not gonna happen"}), content_type='application/json; charset=utf-8')
    return response  

def ket_qua_benh_nhan_view(request):
    phong_chuc_nang = PhongChucNang.objects.all()

    data = {
        'phong_chuc_nang' : phong_chuc_nang,
    }
    return render(request, 'le_tan/ket_qua_benh_nhan.html', context = data)

def chi_tiet_chuoi_kham_benh_nhan(request, **kwargs):
    id_chuoi_kham = kwargs.get('id_chuoi_kham')
    chuoi_kham = ChuoiKham.objects.filter(id=id_chuoi_kham).first()
    if chuoi_kham is not None:
        benh_nhan = chuoi_kham.benh_nhan
    else:
        return render(request, '404.html')
   
    ket_qua_tong_quat = chuoi_kham.ket_qua_tong_quat.all().first()
    if ket_qua_tong_quat is not None:
        ket_qua_chuyen_khoa = ket_qua_tong_quat.ket_qua_chuyen_khoa.all()

        context = {
            'benh_nhan': benh_nhan,
            'chuoi_kham': chuoi_kham,
            'ket_qua_tong_quat': ket_qua_tong_quat,
            'ket_qua_chuyen_khoa': ket_qua_chuyen_khoa,
        }
        return render(request, 'le_tan/chi_tiet_ket_qua_chuoi_kham.html', context=context)
    else:
        context = {
            'benh_nhan': benh_nhan,
            'chuoi_kham': chuoi_kham,
        }
        return render(request, 'le_tan/chi_tiet_ket_qua_chuoi_kham.html', context=context)

def xoa_chuoi_kham(request):
    id_chuoi_kham = request.POST.get('id_chuoi_kham')
    chuoi_kham = ChuoiKham.objects.filter(id=id_chuoi_kham).first()
    chuoi_kham.delete()

    response = {
        'message' : "Xóa Chuỗi Khám Thành Công"
    }
    return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")

def xoa_mau_phieu(request):
    id = request.POST.get('id')
    mau_phieu = MauPhieu.objects.filter(id=id).first()
    mau_phieu.delete()

    response = {
        'message' : "Xóa Mẫu Phiếu Thành Công"
    }

    return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")

def xoa_bai_dang(request):
    id = request.POST.get('id')
    bai_dang = BaiDang.objects.filter(id=id).first()
    bai_dang.delete()
    response = {
        'message' : "Xóa Mẫu Phiếu Thành Công"
    }
    return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")

def chinh_sua_bai_dang_view(request, **kwargs):
    id = kwargs.get('id')
    instance = get_object_or_404(BaiDang, id=id)
    form = BaiDangForm(request.POST or None, instance=instance)
    context = {
        'form': form,
        'id': id
    }

    return render(request, 'le_tan/chinh_sua_bai_dang.html', context=context)

def xoa_tieu_chuan(request, **kwargs):
    id  = kwargs.get('id')
    dich_vu = get_object_or_404(DichVuKham, id=id)
    dich_vu.chi_so = False
    dich_vu.save()

    # chi_so_xet_nghiem = dich_vu.chi_so_xet_nghiem.all()
    response = {
        'message' : "Xóa Tiêu Chuẩn DVKT Thành Công"
    }
    return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")

def xoa_chi_so(request):
    id = request.POST.get('id')
    try:
        chi_so = ChiSoXetNghiem.objects.get(id=id)
        chi_so.delete()
        response = {
            'status': 200,
            'message': "Tiêu chuẩn dịch vụ kĩ thuật đã xóa thành công."
        }
    except ChiSoXetNghiem.DoesNotExist: 
        response = {
            'status': 404,
            'message': "Tiêu chuẩn này không tồn tại"
        }
    return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")

import string
import random

LETTERS = string.ascii_letters
NUMBERS = string.digits  
PUNCTUATION = string.punctuation    

def password_generator(length=8):
    '''
    Generates a random password having the specified length
    :length -> length of password to be generated. Defaults to 8
        if nothing is specified.
    :returns string <class 'str'>
    '''
    # create alphanumerical from string constants
    printable = f'{LETTERS}{NUMBERS}'

    # convert printable from string to list and shuffle
    printable = list(printable)
    random.shuffle(printable)

    # generate random password and convert to string
    random_password = random.choices(printable, k=length)
    random_password = ''.join(random_password)
    return random_password

def resetPassword(request):
    id = request.POST.get('id')
    try:
        import time
        time.sleep(2)
        user = User.objects.get(id=id)
        randPassword = password_generator(12)
        user.set_password(randPassword)
        user.save()
        response = {
            'status': 200,
            'message': "Đặt lại mật khẩu thành công",
            'user': user.ho_ten,
            'password': randPassword,
        }
    except User.DoesNotExist:
        response = {
            'status': 404,
            'message': "Không tồn tại bệnh nhân này"
        }
    return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")
    
def validatePassword(password):
    import re 
    while True:   
        if (len(password)<8): 
            return False
        elif not re.search("[a-z]", password): 
            return False
        elif not re.search("[0-9]", password): 
            return False
        elif re.search("\s", password): 
            return False
        else: 
            return True

def changePassword(request):
    if request.method == "POST":
        id = request.POST.get('id')
        current_password = request.POST.get('current_password')
        new_password = request.POST.get('new_password')
        confirm_password = request.POST.get('confirm_password')
        try:
            user = User.objects.get(id=id)
            if not user.check_password(current_password):
                response = {
                    'status': 400,
                    'message': 'Mật khẩu hiện tại không đúng, vui lòng thử lại'
                }
                return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")
            
            elif new_password != confirm_password:
                response = {
                    'status': 400,
                    'message': 'Mật khẩu hiện tại không đúng, vui lòng thử lại'
                }
                return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")

            elif not validatePassword(new_password):
                response = {
                    'status': 400,
                    'message': 'Mật khẩu mới không đủ mạnh'
                }
                return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")

            else:
                user.set_password(new_password)
                user.save()

        except User.DoesNotExist:
            response = {
                'status': 400,
                'message': 'Người dùng không tồn tại, vui lòng kiểm tra lại'
            }
            return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")

        response = {
            'status': 200,
            'message': 'Cập nhật mật khẩu thành công'
        }
    else:
        response = {
            'status': 400,
            'message': 'Không thể đổi được mật khẩu, vui lòng kiểm tra lại'
        }
    return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")

def chinh_sua_bai_dang(request):
    if request.method == "POST":
        id_bai_dang = request.POST.get('id')
        tieu_de = request.POST.get('tieu_de')
        noi_dung_chinh = request.POST.get('noi_dung_chinh')
        thoi_gian_bat_dau = request.POST.get("thoi_gian_bat_dau")
        thoi_gian_ket_thuc = request.POST.get('thoi_gian_ket_thuc')
        hinh_anh = request.FILES.get('hinh_anh', None)
        noi_dung = request.POST.get('noi_dung')

        thoi_gian_bat_dau = datetime.strptime(thoi_gian_bat_dau, format_2)
        thoi_gian_bat_dau = thoi_gian_bat_dau.strftime("%Y-%m-%d %H:%M")

        thoi_gian_ket_thuc = datetime.strptime(thoi_gian_ket_thuc, format_2)
        thoi_gian_ket_thuc = thoi_gian_ket_thuc.strftime("%Y-%m-%d %H:%M")

        try:
            bai_dang = BaiDang.objects.get(id=id_bai_dang)
            if hinh_anh is not None:
                bai_dang.tieu_de = tieu_de
                bai_dang.noi_dung_chinh = noi_dung_chinh
                bai_dang.thoi_gian_bat_dau = thoi_gian_bat_dau
                bai_dang.thoi_gian_ket_thuc = thoi_gian_ket_thuc
                bai_dang.hinh_anh = hinh_anh
                bai_dang.noi_dung = noi_dung
            else:
                bai_dang.tieu_de = tieu_de
                bai_dang.noi_dung_chinh = noi_dung_chinh
                bai_dang.thoi_gian_bat_dau = thoi_gian_bat_dau
                bai_dang.thoi_gian_ket_thuc = thoi_gian_ket_thuc
                bai_dang.noi_dung = noi_dung

            bai_dang.save()
            response = {
                'status': 200,
                'message': 'Cập nhật bài đăng thành công'
            }

            from actstream import action
            action.send(request.user, verb='đã cập nhật bài đăng')

        except BaiDang.DoesNotExist:
            response = {
                'status': 404,
                'message': 'Không tìm thấy bài đăng'
            }
    else:
        response = {
            'status': 404,
            'message': 'Không cập nhật được bài đăng'
        }
    return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")

def view_danh_sach_nhom_quyen(request):
    import operator
    from functools import reduce
    default_perms_name = ['Can view', 'Can change', 'Can delete', 'Can add']
    query = reduce(operator.or_, (Q(name__contains=x) for x in default_perms_name))
    permissions = Permission.objects.exclude(query)
    context = {
        'permissions': permissions
    }
    return render(request, 'danh_sach_nhom_quyen.html', context)

def view_chinh_sua_nhom_quyen(request, **kwargs):
    import operator
    from functools import reduce
    id_nhom_quyen = kwargs.get('id')
    try:
        group = Group.objects.get(id=id_nhom_quyen)
        group_permissions = group.permissions.all()
        group_permissions_str = [p.name for p in group_permissions]
        code_name_permissions = [p.codename for p in group_permissions]
        default_perms_name = ['Can view', 'Can change', 'Can delete', 'Can add']
        query = reduce(operator.or_, (Q(name__contains=x) for x in default_perms_name))
        permissions = Permission.objects.exclude(query)
        permissions = PermissionSerializer(permissions, many=True)
        permissions_data = json.loads(json.dumps(permissions.data))
        context = {
            'group_permissions': group_permissions_str,
            'code_name_permissions': code_name_permissions,
            'group': group,
            'permissions': permissions_data,

        }
        return render(request, 'chinh_sua_nhom_quyen.html', context)
    except Group.DoesNotExist:
        return render(request, '404.html')

def view_danh_sach_nhan_vien(request):
    province = Province.objects.all()
    groups = Group.objects.all()
    groups_serializer = GroupSerializer(groups, many=True)
    groups_data = json.loads(json.dumps(groups_serializer.data))
    phong_chuc_nang = PhongChucNang.objects.all()
    context = {
        'province': province,
        'groups': groups_data,
        'phong_chuc_nang': phong_chuc_nang,
    }
    return render(request, 'danh_sach_nhan_vien.html', context)

def store_nhom_quyen(request):
    if request.method == "POST":
        ten_nhom_quyen = request.POST.get("ten_nhom_quyen")
        danh_sach_quyen = request.POST.get("danh_sach_quyen")
        groups = json.loads(danh_sach_quyen)

        if ten_nhom_quyen == "":
            response = {
                'status': 404,
                'message': 'Tên nhóm quyền không được để trống'
            }
            return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")

        time.sleep(1)
        new_group, created = Group.objects.get_or_create(name=ten_nhom_quyen)
        for p in groups:
            permission = Permission.objects.get(codename=p)
            new_group.permissions.add(permission)
            
        response = {
            'status': 200,
            'message': 'Thêm nhóm quyền thành công'
        }
    else:
        response = {
            'status': 404,
            'message': 'Không lưu được nhóm quyền'
        }
    return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")

def update_nhom_quyen(request):
    if request.method == "POST":
        id = request.POST.get('id')
        ten_nhom_quyen = request.POST.get('ten_nhom_quyen')
        danh_sach_quyen = request.POST.get("danh_sach_quyen")
        group_permissions = json.loads(danh_sach_quyen)

        if ten_nhom_quyen == "":
            response = {
                'status': 404,
                'message': 'Tên nhóm quyền không được để trống'
            }
            return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")

        try:
            group = Group.objects.get(id=id)
            group.name = ten_nhom_quyen

            group.permissions.clear()

            for codename in group_permissions:
                permission = Permission.objects.get(codename=codename)
                group.permissions.add(permission)

            group.save()

        except Group.DoesNotExist:
            response = {
                'status': 404,
                'message': 'Nhóm quyền không tồn tại'
            }
            return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")
        response = {
            'status': 200,
            'message': 'Cập nhật nhóm quyền thành công'
        }
    else:
        response = {
            'status': 404,
            'message': 'Không lưu được nhóm quyền'
        }
    return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")

def store_uy_quyen(request):
    if request.user.is_superuser or request.user.is_admin:
        if request.method == "POST":
            user_id = request.POST.get('user_id')
            list_groups_id = request.POST.get('list_groups_id')
            groups_id = json.loads(list_groups_id)
            print('---1---')
            print(groups_id)
            try:
                user = User.objects.get(id=user_id)
                print('---2---')
                print(user.ho_ten)
                user.groups.clear()
                print(user.groups.all())
                for id in groups_id:
                    group = Group.objects.get(id=id)
                    user.groups.add(group)
                print(user.groups.all())
            except User.DoesNotExist:
                response = {
                    'status': 404,
                    'message': 'Người dùng không tồn tại'
                }
                return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")

            response = {
                'status': 200,
                'message': 'Ủy quyền thành công'
            }
        else:
            response = {
                'status': 404,
                'message': 'Ủy quyền cho nhân viên không thành công'
            }
    else:
        response = {
            'status': 404,
            'message': 'Bạn không có quyền ủy quyền'
        }
    return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")
    
def create_staff_user(request):
    if request.user.is_superuser or request.user.is_admin:
        if request.method == "POST":
            ho_ten         = request.POST.get('ho_ten')
            username       = request.POST.get('username')
            so_dien_thoai  = request.POST.get('so_dien_thoai')
            password       = request.POST.get("password")
            cmnd_cccd      = request.POST.get('cmnd_cccd')
            ngay_sinh      = request.POST.get('ngay_sinh')
            gioi_tinh      = request.POST.get('gioi_tinh')
            dia_chi        = request.POST.get('dia_chi')
            chuc_nang = request.POST.get('chuc_nang')
            tinh_id = request.POST.get('tinh')
            huyen_id = request.POST.get('huyen')
            xa_id = request.POST.get('xa')

            if User.objects.filter(username=username).exists():
                response = {
                    'status': 400,
                    'message': "Username đã tồn tại, vui lòng chọn username mới"
                }
                return HttpResponse(
                    json.dumps(response),
                    content_type="application/json"
                )
            elif User.objects.filter(so_dien_thoai=so_dien_thoai).exists():
                response = {
                    'status': 400,
                    'message': "Số điện thoại đã tồn tại, vui lòng chọn số điện thoại mới"
                }
                return HttpResponse(
                    json.dumps(response),
                    content_type="application/json"
                )
            elif User.objects.filter(cmnd_cccd=cmnd_cccd).exists():
                response = {
                    'status': 400,
                    'message': "CMND đã tồn tại, vui lòng chọn CMND mới"
                }
                return HttpResponse(
                    json.dumps(response),
                    content_type="application/json"
                )

            if dia_chi == '':
                response = {
                    'status': 400,
                    'message': "Thiếu địa chỉ cụ thể"
                }
                return HttpResponse(
                    json.dumps(response),
                    content_type="application/json"
                )

            print(chuc_nang)

            benh_nhan = User.objects.create_staffuser(
                ho_ten         = ho_ten,
                username       = username, 
                so_dien_thoai  = so_dien_thoai, 
                password       = password,
                cmnd_cccd      = cmnd_cccd,
                gioi_tinh      = gioi_tinh,
            )

            benh_nhan.chuc_nang = chuc_nang

            if ngay_sinh != '':
                ngay_sinh = datetime.strptime(ngay_sinh, format_3)
                ngay_sinh = ngay_sinh.strftime("%Y-%m-%d")
                benh_nhan.ngay_sinh = ngay_sinh

            if dia_chi != "":
                benh_nhan.dia_chi = dia_chi
            
            if tinh_id != 'null':      
                tinh = Province.objects.filter(id=tinh_id).first()
                benh_nhan.tinh = tinh
            else:
                response = {
                    'status': 400,
                    'message': "Không thể thiếu tỉnh thành"
                }
                return HttpResponse(
                    json.dumps(response),
                    content_type="application/json"
                )
       
            if huyen_id != 'null':
                huyen = District.objects.filter(id=huyen_id).first()
                benh_nhan.huyen = huyen
            else:
                response = {
                    'status': 400,
                    'message': "Không thể thiếu quận huyện"
                }
                return HttpResponse(
                    json.dumps(response),
                    content_type="application/json"
                )

            
            if xa_id != 'null': 
                xa = Ward.objects.filter(id=xa_id).first()  
                benh_nhan.xa = xa
            else:
                response = {
                    'status': 400,
                    'message': "Không thể thiếu phường xã"
                }
                return HttpResponse(
                    json.dumps(response),
                    content_type="application/json"
                )

            benh_nhan.save()

            response = {
                'status': 200,
                'message': "Thêm nhân viên thành công"
            }

            return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")
        else:
            response = {
                'status': 400,
                'message': "Xảy ra lỗi, không thể thêm nhân viên"
            }
            return HttpResponse(
                json.dumps(response),
                content_type="application/json"
            )
    else:
        response = {
            'status': 400,
            'message': "Bạn Không Có Quyền Thêm Nhân Viên"
        }
        return HttpResponse(
            json.dumps(response),
            content_type="application/json", charset="utf-8"
        )

def update_nhan_vien(request, **kwargs):
    id = kwargs.get('id')
    instance = get_object_or_404(User, id=id)
    form = UserForm(request.POST or None, instance=instance)
    phong_chuc_nang = PhongChucNang.objects.all()
    provinces = Province.objects.all()
    data = {
        'benh_nhan': instance,
        'form': form,
        'id': id,
        'phong_chuc_nang': phong_chuc_nang,
        'province': provinces
    }
    return render(request, 'update_nhan_vien.html', context=data)

def cap_nhat_thong_tin_nhan_vien(request):
    if request.method == "POST":
        id             = request.POST.get('id')
        ho_ten         = request.POST.get('ho_ten')
        so_dien_thoai  = request.POST.get('so_dien_thoai')
        username       = request.POST.get('username', None)
        ngay_sinh      = request.POST.get('ngay_sinh')
        cmnd_cccd      = request.POST.get('cmnd_cccd')
        dia_chi        = request.POST.get('dia_chi')

        tinh_id = request.POST.get('tinh')      
        tinh = Province.objects.filter(id=tinh_id).first()       
        huyen_id = request.POST.get('huyen')       
        huyen = District.objects.filter(id=huyen_id).first()
        xa_id = request.POST.get('xa')
        xa = Ward.objects.filter(id=xa_id).first()

        ngay_sinh = datetime.strptime(ngay_sinh, format_3)
        ngay_sinh = ngay_sinh.strftime("%Y-%m-%d")
    
        nhan_vien = get_object_or_404(User, id=id)
        nhan_vien.ho_ten            = ho_ten
        nhan_vien.so_dien_thoai     = so_dien_thoai
        nhan_vien.username          = username
        nhan_vien.cmnd_cccd         = cmnd_cccd
        nhan_vien.dia_chi           = dia_chi
        nhan_vien.ngay_sinh         = ngay_sinh

        nhan_vien.tinh = tinh
        nhan_vien.huyen = huyen
        nhan_vien.xa = xa

        nhan_vien.save()

        response = {
            'status': 200,
            'message': 'Cập Nhật Thông Tin Thành Công'
        }
    else:
        response = {
            'status': 404,
            'message': 'Có lỗi xảy ra'
        }
    return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")

def update_staff_user(request):
    if request.user.is_superuser or request.user.is_admin:
        if request.method == "POST":
            pass


def import_dich_vu_kham_view(request):
    import openpyxl
    if request.method == "GET":
        return render(request, 'phong_tai_chinh/import_dich_vu_kham.html')
    else:
        excel_file = request.FILES["excel_file"]
        wb = openpyxl.load_workbook(excel_file)
        worksheet = wb["Sheet1"]

        excel_data = list()
        result = []

        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)
            # excel_data.pop(0)
        list_title = excel_data[0]
        excel_data.pop(0)

        for row in excel_data:
            res = dict(zip(list_title, row))
            result.append(res)

        response = {
            'status': 200,
            'data': result
        }
        return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")

def import_thuoc_view(request):
    import openpyxl
    if request.method == "GET":
        return render(request, 'phong_tai_chinh/import_thuoc.html')
    else:
        excel_file = request.FILES["excel_file"]
        wb = openpyxl.load_workbook(excel_file)
        worksheet = wb["Sheet1"]

        excel_data = list()
        result = []

        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)
            # excel_data.pop(0)
        list_title = excel_data[0]
        excel_data.pop(0)

        for row in excel_data:
            res = dict(zip(list_title, row))
            result.append(res)

        response = {
            'status': 200,
            'data': result
        }
        return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")

def store_dich_vu_kham_excel(request):
    import openpyxl
    if request.method == "POST":
        excel_file = request.FILES["excel_file"]
        wb = openpyxl.load_workbook(excel_file)
        worksheet = wb["Sheet1"]

        excel_data = []
        for row in worksheet.iter_rows():
            row_data = []
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)
        list_title = excel_data[0]
        excel_data.pop(0)

        bulk_create_data = []
        for row in excel_data:
            res = dict(zip(list_title, row))
            stt = res['STT']
            ma_dvkt = res['MA_DVKT']
            ten_dvkt = res['TEN_DVKT']
            don_gia_bhyt = res['DON_GIA_BHYT']
            don_gia = res['DON_GIA']
            quyet_dinh = res['QUYET_DINH']
            cong_bo = res['CONG_BO']
            ma_gia = res['MA_GIA']
            phong_chuc_nang = res['PHONG_CHUC_NANG']
            ma_cosokcb = res['MA_COSOKCB']
            bao_hiem = True

            if don_gia is not None:
                don_gia = int(don_gia)
            else:
                don_gia = 0

            if don_gia_bhyt is not None:
                don_gia_bhyt = int(don_gia_bhyt)
            else:
                don_gia_bhyt = 0

            content_type = ContentType.objects.get_for_model(PhongChucNang)

            group_phong_chuc_nang = PhongChucNang.objects.get_or_create(ten_phong_chuc_nang = phong_chuc_nang)[0]
            group_phong_chuc_nang.save()

            new_group, created = Group.objects.get_or_create(name=f"Nhóm {phong_chuc_nang}")
            codename_perm = f'can_view_{group_phong_chuc_nang.slug}'

            if not Permission.objects.filter(codename = codename_perm).exists():
                permission = Permission.objects.create(
                    codename=codename_perm,
                    name = f'Xem {phong_chuc_nang}',
                    content_type=content_type
                )
                new_group.permissions.add(permission)

            model = DichVuKham(
                ma_dvkt         = ma_dvkt,
                ten_dvkt        = ten_dvkt,
                ma_gia          = ma_gia,
                don_gia         = Decimal(don_gia),
                don_gia_bhyt = Decimal(don_gia_bhyt),
                bao_hiem        = bao_hiem,
                quyet_dinh      = quyet_dinh,
                cong_bo         = cong_bo,
                ma_cosokcb      = ma_cosokcb,
                phong_chuc_nang = group_phong_chuc_nang
            )
            bulk_create_data.append(model)
        DichVuKham.objects.bulk_update_or_create(bulk_create_data, [
            'stt',
            'ma_dvkt',
            'ten_dvkt',
            'ma_gia',
            'don_gia',
            'don_gia_bhyt',
            'bao_hiem',
            'quyet_dinh',
            'cong_bo',
            'ma_cosokcb',
            'phong_chuc_nang' 
        ], match_field = 'ma_dvkt', batch_size=10)

        response = {
            'status': 200,
            'message': 'Thêm Dịch Vụ Khám Thành Công',
            'url' : '/danh_sach_dich_vu_kham/'
        }
        return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")
    else:
        response = {
            'status': 404,
            'message': 'Thêm Dịch Vụ Khám Thất Bại',
        }
        return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")

def store_thuoc_excel(request):
    import openpyxl
    if request.method == "POST":
        excel_file = request.FILES["excel_file"]
        wb = openpyxl.load_workbook(excel_file)
        worksheet = wb["Sheet1"]

        excel_data = []
        for row in worksheet.iter_rows():
            row_data = []
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)
        list_title = excel_data[0]
        excel_data.pop(0)

        bulk_create_data = []
        for row in excel_data:
            res = dict(zip(list_title, row))
            stt = res['STT']
            ma_hoat_chat = res['MA_HOAT_CHAT']
            hoat_chat = res['HOAT_CHAT']
            ma_duong_dung = res['MA_DUONG_DUNG']
            duong_dung = res['DUONG_DUNG']
            ham_luong = res['HAM_LUONG']
            ma_thuoc = res['MA_THUOC']
            ten_thuoc = res['TEN_THUOC']
            so_dang_ky = res['SO_DANG_KY']
            dong_goi = res['DONG_GOI']
            don_vi_tinh = res['DON_VI_TINH']
            don_gia = res['DON_GIA']
            don_gia_tt = res['DON_GIA_TT']
            
            ma_cskcb = res['MA_CSKCB']
            hang_sx = res['HANG_SX']
            nuoc_sx = res['NUOC_SX']
            nha_thau = res['NHA_THAU']
            quyet_dinh = res['QUYET_DINH']
            cong_bo = res['CONG_BO']
            loai_thuoc = res['LOAI_THUOC']
            loai_thau = res['LOAI_THAU']
            nhom_thau = res['NHOM_THAU']
            bao_hiem = True

            group_nhom_thau = NhomThau.objects.get_or_create(ten_nhom_thau=nhom_thau)[0]
            group_cong_ty = CongTy.objects.get_or_create(ten_cong_ty=nha_thau)[0] 
                    
            model = Thuoc(
                stt = stt,
                ma_thuoc          = ma_thuoc,
                ma_hoat_chat      = ma_hoat_chat, 
                ten_hoat_chat     = hoat_chat, 
                duong_dung        = duong_dung,
                ham_luong         = ham_luong,
                ten_thuoc         = ten_thuoc,
                so_dang_ky        = so_dang_ky, 
                dong_goi          = dong_goi,
                don_vi_tinh       = don_vi_tinh,
                don_gia           = Decimal(don_gia),
                don_gia_tt        = Decimal(don_gia_tt),
                so_lo             = "",
                so_luong_kha_dung = 0,
                ma_cskcb          = ma_cskcb, 
                hang_sx           = hang_sx,
                nuoc_sx           = nuoc_sx,
                quyet_dinh        = quyet_dinh, 
                loai_thuoc        = loai_thuoc, 
                cong_bo           = cong_bo,
                loai_thau         = loai_thau,
                nhom_thau         = group_nhom_thau,
                cong_ty           = group_cong_ty,
                bao_hiem          = bao_hiem,
            )

            bulk_create_data.append(model)

        Thuoc.objects.bulk_update_or_create(bulk_create_data, [
            'stt',
            'ma_thuoc',
            'ma_hoat_chat', 
            'ten_hoat_chat', 
            'duong_dung', 
            'ham_luong',
            'ma_thuoc',
            'ten_thuoc', 
            'so_dang_ky', 
            'dong_goi', 
            'don_vi_tinh', 
            'don_gia', 
            'don_gia_tt',
            'so_luong_kha_dung',
            'ma_cskcb',
            'hang_sx',
            'nuoc_sx',
            'quyet_dinh',
            'cong_bo',
            'loai_thau',
            'nhom_thau', 

        ], match_field = 'ma_thuoc', batch_size=10)

        response = {
            'status': 200,
            'message': 'Thêm Thuốc Thành Công',
            'url' : '/phong_tai_chinh/danh_sach_thuoc/'
        }
        return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")
    else:
        response = {
            'status': 404,
            'message': 'Thêm Thuốc Thất Bại',
        }
        return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")

def store_thuoc_dich_vu_excel(request):
    import openpyxl
    import dateutil.parser
    if request.method == "POST":
        excel_file = request.FILES["excel_file"]
        wb = openpyxl.load_workbook(excel_file)
        worksheet = wb["Sheet1"]

        excel_data = []
        for row in worksheet.iter_rows():
            row_data = []
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)
        list_title = excel_data[0]
        excel_data.pop(0)

        bulk_create_data = []
        bulk_create_data_nhom_thuoc = []

        for row in excel_data:
            res = dict(zip(list_title, row))

            thuoc_data = {}
            thuoc_data = res.copy()
            del thuoc_data['NHOM_THUOC']

            nhom_thuoc = res['NHOM_THUOC']
        
            if nhom_thuoc not in bulk_create_data_nhom_thuoc:
                bulk_create_data_nhom_thuoc.append(nhom_thuoc)

            if thuoc_data not in bulk_create_data:
                bulk_create_data.append(thuoc_data)

        list_model_nhom_thuoc = []
        list_model_thuoc = []

        for data in bulk_create_data_nhom_thuoc:
            model = NhomThuoc(
                ten_nhom=data
            )
            list_model_nhom_thuoc.append(model)

        for data in bulk_create_data:
            group_cong_ty = CongTy.objects.get_or_create(ten_cong_ty=data['CONG_TY'])[0]
            ngay_san_xuat = data['NGAY_SAN_XUAT']
            ngay_san_xuat = dateutil.parser.parse(ngay_san_xuat)
            han_su_dung = data['HAN_SU_DUNG']
            han_su_dung = dateutil.parser.parse(han_su_dung)
            model = Thuoc(
                ma_thuoc = data['MA_THUOC'],
                ten_thuoc = data['TEN_THUOC'],
                ten_hoat_chat = data['HOAT_CHAT'], 
                ham_luong = data['HAM_LUONG'],
                so_dang_ky = data['SO_DANG_KY'],
                nuoc_sx = data['NUOC_SX'], 
                dong_goi = data['DONG_GOI'],
                don_gia = Decimal(data['DON_GIA']),
                don_gia_tt = Decimal(data['DON_GIA_TT']),
                don_vi_tinh = data['DON_VI_TINH'],
                so_lo = data['SO_LO'],
                duong_dung = data['DUONG_DUNG'],
                ngay_san_xuat = ngay_san_xuat,
                han_su_dung = han_su_dung,
                cong_ty = group_cong_ty,
            )
            list_model_thuoc.append(model)

        NhomThuoc.objects.bulk_create(list_model_nhom_thuoc, batch_size=10)

        Thuoc.objects.bulk_update_or_create(list_model_thuoc, [
            'ma_thuoc',
            'ten_thuoc',
            'ten_hoat_chat',
            'ham_luong',
            'so_dang_ky',
            'nuoc_sx', 
            'dong_goi',
            'don_gia',
            'don_gia_tt', 
            'don_vi_tinh',
            'so_lo',
            'duong_dung',
            'ngay_san_xuat',
            'han_su_dung',
            'cong_ty',
        ], match_field = 'ma_thuoc', batch_size=50)

        for row in excel_data:
            res = dict(zip(list_title, row))

            nhom_thuoc = res['NHOM_THUOC']
            ma_thuoc = res['MA_THUOC']

            group_nhom_thuoc = NhomThuoc.objects.filter(ten_nhom=nhom_thuoc).first()
            thuoc = Thuoc.objects.filter(ma_thuoc=ma_thuoc).first()
            group_nhom_thuoc.nhom_thuoc.add(thuoc)

        response = {
            'status': 200,
            'message': 'Thêm Thuốc Thành Công',
            'url' : '/phong_tai_chinh/danh_sach_thuoc/'
        }
    else:
        response = {
            'status': 404,
            'message': 'Thêm Thuốc Thất Bại',
        }
    return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")

def nhap_them_thuoc(request):
    phong_chuc_nang = PhongChucNang.objects.all()
    user_id = request.user.id
    data = {
        'phong_chuc_nang' : phong_chuc_nang,
        'user_id' : user_id
    }
    return render(request, 'phong_tai_chinh/nhap_them_thuoc.html', context = data)

def store_nhap_thuoc(request):
    if request.method == "POST":
        request_data = request.POST.get('data', None)
        user = request.POST.get('user', None)
        data = json.loads(request_data)

        now = datetime.now()
        date_time = now.strftime("%m%d%y%H%M%S")

        user = User.objects.get(id=user)
        subName = getSubName(user.ho_ten)
        ma_hoa_don = subName + '-' + date_time
        hoa_don_nhap = HoaDonNhapHang.objects.get_or_create(nguoi_phu_trach=user, ma_hoa_don=ma_hoa_don)[0]

        bulk_create_data = []
        for i in data:
            so_luong = i['obj']['so_luong']
            id = i['obj']['id']

            thuoc = Thuoc.objects.filter(id=id)

            if thuoc[0].so_luong_kha_dung is not None:
                thuoc.update(so_luong_kha_dung=F('so_luong_kha_dung') + so_luong)
            else:
                thuoc.update(so_luong_kha_dung = so_luong)

            ThuocLog.objects.create(thuoc=thuoc[0], ngay=timezone.now(), quy_trinh=ThuocLog.IN, so_luong=so_luong)

            nhap_hang = NhapHang(hoa_don=hoa_don_nhap, thuoc=thuoc[0], so_luong=i['obj']['so_luong'], bao_hiem=i['obj']['bao_hiem'])

            bulk_create_data.append(nhap_hang)

        NhapHang.objects.bulk_create(bulk_create_data)

        response = {'status': 200, 'message': 'Nhập Hàng Thành Công'}
    else:
       response = {'message': 'oke'} 
    return HttpResponse(json.dumps(response), content_type='application/json; charset=utf-8')

def bao_cao_thuoc(request):
    phong_chuc_nang = PhongChucNang.objects.all()

    data = {
        'phong_chuc_nang' : phong_chuc_nang,
    }
    return render(request, 'phong_tai_chinh/bao_cao_thuoc.html', context=data)

def xuat_bao_cao_ton(request, *args, **kwargs):
    if request.method == "POST":
        range_start = request.POST.get('range_start')
        range_end = request.POST.get('range_end')

        start = datetime.strptime(range_start, "%d-%m-%Y")
        tomorrow_start = start + timedelta(1)

        if range_end == '':
            tong_nhap_hang = NhapHang.objects.filter(thoi_gian_tao__gte=start, thoi_gian_tao__lt=tomorrow_start).exclude(bao_hiem=False).values("thuoc__ten_thuoc").annotate(id=F('thuoc__id')).annotate(so_luong=Sum('so_luong')).annotate(c = Count('thuoc__ten_thuoc')).annotate(bao_hiem=F('bao_hiem'))

            tong_xuat_hang = KeDonThuoc.objects.filter(thoi_gian_tao__gte=start, thoi_gian_tao__lt=tomorrow_start).exclude(bao_hiem=False).values('thuoc__ten_thuoc').annotate(id=F('thuoc__id')).annotate(so_luong=Sum('so_luong')).annotate(c = Count('thuoc__ten_thuoc')).annotate(bao_hiem=F('bao_hiem'))

            list_ton = []

            for i in tong_nhap_hang:
                for j in tong_xuat_hang:
                    if j['id'] in i.values():
                        so_luong_nhap = i['so_luong']
                        so_luong_xuat = j['so_luong']
                        so_luong_ton = so_luong_nhap - so_luong_xuat
                        ten_thuoc = j['thuoc__ten_thuoc']
                        d = {}
                        d['so_luong'] = so_luong_ton
                        d['ten_thuoc'] = ten_thuoc
                        list_ton.append(d)

            response = {
                'list_ton' : list_ton,
            }
            return HttpResponse(json.dumps(response), content_type='application/json; charset=utf-8')
        
        else:
            end = datetime.strptime(range_end, "%d-%m-%Y")
            tong_nhap_hang = NhapHang.objects.filter(thoi_gian_tao__gte=start, thoi_gian_tao__lt=end).exclude(bao_hiem=False).values("thuoc__ten_thuoc").annotate(id=F('thuoc__id')).annotate(so_luong=Sum('so_luong')).annotate(c = Count('thuoc__ten_thuoc')).annotate(bao_hiem=F('bao_hiem'))

            tong_xuat_hang = KeDonThuoc.objects.filter(thoi_gian_tao__gte=start, thoi_gian_tao__lt=end).exclude(bao_hiem=False).values('thuoc__ten_thuoc').annotate(id=F('thuoc__id')).annotate(so_luong=Sum('so_luong')).annotate(c = Count('thuoc__ten_thuoc')).annotate(bao_hiem=F('bao_hiem'))

            list_ton = []

            for i in tong_nhap_hang:
                for j in tong_xuat_hang:
                    if j['id'] in i.values():
                        so_luong_nhap = i['so_luong']
                        so_luong_xuat = j['so_luong']
                        so_luong_ton = so_luong_nhap - so_luong_xuat
                        ten_thuoc = j['thuoc__ten_thuoc']
                        d = {}
                        d['so_luong'] = so_luong_ton
                        d['ten_thuoc'] = ten_thuoc
                        list_ton.append(d)

            response = {
                'list_ton' : list_ton,
            }
            return HttpResponse(json.dumps(response), content_type='application/json; charset=utf-8')
    else:
        response = {'message': 'oke'} 
    return HttpResponse(json.dumps(response), content_type='application/json; charset=utf-8')

def xuat_bao_cao_ton_dich_vu(request):
    if request.method == "POST":
        range_start = request.POST.get('range_start')
        range_end = request.POST.get('range_end')

        start = datetime.strptime(range_start, "%d-%m-%Y")
        if range_end == '':
            end = start + timedelta(1)
            tong_nhap_hang = NhapHang.objects.filter(thoi_gian_tao__gte=start, thoi_gian_tao__lte=end).exclude(bao_hiem=True).values("thuoc__ten_thuoc").annotate(id=F('thuoc__id')).annotate(so_luong=Sum('so_luong')).annotate(c = Count('thuoc__ten_thuoc')).annotate(bao_hiem=F('bao_hiem'))

            tong_xuat_hang = KeDonThuoc.objects.filter(thoi_gian_tao__gte=start, thoi_gian_tao__lte=end).exclude(bao_hiem=True).values('thuoc__ten_thuoc').annotate(id=F('thuoc__id')).annotate(so_luong=Sum('so_luong')).annotate(c = Count('thuoc__ten_thuoc')).annotate(bao_hiem=F('bao_hiem'))

            list_ton = []
        else:
            end = datetime.strptime(range_end, "%d-%m-%Y")
            if range_start == range_end:
                end = end + timedelta(1)
            tong_nhap_hang = NhapHang.objects.filter(thoi_gian_tao__gte=start, thoi_gian_tao__lte=end).exclude(bao_hiem=True).values("thuoc__ten_thuoc").annotate(id=F('thuoc__id')).annotate(so_luong=Sum('so_luong')).annotate(c = Count('thuoc__ten_thuoc')).annotate(bao_hiem=F('bao_hiem'))

            tong_xuat_hang = KeDonThuoc.objects.filter(thoi_gian_tao__gte=start, thoi_gian_tao__lte=end).exclude(bao_hiem=True).values('thuoc__ten_thuoc').annotate(id=F('thuoc__id')).annotate(so_luong=Sum('so_luong')).annotate(c = Count('thuoc__ten_thuoc')).annotate(bao_hiem=F('bao_hiem'))

            list_ton = []

        for i in tong_nhap_hang:
            for j in tong_xuat_hang:
                if j['id'] in i.values():
                    so_luong_nhap = i['so_luong']
                    so_luong_xuat = j['so_luong']
                    so_luong_ton = so_luong_nhap - so_luong_xuat
                    ten_thuoc = j['thuoc__ten_thuoc']
                    d = {}
                    d['so_luong'] = so_luong_ton
                    d['ten_thuoc'] = ten_thuoc
                    list_ton.append(d)

        response = {
            'list_ton' : list_ton,
        }
        return HttpResponse(json.dumps(response), content_type='application/json; charset=utf-8')
    else:
        response = {'message': 'oke'}
        return HttpResponse(json.dumps(response), content_type='application/json; charset=utf-8')

@login_required(login_url='/dang_nhap/')
def hoa_don_tpcn(request, **kwargs):
    id_don_thuoc = kwargs.get('id_don_thuoc')
    don_thuoc = DonThuoc.objects.get(id = id_don_thuoc)
    danh_sach_thuoc = don_thuoc.ke_don.all()

    bao_hiem = []
    tong_tien = []
    for thuoc_instance in danh_sach_thuoc:
        if thuoc_instance.bao_hiem:
            gia = int(thuoc_instance.thuoc.don_gia_tt) * \
                thuoc_instance.so_luong
            bao_hiem.append(gia)
        else:
            gia = int(thuoc_instance.thuoc.don_gia_tt) * \
                thuoc_instance.so_luong

        tong_tien.append(gia)

    total_spent = sum(tong_tien)
    tong_bao_hiem = sum(bao_hiem)
    thanh_tien = total_spent - tong_bao_hiem

    total_spent = sum(tong_tien)
    tong_tien.clear()
    bao_hiem.clear()

    phong_chuc_nang = PhongChucNang.objects.all()
    phong_kham  = PhongKham.objects.all().first()

    data = {
        'danh_sach_thuoc': danh_sach_thuoc,
        'tong_tien'      : total_spent,
        'don_thuoc'      : don_thuoc,
        'phong_chuc_nang': phong_chuc_nang,
        'thanh_tien'     : thanh_tien,
        'tong_bao_hiem'  : tong_bao_hiem,
        'phong_kham'     : phong_kham,
    }
    return render(request, 'phong_tai_chinh/hoa_don_thuc_pham_ho_tro.html', context=data)

def my_activities(request):
    phong_chuc_nang = PhongChucNang.objects.all()
    from actstream.models import actor_stream
    activities = actor_stream(request.user)
    bac_si = request.user

    is_bac_si_lam_sang = request.user.is_bac_si_lam_sang()
    starting_day = datetime.now() - timedelta(days=7)

    context = {
        'phong_chuc_nang': phong_chuc_nang,
        'activities': activities,
        'is_bac_si_lam_sang': is_bac_si_lam_sang,
    }
    return render(request, 'user_activities.html', context)

def xoa_nhom_quyen(request):
    if request.method == "POST":
        id = request.POST.get('id')
        group = get_object_or_404(Group, id=id)
        group.delete()
        response = {
            'status': 200,
            'message': 'Nhóm quyền đã được xóa thành công'
        }
    else:
        response = {
            'status': 404, 
            'message': 'Xóa nhóm quyền không thành công'
        }
    return HttpResponse(json.dumps(response), content_type='application/json; charset=utf-8')

import re
def validation_check(input_string):
    if re.match(r"(84|0[3|5|7|8|9])+([0-9]{8})\b", str(input_string)):
        return True
    else:
        return False

@csrf_exempt
def check_so_dien_thoai_exists(request):
    if request.method == "POST":
        so_dien_thoai = request.POST.get('so_dien_thoai')
        if validation_check(so_dien_thoai):
            if User.objects.filter(so_dien_thoai=so_dien_thoai).exists():
                response = {
                    'status': 200,
                    'message': 'Số Điện Thoại Đã Tồn Tại'
                }
                return HttpResponse(json.dumps(response), content_type='application/json; charset=utf-8')
            else:
                response = {
                    'status': 404,
                    'message': 'Số Điện Thoại Khả Dụng'
                }
                return HttpResponse(json.dumps(response), content_type='application/json; charset=utf-8')
        else:
            response = {
                'status': 406,
                'message': 'Số Điện Thoại Không Hợp Lệ'
            }
            return HttpResponse(json.dumps(response), content_type='application/json; charset=utf-8')
    else:
        response = {
            'message': "not ok"
        }
        return HttpResponse(json.dumps(response), content_type='application/json; charset=utf-8')

@csrf_exempt
def check_username_exists(request):
    if request.method == "POST":
        username = request.POST.get('username')
        if User.objects.filter(username=username).exists():
            response = {
                'status': 200,
                'message': 'Username Đã Tồn Tại'
            }
            return HttpResponse(json.dumps(response), content_type='application/json; charset=utf-8')
        else:
            response = {
                'status': 404,
                'message': 'Username Khả Dụng'
            }
            return HttpResponse(json.dumps(response), content_type='application/json; charset=utf-8')
    else:
        response = {
            'message': "not ok"
        }
        return HttpResponse(json.dumps(response), content_type='application/json; charset=utf-8')

def them_mau_hoa_don(request):
    phong_chuc_nang = PhongChucNang.objects.all()
    context = {
        'phong_chuc_nang': phong_chuc_nang
    }
    return render(request, 'le_tan/them_mau_hoa_don.html', context)

def create_mau_hoa_don(request):
    if request.method == "POST":
        ten_mau_phieu = request.POST.get('ten_mau_hoa_don')
        codename = request.POST.get('codename')
        noi_dung = request.POST.get('noi_dung')

        if MauPhieu.objects.filter(codename=codename).exists():
            response = {
                'status': 404,
                'message': "Codename Cho Mẫu Hóa Đơn Này Đã Tồn Tại, Vui Lòng Chọn Codename Khác"
            }
            return HttpResponse(json.dumps(response), content_type='application/json; charset=utf-8')


        mau_hoa_don = MauPhieu(
            ten_mau = ten_mau_phieu,
            codename = codename,
            noi_dung = noi_dung,
        )
        mau_hoa_don.save()

        from actstream import action
        action.send(request.user, verb='đã thêm mới mẫu hóa đơn', target=mau_hoa_don)

        response = {
            'status': 200,
            'message': "Thêm Mới Mẫu Hóa Đơn Thành Công"
        }
    else:
        response = {
            'status': 404,
            'message': "Không Thể Gửi Lên Dữ Liệu, Vui Lòng Kiểm Tra Lại!"
        }

    return HttpResponse(json.dumps(response), content_type='application/json; charset=utf-8')

def update_mau_hoa_don(request):
    if request.method == "POST":
        id = request.POST.get('id_mau_hoa_don')
        ten_mau_hoa_don = request.POST.get('ten_mau_hoa_don')
        codename = request.POST.get('codename')
        noi_dung = request.POST.get('noi_dung')

        mau_hoa_don = get_object_or_404(MauPhieu, id=id)

        mau_hoa_don.ten_mau = ten_mau_hoa_don
        mau_hoa_don.codename = codename 
        mau_hoa_don.noi_dung = noi_dung
        mau_hoa_don.save()

        from actstream import action
        action.send(request.user, verb='đã chỉnh sửa mẫu hóa đơn', target=mau_hoa_don)

        response = {
            'status': 200,
            'message': 'Cập Nhật Mẫu Hóa Đơn Thành Công'
        }
    else:
        response = {
            'status': 404,
            'message': "Không Thể Gửi Lên Dữ Liệu, Vui Lòng Kiểm Tra Lại!"
        }
    return HttpResponse(json.dumps(response), content_type='application/json; charset=utf-8')

# def thay_doi_phan_khoa(request):
#     if request.method == "POST":
#         request_data = request.POST.get('data')
#         id_chuoi_kham = request.POST.get('id_chuoi_kham')
#         data = json.loads(request_data)
#         bao_hiem = False

#         chuoi_kham = get_object_or_404(ChuoiKham, id=id_chuoi_kham)
#         danh_sach_phan_khoa = chuoi_kham.phan_khoa_kham.all().delete()
#         benh_nhan = chuoi_kham.benh_nhan
#         print(data)
    
#         now       = datetime.now()
#         date_time = now.strftime("%m%d%y%H%M%S")
#         subName = getSubName(benh_nhan.ho_ten)
#         ma_hoa_don = "HD" + "-" + subName + '-' + date_time

#         trang_thai_phan_khoa = TrangThaiKhoaKham.objects.get_or_create(trang_thai_khoa_kham='Chờ Khám')[0]
#         bulk_create_data = []
#         for i in data:
#             if i['obj']['bao_hiem'] == "True":
#                 bao_hiem = True

#             index = data.index(i)
#             priority = index + 1
#             dich_vu = DichVuKham.objects.only('id').filter(id=i['obj']['id']).first()
#             bac_si = request.user
#             bulk_create_data.append(
#                 PhanKhoaKham(
#                     benh_nhan=benh_nhan, 
#                     dich_vu_kham=dich_vu, 
#                     bao_hiem=i['obj']['bao_hiem'], 
#                     bac_si_lam_sang=bac_si, 
#                     chuoi_kham=chuoi_kham, 
#                     priority=priority,
#                     trang_thai=trang_thai_phan_khoa,
#                     )
#                 )

#         hoa_don_chuoi_kham = chuoi_kham.hoa_don_dich_vu.delete()
#         hoa_don = HoaDonChuoiKham.objects.create(chuoi_kham=chuoi_kham, ma_hoa_don=ma_hoa_don, bao_hiem = bao_hiem)

#         PhanKhoaKham.objects.bulk_create(bulk_create_data)

#         from actstream import action
#         action.send(request.user, verb='đã cập nhật phân khoa khám cho bệnh nhân', target=benh_nhan)
#         response = {
#             'status': 200,
#             'message': "Cập Nhật Phân Khoa Thành Công"
#         }
#     else:
#         response = {
#             'status': 404,
#             'message': "Cập Nhật Phân Khoa Không Thành Công"
#         }
#     return HttpResponse(json.dumps(response), content_type='application/json; charset=utf-8')

def thay_doi_phan_khoa(request):
    if request.method == "POST":
        request_data = request.POST.get('data')
        id_chuoi_kham = request.POST.get('id_chuoi_kham')
        data = json.loads(request_data)

        chuoi_kham = get_object_or_404(ChuoiKham, id=id_chuoi_kham)
        danh_sach_phan_khoa = chuoi_kham.phan_khoa_kham.all()
        danh_sach_dich_vu_phan_khoa = [phan_khoa.dich_vu_kham for phan_khoa in danh_sach_phan_khoa]
        benh_nhan = chuoi_kham.benh_nhan

        trang_thai_phan_khoa = TrangThaiKhoaKham.objects.get_or_create(trang_thai_khoa_kham='Chờ Khám')[0]
        bulk_create_data = []

        counter = 0
        old_priority = danh_sach_phan_khoa.aggregate(Max('priority'))

        for i in data:
            dich_vu = DichVuKham.objects.only('id').filter(id=i['obj']['id']).first()

            bac_si = request.user

            if dich_vu not in danh_sach_dich_vu_phan_khoa:
                counter += 1
                priority = int(old_priority['priority__max']) + counter
                model = PhanKhoaKham(
                    benh_nhan=benh_nhan,
                    dich_vu_kham=dich_vu,
                    bac_si_lam_sang=bac_si,
                    bao_hiem=i['obj']['bao_hiem'],
                    chuoi_kham=chuoi_kham,
                    trang_thai=trang_thai_phan_khoa,
                    priority = priority,
                )
                bulk_create_data.append(model)

        PhanKhoaKham.objects.bulk_create(bulk_create_data)

        from actstream import action
        action.send(request.user, verb='đã cập nhật phân khoa khám cho bệnh nhân', target=benh_nhan)
        response = {
            'status': 200,
            'message': "Cập Nhật Phân Khoa Thành Công"
        }
    else:
        response = {
            'status': 404,
            'message': "Cập Nhật Phân Khoa Không Thành Công"
        }
    return HttpResponse(json.dumps(response), content_type='application/json; charset=utf-8')

def thay_doi_phan_khoa_hoa_don_dich_vu(request):
    if request.method == "POST":
        request_data = request.POST.get('data')
        id_chuoi_kham = request.POST.get('id_chuoi_kham')
        data = json.loads(request_data)

        chuoi_kham = get_object_or_404(ChuoiKham, id=id_chuoi_kham)
        chuoi_kham.phan_khoa_kham.all().delete()
        
        benh_nhan = chuoi_kham.benh_nhan

        trang_thai_phan_khoa = TrangThaiKhoaKham.objects.get_or_create(trang_thai_khoa_kham='Chờ Khám')[0]
        bulk_create_data = []

        
        for i, j  in enumerate(data):
            priority = i + 1
            dich_vu = DichVuKham.objects.only('id').filter(id=j['obj']['id']).first()

            bac_si = request.user

            model = PhanKhoaKham(
                benh_nhan=benh_nhan,
                dich_vu_kham=dich_vu,
                bac_si_lam_sang=bac_si,
                bao_hiem=j['obj']['bao_hiem'],
                chuoi_kham=chuoi_kham,
                trang_thai=trang_thai_phan_khoa,
                priority = priority,
            )
            bulk_create_data.append(model)

        PhanKhoaKham.objects.bulk_create(bulk_create_data)

        from actstream import action
        action.send(request.user, verb='đã cập nhật phân khoa khám cho bệnh nhân', target=benh_nhan)
        response = {
            'status': 200,
            'message': "Cập Nhật Phân Khoa Thành Công"
        }
    else:
        response = {
            'status': 404,
            'message': "Cập Nhật Phân Khoa Không Thành Công"
        }
    return HttpResponse(json.dumps(response), content_type='application/json; charset=utf-8')

@login_required(login_url='/dang_nhap/')
def hoa_don_lam_sang(request, **kwargs):
    mau_hoa_don = MauPhieu.objects.filter(codename='hoa_don_lam_sang').first()
    phong_chuc_nang = PhongChucNang.objects.all()
    id_lich_hen = kwargs.get('id')
    lich_hen = get_object_or_404(LichHenKham, id=id_lich_hen)
    benh_nhan = lich_hen.benh_nhan
    hoa_don_lam_sang = lich_hen.hoa_don_lam_sang.all().last()
    thoi_gian_thanh_toan = hoa_don_lam_sang.thoi_gian_tao
    tong_tien = hoa_don_lam_sang.tong_tien
    nguoi_thanh_toan = hoa_don_lam_sang.nguoi_thanh_toan

    data_dict = {}
    data_dict['{benh_nhan}'] = benh_nhan.ho_ten.upper()
    data_dict['{so_dien_thoai}'] = benh_nhan.get_so_dien_thoai()
    data_dict['{dia_chi}'] = benh_nhan.get_dia_chi()
    data_dict['{tien_thanh_toan}'] = "{:,}".format(int(tong_tien))
    data_dict['{gia_tien}'] = "{:,}".format(int(tong_tien))
    data_dict['{nguoi_thanh_toan}'] = nguoi_thanh_toan.ho_ten
    data_dict['{thoi_gian_thanh_toan}'] = f"{thoi_gian_thanh_toan.strftime('%H:%m')} Ngày {thoi_gian_thanh_toan.strftime('%d')} Tháng {thoi_gian_thanh_toan.strftime('%m')} Năm {thoi_gian_thanh_toan.strftime('%Y')}"

    context = {
        'mau_hoa_don': mau_hoa_don,
        'data_dict': data_dict,
        'phong_chuc_nang': phong_chuc_nang,
    }

    return render(request, 'phong_tai_chinh/hoa_don_lam_sang.html', context)

def bao_cao_xuat_nhap_ton_thuoc_view(request):
    phong_chuc_nang = PhongChucNang.objects.all()
    context = {
        'phong_chuc_nang': phong_chuc_nang
    }

    return render(request, 'phong_tai_chinh/bao_cao_xuat_nhap_ton_tong_hop_thuoc.html', context)

def danh_sach_thuoc_sap_het_date_view(request):
    phong_chuc_nang = PhongChucNang.objects.all()
    context = {
        'phong_chuc_nang': phong_chuc_nang
    }
    return render(request, 'phong_tai_chinh/bao_cao_thuoc_sap_het_date.html', context)

@csrf_exempt
def export_excel_danh_sach_xnt_tong_hop_thuoc(request):
    if request.method == 'POST':
        range_start = request.POST.get('startDate', None)
        range_end   = request.POST.get('endDate', None)
        start = datetime.strptime(range_start, "%d-%m-%Y")

        _start = "".join(range_start.split('-'))
        _end = "".join(range_end.split('-'))

        if range_end == '':
            end = start + timedelta(1)
        else:
            end = datetime.strptime(range_end, "%d-%m-%Y")
            if range_start == range_end:
                end = end + timedelta(1)

        danh_sach_thuoc = Thuoc.objects.all()

        medicines = danh_sach_thuoc.annotate(
            so_luong_nhap = Subquery(
                NhapHang.objects.filter(
                    thuoc = OuterRef('pk'),
                ).filter(thoi_gian_tao__gte=start, thoi_gian_tao__lt=end).values_list(
                    Func(
                        'so_luong',
                        function='SUM',
                    )
                )
            )
        ).annotate(
            so_luong_xuat = Subquery(
                KeDonThuoc.objects.filter(
                    thuoc = OuterRef('pk'),
                ).filter(thoi_gian_tao__gte=start, thoi_gian_tao__lt=end).values_list(
                    Func(
                        'so_luong',
                        function='SUM',
                    )
                )
            )
        ).annotate(
            ton_dau_ky=models.Case(
                models.When(Q(so_luong_nhap__isnull=True) & Q(so_luong_xuat__isnull=True), then=F('so_luong_kha_dung')),
                models.When(so_luong_nhap__isnull=True, then=F('so_luong_kha_dung') + F('so_luong_xuat')),
                models.When(so_luong_xuat__isnull=True, then=F('so_luong_kha_dung') - F('so_luong_nhap')),
                default = F('so_luong_kha_dung') - F('so_luong_nhap') + F('so_luong_xuat'),
                output_field=models.IntegerField(),
            )
        ).annotate(
            ton_cuoi_ky=models.Case(
                models.When(Q(so_luong_nhap__isnull=True) & Q(so_luong_xuat__isnull=True), then=F('ton_dau_ky')),
                models.When(so_luong_nhap__isnull=True, then=F('ton_dau_ky') - F('so_luong_xuat')),
                models.When(so_luong_xuat__isnull=True, then=F('ton_dau_ky') + F('so_luong_nhap')),
                default = F('ton_dau_ky') + F('so_luong_nhap') - F('so_luong_xuat'),
            )
        ).annotate(
            thanh_tien_nhap=F('so_luong_nhap') * F('don_gia')
        ).annotate(
            thanh_tien_xuat=F('so_luong_xuat') * F('don_gia_tt')
        ).values(
            'ma_thuoc',
            'ten_thuoc',
            'ten_hoat_chat',
            'don_vi_tinh',
            'ham_luong',
            'cong_ty__ten_cong_ty',
            'nuoc_sx',
            'so_lo',
            'han_su_dung',
            'don_gia',
            'don_gia_tt',
            'gia_bhyt',      
            'so_luong_nhap',
            'so_luong_xuat',
            'ton_dau_ky',
            'thanh_tien_nhap',
            'thanh_tien_xuat',
            'ton_cuoi_ky',
        )
        
        xlsx_data = ExportExcelBaoCaoXNTThuoc(list(medicines))
        response = HttpResponse(
            content_type='application/ms-excel'
        )
        response['Content-Disposition'] = f'attachment; filename="BCXNT_THUOC_TONGHOP({_start}-{_end}).xlsx"'
        response.write(xlsx_data)
       
    else:
        response = HttpResponse(json.dumps({'message': "It's not gonna happen"}), content_type='application/json; charset=utf-8')
    return response

@csrf_exempt
def export_excel_danh_sach_thuoc_het_date(request):
    if request.method == "POST":
        num_of_months = request.POST.get('num_of_months')
        num_of_months_later = date.today() + relativedelta(months=+int(num_of_months))

        medicines = Thuoc.objects.filter(han_su_dung__lte=num_of_months_later)
        serializer = ThuocSerializer(medicines, many=True)

        xlsx_data = ExportExcelDSThuocSapHetDate(serializer.data)
        response = HttpResponse(
            content_type='application/ms-excel'
        )
        response['Content-Disposition'] = f'attachment; filename="DS_THUOC_SAP_HET_DATE.xlsx"'
        response.write(xlsx_data)
       
    else:
        response = HttpResponse(json.dumps({'message': "It's not gonna happen"}), content_type='application/json; charset=utf-8')
    return response

@csrf_exempt
def hoan_thanh_kham(request):
    if request.method == "POST":
        id_chuoi_kham = request.POST.get('id_chuoi_kham')
        try:
            chuoi_kham = ChuoiKham.objects.get(id=id_chuoi_kham)
            lich_hen = chuoi_kham.lich_hen

            trang_thai_chuoi_kham = TrangThaiChuoiKham.objects.get_or_create(trang_thai_chuoi_kham = "Hoàn Thành")[0]
            trang_thai_lich_hen = TrangThaiLichHen.objects.get_or_create(ten_trang_thai = "Hoàn Thành")[0]

            chuoi_kham.trang_thai = trang_thai_chuoi_kham
            lich_hen.trang_thai = trang_thai_lich_hen
            lich_hen.thanh_toan_sau = False
            lich_hen.thoi_gian_ket_thuc = timezone.localtime(timezone.now())
            chuoi_kham.save()
            lich_hen.save()

        except ChuoiKham.DoesNotExist:
            response = {
                'status': 404,
                'message': "Chuỗi Khám Không Tồn Tại",
            }
            return HttpResponse(json.dumps(response), content_type='application/json; charset=utf-8')
        
        response = {
            'status': 200,
            'message': "Hoàn Thành Chuỗi Khám"
        }
    else:
        response = {
            'status': 404,
            'message': 'Xảy ra lỗi trong quá trình xử lí'
        }
    
    return HttpResponse(json.dumps(response), content_type='application/json; charset=utf-8')


@csrf_exempt
def hoan_thanh_kham_hoa_don(request):
    if request.method == "POST":
        id_chuoi_kham = request.POST.get('id_chuoi_kham')
        try:
            chuoi_kham = ChuoiKham.objects.get(id=id_chuoi_kham)
            lich_hen = chuoi_kham.lich_hen
            hoa_don_dich_vu = chuoi_kham.hoa_don_dich_vu

            trang_thai_chuoi_kham = TrangThaiChuoiKham.objects.get_or_create(trang_thai_chuoi_kham = "Hoàn Thành")[0]
            trang_thai_lich_hen = TrangThaiLichHen.objects.get_or_create(ten_trang_thai = "Hoàn Thành")[0]

            chuoi_kham.trang_thai = trang_thai_chuoi_kham
            lich_hen.trang_thai = trang_thai_lich_hen
            lich_hen.thoi_gian_ket_thuc = timezone.localtime(timezone.now())
            lich_hen.thanh_toan_sau = False
            if hoa_don_dich_vu is not None:
                hoa_don_dich_vu.tong_tien = 0
                hoa_don_dich_vu.discount = 0
                hoa_don_dich_vu.save()

            chuoi_kham.save()
            lich_hen.save()

        except ChuoiKham.DoesNotExist:
            response = {
                'status': 404,
                'message': "Chuỗi Khám Không Tồn Tại",
            }
            return HttpResponse(json.dumps(response), content_type='application/json; charset=utf-8')
        
        response = {
            'status': 200,
            'message': "Hoàn Thành Chuỗi Khám"
        }
    else:
        response = {
            'status': 404,
            'message': 'Xảy ra lỗi trong quá trình xử lí'
        }
    
    return HttpResponse(json.dumps(response), content_type='application/json; charset=utf-8')

@csrf_exempt
def hoan_thanh_chuoi_kham_hoa_don_thuoc(request):
    if request.method == 'POST':
        id = request.POST.get('id')
        trang_thai_don_thuoc = TrangThaiDonThuoc.objects.get_or_create(trang_thai = "Hoàn Thành")[0]
        don_thuoc = get_object_or_404(DonThuoc, id=id)
        if don_thuoc.chuoi_kham is not None:
            now             = datetime.now()
            date_time       = now.strftime("%m%d%y%H%M%S")
            ma_hoa_don      = "HDT-" + date_time

            chuoi_kham = don_thuoc.chuoi_kham
            lich_hen = chuoi_kham.lich_hen
            trang_thai_chuoi_kham = TrangThaiChuoiKham.objects.get_or_create(trang_thai_chuoi_kham = "Hoàn Thành")[0]
            trang_thai_lich_hen = TrangThaiLichHen.objects.get_or_create(ten_trang_thai = "Hoàn Thành")[0]
            chuoi_kham.trang_thai = trang_thai_chuoi_kham
            lich_hen.trang_thai = trang_thai_lich_hen
            lich_hen.thoi_gian_ket_thuc = timezone.localtime(timezone.now())
            lich_hen.save()
            chuoi_kham.save()
            don_thuoc.trang_thai = trang_thai_don_thuoc
            don_thuoc.save()

            hoa_don_thuoc, created = HoaDonThuoc.objects.get_or_create(don_thuoc=don_thuoc, ma_hoa_don=ma_hoa_don, tong_tien=0, nguoi_thanh_toan=request.user)
            hoa_don_thuoc.save()
            
            channel_layer = get_channel_layer()
            async_to_sync(channel_layer.group_send)(
                f"process_accomplished_user_{don_thuoc.benh_nhan.id}", {
                    'type':'process_accomplished_notification'
                }
            )
        else:
            don_thuoc.trang_thai = trang_thai_don_thuoc
            don_thuoc.save()
            response={
                'status' : 200,
                'message' : 'Đã Nhận Đơn Thuốc'
            }
            return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")

        response={
            'status' : 200,
            'message' : 'Đã Nhận Đơn Thuốc'
        }
        return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")

def xoa_ket_qua_chuyen_khoa(request):
    if request.method == "POST":
        id = request.POST.get('id')
        try:
            ket_qua_chuyen_khoa = KetQuaChuyenKhoa.objects.get(id=id)
            ket_qua_chuyen_khoa.delete()
        except KetQuaChuyenKhoa.DoesNotExist:
            response = {
                'status': 404,
                'message': "Kết Quả Chuyên Khoa Không Tồn Tại"
            }
            return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")
        response = {
            'status': 200,
            'message': "Kết Quả Chuyên Khoa Đã Được Xóa Thành Công"
        }
    else:
        response = {
            'status': 404,
            'message': "Xảy Ra Lỗi Trong Quá Trình Xử Lý"
        }
    return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")

def menu_view(request):
    phong_chuc_nang = PhongChucNang.objects.all()
    context = {
        'phong_chuc_nang': phong_chuc_nang
    }

    return render(request, 'menu.html', context)

def xoa_phan_khoa_kham(request):
    if request.method == "POST":
        id = request.POST.get('id')
        try:
            phan_khoa = PhanKhoaKham.objects.get(id=id)
            phan_khoa.delete()
            response = {
                'status': 200,
                'message': 'Đã Thu Hồi Chỉ Định'
            }
        except PhanKhoaKham.DoesNotExist:
            response = {
                'status': 404,
                'message': 'Không Tìm Thấy Chỉ Định'
            }
        return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")

    else:
        response = {
            'status': 404,
            'message': "Xảy Ra Lỗi Trong Quá Trình Thực Hiện"
        }
        return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")

def chinh_sua_phieu_ket_qua_view(request, **kwargs):
    phong_chuc_nang = PhongChucNang.objects.all()
    id_phan_khoa = kwargs.get('id_phan_khoa')
    try:
        phan_khoa = PhanKhoaKham.objects.get(id=id_phan_khoa)
        ket_qua_chuyen_khoa = phan_khoa.ket_qua_chuyen_khoa.all().last()
        html_ket_qua = ket_qua_chuyen_khoa.html_ket_qua.all().last()

    except PhanKhoaKham.DoesNotExist:
        pass
    context = {
        'phong_chuc_nang': phong_chuc_nang,
        'mau_phieu': html_ket_qua.noi_dung,
        'ket_qua_chuyen_khoa': ket_qua_chuyen_khoa,
        'phan_khoa_kham': phan_khoa,
        'html_ket_qua': html_ket_qua
    }
    return render(request, 'bac_si_chuyen_khoa/update_phieu_ket_qua.html', context)

def store_update_phieu_ket_qua(request):
    if request.method == "POST":
        id_html_ket_qua = request.POST.get('id_html_ket_qua')
        noi_dung = request.POST.get('noi_dung')
        try:
            html_ket_qua = HtmlKetQua.objects.get(id=id_html_ket_qua)
            html_ket_qua.noi_dung = noi_dung
            html_ket_qua.save()
            response = {
                'status': 200,
                'message': "Cập Nhật Phiếu Kết Quả Thành Cống"
            }
        except HtmlKetQua.DoesNotExist:
            response = {
                'status': 404, 
                'message': "Không Tìm Thấy Phiếu Kết Quả"
            }
        return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")
    else:
        response = {
            'status': 404,
            'message': "Xảy Ra Lỗi Trong Quá Trình Xử Lí"
        }
        return HttpResponse(json.dumps(response), content_type="application/json, charset=utf-8")

